"""
DocxTpl Automatisierung v7.0 - Kombinierte Logik und GUI mit Tabs

Dieses Skript kombiniert eine PySide6 GUI mit der Logik zur Dokumentenerstellung
in einer einzigen Datei. Es automatisiert die Erstellung von Word-Dokumenten
aus Vorlagen basierend auf Daten aus einer Excel-Datei.

Wichtigste Merkmale:
- Grafische Benutzeroberfläche mit Tabs zur einfachen Bedienung.
- Zusammenführung von GUI und Berichtsgenerator in einer Datei.
- Robuste Ersetzung von Platzhaltern in Word-Dokumenten mit docxtpl.
- Beibehaltung der Formatierung (fett, kursiv, etc.) von Platzhaltern.
- Asynchrone Verarbeitung, um ein Einfrieren der GUI zu verhindern.
- Live-Logs in eigenem Tab und anpassbare Themes.
- SVG-Konvertierung und QR-Code-Generierung.
- Konfigurierbare Ausgabekategorien direkt in der GUI.
- Fallback-Logik für Daten aus dem zweiten Excel-Blatt.
"""

# ==============================================================================
# VERSION
# ==============================================================================
VERSION = "7.0.1"
BUILD_DATE = "2025-01-27"

import sys
import os
import json
import pandas as pd
import traceback
import io
import shutil
import re
from docx.shared import Cm
import qrcode
from datetime import datetime, timezone
import tempfile
from openpyxl import load_workbook

from docxtpl import DocxTemplate, InlineImage
from jinja2.exceptions import TemplateSyntaxError
from docx.opc.exceptions import PackageNotFoundError

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QFileDialog, QProgressBar, QTableView, QMessageBox, QFrame,
    QSpinBox, QTextEdit, QMainWindow, QTabWidget, QGroupBox, QScrollArea, QCheckBox,
    QSplitter, QMenu, QComboBox
)
from PySide6.QtCore import Qt, QAbstractTableModel, QThread, Signal, QObject
from PySide6.QtGui import QIcon, QAction, QActionGroup, QTextCursor

# ==============================================================================
# BERICHT-GENERATOR LOGIK
# ==============================================================================
# Dieser Abschnitt enthält die gesamte Logik für die Dokumentenerstellung.
# Er ist unabhängig von der GUI und könnte auch als eigenständiges Skript
# importiert und verwendet werden. Die Funktionen hier kümmern sich um das
# Laden der Daten, die Verarbeitung der Vorlagen und das Speichern der
# fertigen Dokumente.

def _log_handler(msg, level="INFO", log_callback=None, is_dark_mode=False):
    """
    Zentraler Log-Handler, der Nachrichten formatiert und ausgibt.
    Unterstützt zwei Ausgabemodi:
    - GUI (log_callback vorhanden): Formatiert die Nachricht als HTML für eine
      farbige und strukturierte Darstellung in einem QTextEdit-Widget.
    - Konsole (log_callback ist None): Gibt die Nachricht als einfachen Text aus.
    """
    timestamp = datetime.now().strftime('%H:%M:%S')
    level_upper = level.upper()

    # Nachricht immer zuerst für die Konsole formatieren
    plain_message = msg if level_upper == "SEP" else f"[{timestamp}] [{level_upper:^7}] {msg}"
    if not plain_message.endswith('\n'):
        plain_message += '\n'

    if log_callback:
        # GUI-Logging: HTML aus den ursprünglichen Daten erstellen
        color_map = {
            "INFO": "#0077CC",      # Blau
            "SUCCESS": "#2E8B57",   # Seegrün
            "WARN": "#FFA500",      # Orange
            "ERROR": "#D22B2B",      # Ziegelrot
            "FATAL": "#8B0000",      # Dunkelrot
            "SEP": "#808080"        # Grau
        }
        color = color_map.get(level_upper, "black")

        # KEIN Escaping mehr! msg wird als Plaintext behandelt, aber HTML-Formatierung wird hinzugefügt
        msg_html = msg.replace("\n", "<br>")
        # Dateinamen fett machen (z.B. 'template.docx')
        msg_html = re.sub(r"'([^']+\.\w+)'", r"<strong>'\1'</strong>", msg_html)
        # Platzhalter-Sets lila machen (z.B. {'var1', 'var2'})
        msg_html = re.sub(r"(\s*\{.*?\})", r'<span style="color: #8A2BE2;">\1</span>', msg_html)

        if level_upper == "SEP":
            html_message = f'<div style="display:block; color: {color}; text-align: center; font-family: monospace; margin: 5px 0;">{msg_html}</div>'
        else:
            beschr_color = "#FFF" if is_dark_mode else "#333"
            html_message = (
                f'<br><div style="display:block; font-family: Consolas, Courier New, monospace; line-height: 1.4;">'
                f'<span style="color: #808080;">{timestamp}</span> '
                f'<span style="color: {color}; font-weight: bold;">[{level_upper:^7}]</span> '
                f'<span style="color: {beschr_color};">{msg_html}</span>'
                f'</div>'
            )
        log_callback(html_message)
    else:
        # Reines Konsolen-Logging
        print(plain_message, end='')

def svg_to_png_file_pyside(svg_path, png_path, log_callback=None, scale=3, compression=-1):
    """
    Konvertiert eine SVG-Datei in eine PNG-Datei unter Verwendung von PySide6.
    Dies ist nützlich, da python-docx SVG-Bilder nicht direkt einbetten kann.
    
    Args:
        svg_path (str): Pfad zur SVG-Quelldatei.
        png_path (str): Pfad zur PNG-Zieldatei.
        log_callback (callable, optional): Callback für Log-Nachrichten.
        scale (int): Skalierungsfaktor für die Auflösung der PNG-Datei.
        compression (int): PNG-Kompressionslevel (-1 für Standard).
        
    Returns:
        bool: True bei Erfolg, False bei einem Fehler.
    """
    try:
        from PySide6.QtSvg import QSvgRenderer
        from PySide6.QtGui import QImage, QPainter
        from PySide6.QtCore import QByteArray, QSize

        if not os.path.exists(svg_path):
            _log_handler(f"SVG-Datei nicht gefunden: {svg_path}", "ERROR", log_callback); return False
        with open(svg_path, 'rb') as f: svg_data = f.read()
        renderer = QSvgRenderer(QByteArray(svg_data))
        if not renderer.isValid():
            _log_handler(f"SVG-Datei ist ungültig: {svg_path}", "ERROR", log_callback); return False
        size = renderer.defaultSize()
        if size.isEmpty(): size.setWidth(300); size.setHeight(150)
        scaled_size = QSize(int(size.width() * scale), int(size.height() * scale))
        image = QImage(scaled_size, QImage.Format_ARGB32); image.fill(Qt.transparent)
        painter = QPainter(image); renderer.render(painter); painter.end()
        if not image.save(png_path, "PNG", compression):
            _log_handler(f"Speichern der PNG-Datei fehlgeschlagen für: {png_path}", "ERROR", log_callback); return False
        return True
    except Exception as e:
        _log_handler(f"FEHLER bei SVG-Konvertierung: {e}", "ERROR", log_callback); return False

def lade_excel_daten(pfad, header_row=3, log_callback=None):
    """
    Lädt Daten aus einer Excel-Datei.
    - Liest das erste Tabellenblatt in einen Pandas DataFrame. Jede Zeile, die eine
      'anlage_seriennummer' enthält, wird als ein Datensatz (dictionary) behandelt.
    - Liest das zweite Tabellenblatt als Quelle für Fallback-Textmarken. Dies ist
      nützlich für allgemeine Informationen (z.B. Projektname, Datum), die für
      alle Dokumente gleich sind.
      
    Args:
        pfad (str): Pfad zur Excel-Datei.
        header_row (int): Die Zeilennummer, die die Spaltenüberschriften enthält.
        log_callback (callable, optional): Callback für Log-Nachrichten.
        
    Returns:
        tuple: Ein Tupel mit (Liste von Datensätzen, Dictionary mit Fallback-Marken).
    """
    _log_handler(f"Lade Excel-Daten aus: {pfad}", "INFO", log_callback)
    if not os.path.isfile(pfad): raise FileNotFoundError(f"FEHLER: Excel-Datei nicht gefunden: {pfad}")
    try:
        df = pd.read_excel(pfad, header=header_row - 1).dropna(how='all')
        df.columns = [str(col).strip() for col in df.columns]
        df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
        datensaetze = [row.to_dict() for _, row in df.iterrows() if 'anlage_seriennummer' in row and pd.notna(row['anlage_seriennummer'])]
        _log_handler(f"Erfolgreich {len(datensaetze)} Datensätze (Zeilen) geladen.", "SUCCESS", log_callback)
        
        fallback_marken = {}
        wb = load_workbook(pfad, read_only=True)
        if len(wb.sheetnames) > 1:
            ws = wb[wb.sheetnames[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 3 and row[1]:
                    key, value = str(row[1]).strip(), str(row[2]).strip() if row[2] is not None else ""
                    fallback_marken[key] = value
            _log_handler(f"Erfolgreich {len(fallback_marken)} Fallback-Textmarken aus dem zweiten Blatt geladen.", "SUCCESS", log_callback)
        
        return datensaetze, fallback_marken
    except Exception as e:
        raise Exception(f"FEHLER beim Lesen der Excel-Datei {pfad}: {e}") from e

def ersetze_platzhalter_mit_docxtpl(doc_path, context, svg_png_map, log_callback=None):
    """
    Ersetzt Platzhalter in einem Word-Dokument (.docx) mit den Daten aus dem Kontext.
    Nutzt docxtpl, was eine Jinja2-ähnliche Syntax erlaubt.
    
    Verarbeitungsreihenfolge (nach User-Wunsch für bessere Nachvollziehbarkeit):
    1. Bild-Platzhalter (_img)
    2. QR-Code-Platzhalter (_qr, _link)
    3. Alle anderen Text-Platzhalter
    
    Spezielle Logik:
    - Bilder: Platzhalter mit Suffix '_img' werden durch Bilder ersetzt. Die Bildgröße
      kann über einen zusätzlichen Platzhalter mit Suffix '_img_size' gesteuert werden.
    - QR-Codes: Platzhalter mit Suffix '_qr' oder '_link' werden in QR-Codes umgewandelt.
      Auch hier kann die Größe mit einem '_qr_size' oder '_link_size' Suffix angepasst werden.
    - SVG-Unterstützung: Verwendet die vorab konvertierten PNGs aus der svg_png_map.
    
    Args:
        doc_path (str): Pfad zur Word-Vorlagendatei.
        context (dict): Dictionary mit den Daten zum Ersetzen.
        svg_png_map (dict): Mapping von SVG-Pfaden zu temporären PNG-Pfaden.
        log_callback (callable, optional): Callback für Log-Nachrichten.
        
    Returns:
        DocxTemplate: Das bearbeitete Dokumentenobjekt.
    """
    doc = DocxTemplate(doc_path)
    try:
        undeclared_variables = doc.get_undeclared_template_variables()
    except Exception:
        undeclared_variables = set(context.keys())

    render_context = context.copy()
    bilder_ordner = context.get('_bilder_ordner')

    # Platzhalter nach Typ sortieren, um eine definierte Verarbeitungsreihenfolge zu gewährleisten.
    image_keys = sorted([k for k in undeclared_variables if k.endswith('_img')])
    qr_keys = sorted([k for k in undeclared_variables if k.endswith(('_qr', '_link'))])
    text_keys = sorted([k for k in undeclared_variables if k not in image_keys and k not in qr_keys])

    # Gewünschte Verarbeitungsreihenfolge: Bilder -> QR-Codes -> Text.
    processing_order = image_keys + qr_keys + text_keys

    for key in processing_order:
        value = context.get(key)
        if not isinstance(value, str) or not value.strip():
            continue

        # 1. Bild-Platzhalter verarbeiten
        if key in image_keys and bilder_ordner:
            width_cm = 15
            size_val = context.get(key + '_size', context.get(key + '_Size'))
            if pd.notna(size_val):
                try:
                    width_cm = float(size_val)
                except (ValueError, TypeError):
                    pass
            bild_pfad = os.path.join(bilder_ordner, value)
            if os.path.exists(bild_pfad):
                try:
                    render_context[key] = InlineImage(doc, svg_png_map.get(bild_pfad, bild_pfad), width=Cm(width_cm))
                except Exception as e:
                    render_context[key] = f"[Bild-Fehler]"
                    _log_handler(f"Bild {value} Fehler: {e}", "ERROR", log_callback)
            else:
                render_context[key] = f"[Bild nicht gefunden]"
                _log_handler(f"Bild {value} nicht gefunden", "WARN", log_callback)

        # 2. QR-Code-Platzhalter verarbeiten
        elif key in qr_keys:
            width_cm = 4
            size_val = context.get(key + '_size', context.get(key + '_Size'))
            if pd.notna(size_val):
                try:
                    width_cm = float(size_val)
                except (ValueError, TypeError):
                    pass
            try:
                qr_img = qrcode.make(value)
                img_bytes = io.BytesIO()
                qr_img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                render_context[key] = InlineImage(doc, img_bytes, width=Cm(width_cm))
            except Exception as e:
                render_context[key] = f"[QR-Fehler]"
                _log_handler(f"QR-Code für '{value}' Fehler: {e}", "ERROR", log_callback)

    # 3. Alle Platzhalter (inkl. der vorbereiteten Bilder/QRs) von docxtpl rendern lassen.
    doc.render(render_context, autoescape=True)
    return doc

def _speichere_dokument(doc, vorlage_pfad, eintrag, is_anlage, haupt_export_ordner, categories, log_callback, is_dark_mode=False):
    """
    Speichert ein generiertes Word-Dokument im richtigen Ordner und mit dem korrekten Namen.
    
    - Kategorisierung: Der Dateiname der Vorlage wird geprüft. Beginnt er mit einem
      der in der GUI definierten Präfixe (z.B. 'B_'), wird das Dokument in den
      zugehörigen Kategorie-Ordner (z.B. 'Beschilderung') gespeichert.
    - Benennung:
        - Anlagenspezifische Dokumente: '{seriennummer}_{vorlagenname}.docx'
        - Allgemeine Dokumente: '{vorlagenname}.docx'
    """
    vorlage_name = os.path.basename(vorlage_pfad)
    cleaned_name = vorlage_name
    export_kategorie_ordner = haupt_export_ordner
    for prefix, cat_name in categories.items():
        if vorlage_name.lower().startswith(prefix.lower()):
            cleaned_name = vorlage_name[len(prefix):]; export_kategorie_ordner = os.path.join(haupt_export_ordner, cat_name); break
    os.makedirs(export_kategorie_ordner, exist_ok=True)
    seriennummer = str(eintrag.get('anlage_seriennummer', 'Unbekannt')).strip()
    ziel_name = f"{seriennummer}_{cleaned_name}" if is_anlage else cleaned_name
    
    # Speichere das Dokument
    ziel_pfad = os.path.join(export_kategorie_ordner, ziel_name)
    doc.save(ziel_pfad)
    
    if is_anlage:
        _log_handler(f"-> Speichere anlagenspez. Dokument: '{ziel_name}' für Anlage: '{seriennummer}' in Ordner: '{os.path.basename(export_kategorie_ordner)}'", "SUCCESS", log_callback, is_dark_mode)
    else:
        _log_handler(f"-> Speichere allgemeines Dokument: '{ziel_name}' in Ordner: '{os.path.basename(export_kategorie_ordner)}'", "SUCCESS", log_callback, is_dark_mode)

def verarbeite_vorlagen_trockenlauf(vorlagen_ordner, excel_path, log_callback, header_row, bilder_ordner=None):
    """
    Führt eine "Trockenlauf"-Validierung durch, ohne Dokumente zu erstellen.
    Prüft, ob alle referenzierten Bilder und Platzhalter vorhanden sind.
    """
    def _log(msg, level="INFO"): _log_handler(msg, level, log_callback)
    _log("="*60+"\n", "SEP"); _log("Starte Trockenlauf (Konfigurationsprüfung)...", "INFO"); _log("="*60+"\n", "SEP")
    
    is_ok = True
    
    try:
        # 1. Excel-Daten laden
        datensaetze, fallback_marken = lade_excel_daten(excel_path, header_row, log_callback)
        if not datensaetze:
            _log("Keine Datensätze in Excel gefunden. Abbruch.", "WARN")
            return False
        
        excel_columns = set(datensaetze[0].keys()) | set(fallback_marken.keys())
        _log(f"Gefundene Spalten/Marken in Excel: {len(excel_columns)}", "INFO")
        _log(f"Gefundene Fallback-Textmarken (Blatt 2): {sorted(list(fallback_marken.keys()))}", "INFO")

        # 2. Vorlagen sammeln
        all_templates = []
        for root, _, files in os.walk(vorlagen_ordner):
            for file in files:
                if file.endswith('.docx') and not file.startswith('~'):
                    all_templates.append(os.path.join(root, file))
        _log(f"Gefundene Word-Vorlagen: {len(all_templates)}", "INFO")

        # 3. Vorlagen und Bilder validieren
        _log("\n--- Validierung der Vorlagen ---\n", "SEP")
        all_template_vars = set()
        # 'datetime_utc' wird automatisch hinzugefügt und ist daher kein Fehler, wenn es in Excel fehlt.
        valid_placeholders = excel_columns | {'datetime_utc'}

        for template_path in all_templates:
            try:
                doc = DocxTemplate(template_path)
                template_vars = doc.get_undeclared_template_variables()
                all_template_vars.update(template_vars)
                
                missing_vars = template_vars - valid_placeholders
                if missing_vars:
                    is_ok = False
                    _log(f"Vorlage '{os.path.basename(template_path)}': Fehlende Platzhalter in Excel: {missing_vars}", "ERROR")
                else:
                    _log(f"Vorlage '{os.path.basename(template_path)}': OK", "SUCCESS")

                # NEU: Für _img, _qr, _link Platzhalter prüfe nur in Excel/Fallback auf _size
                img_qr_keys = {k for k in template_vars if k.endswith(('_img', '_qr', '_link'))}
                if img_qr_keys:
                    _log(f"Platzhalter-Details für '{os.path.basename(template_path)}':", "INFO")
                for key in img_qr_keys:
                    size_key_lower = f"{key}_size"
                    size_key_camel = f"{key}_Size"
                    # Wert und Größe suchen (Excel/Fallback)
                    value = None
                    size = None
                    # Suche Wert in allen Datensätzen und Fallback
                    for eintrag in datensaetze + [fallback_marken]:
                        if key in eintrag and eintrag[key]:
                            value = eintrag[key]
                        if size_key_lower in eintrag and eintrag[size_key_lower]:
                            size = eintrag[size_key_lower]
                        elif size_key_camel in eintrag and eintrag[size_key_camel]:
                            size = eintrag[size_key_camel]
                    # Standardgrößen
                    if key.endswith('_img'):
                        default_size = 15
                        typ = 'Bild'
                    else:
                        default_size = 4
                        typ = 'QR/Link'
                    used_size = size if size is not None else default_size
                    value_str = value if value is not None else '[kein Wert in Excel/Fallback]'
                    _log(f"  - {typ}: '{key}' → Wert: {value_str}, Größe: {used_size} cm", "INFO")
                    if size is None:
                        _log(f"Hinweis: Für Platzhalter '{key}' wird die Standardgröße verwendet, da kein '{size_key_lower}' oder '{size_key_camel}' in Excel/Fallback gefunden wurde.", "INFO")

            except Exception as e:
                is_ok = False
                _log(f"Vorlage '{os.path.basename(template_path)}': Konnte nicht gelesen werden. Fehler: {e}", "FATAL")

        _log("\n--- Validierung der Bilder (aus Excel-Daten) ---\n", "SEP")
        if bilder_ordner:
            image_references = set()
            for eintrag in datensaetze + [fallback_marken]:
                for key, value in eintrag.items():
                    if isinstance(value, str) and (value.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.svg'))):
                        image_references.add(value)
            
            _log(f"Gefundene Bild-Referenzen in Excel: {len(image_references)}", "INFO")
            for img_name in image_references:
                img_path = os.path.join(bilder_ordner, img_name)
                if not os.path.exists(img_path):
                    is_ok = False
                    _log(f"Bild '<strong>{img_name}</strong>' nicht gefunden", "WARN")
            if is_ok and image_references:
                 _log("Alle referenzierten Bilder wurden gefunden.", "SUCCESS")

        else:
            _log("Kein Bilder-Ordner angegeben, Prüfung übersprungen.", "WARN")

        _log("\n" + "="*60, "SEP")
        if is_ok:
            _log("Trockenlauf erfolgreich! Konfiguration scheint gültig.", "SUCCESS")
        else:
            _log("Trockenlauf mit Fehlern beendet. Bitte prüfen Sie die obigen Meldungen.", "ERROR")
        
        return is_ok

    except Exception as e:
        _log_handler(f"FATALER FEHLER im Trockenlauf: {e}\n{traceback.format_exc()}", "FATAL", log_callback)
        return False

def verarbeite_vorlagen(vorlagen_ordner, export_ordner, excel_path, log_callback, progress_callback, file_callback, header_row, svg_scale, png_compression, categories, datetime_utc_format, bilder_ordner=None, worker_thread=None, dry_run=False, is_dark_mode=False):
    """
    Die zentrale Orchestrierungsfunktion, die den gesamten Prozess steuert.
    
    Ablauf:
    1.  Lädt die Daten aus der Excel-Datei (Hauptdaten und Fallback-Marken).
    2.  Sucht nach allen SVG-Bildern, die in den Daten referenziert werden.
    3.  Konvertiert alle gefundenen, einzigartigen SVGs in PNGs und speichert sie
        in einem temporären Ordner, um die Performance zu verbessern (jedes SVG
        wird nur einmal konvertiert).
    4.  Durchsucht den Vorlagen-Ordner und teilt die gefundenen .docx-Dateien in
        'anlagen'-spezifische und 'allgemein'e Vorlagen auf.
    5.  Erstellt einen eindeutigen Export-Ordner für den aktuellen Durchlauf,
        benannt mit Zeitstempel und Projektname.
    6.  Verarbeitet die 'Anlagen'-Vorlagen: Iteriert über jeden Datensatz aus der
        Excel-Datei und füllt für jeden Datensatz jede Vorlage aus.
    7.  Verarbeitet die 'Allgemein'-Vorlagen: Füllt diese Vorlagen einmal aus,
        wobei der erste Datensatz und die Fallback-Marken als Datenquelle dienen.
    8.  Verwendet bei der Erstellung eine Fallback-Logik: Wenn ein Wert in einem
        Haupt-Datensatz fehlt, wird der entsprechende Wert aus den Fallback-Marken
        (zweites Excel-Blatt) verwendet.
    9.  Löscht nach Abschluss das temporäre Verzeichnis mit den PNGs.
    """
    if dry_run:
        return verarbeite_vorlagen_trockenlauf(vorlagen_ordner, excel_path, log_callback, header_row, bilder_ordner)

    def _log(msg, level="INFO"): _log_handler(msg, level, log_callback)
    _log("="*60+"\n", "SEP"); _log("Starte Dokumentenerstellung...", "INFO"); _log("="*60+"\n", "SEP")
    temp_dir = None
    try:
        datensaetze, fallback_marken = lade_excel_daten(excel_path, header_row, log_callback)
        if not datensaetze: _log("Keine Datensätze gefunden. Abbruch.", "WARN"); return
            
        temp_dir = tempfile.mkdtemp(prefix="svg_cache_")
        unique_svg_paths = set()
        if bilder_ordner:
            for eintrag in datensaetze + [fallback_marken]:
                for key, value in eintrag.items():
                    if isinstance(value, str) and value.lower().endswith('.svg'):
                        full_path = os.path.join(bilder_ordner, value)
                        if os.path.exists(full_path): unique_svg_paths.add(full_path)
        svg_png_map = {}
        if unique_svg_paths:
            _log(f"Konvertiere {len(unique_svg_paths)} einzigartige SVGs...", "INFO")
            for svg_path in unique_svg_paths:
                png_filename = os.path.splitext(os.path.basename(svg_path))[0] + ".png"
                temp_png_path = os.path.join(temp_dir, png_filename)
                if svg_to_png_file_pyside(svg_path, temp_png_path, _log, svg_scale, png_compression):
                    svg_png_map[svg_path] = temp_png_path
            _log("SVG-Konvertierung abgeschlossen.", "SUCCESS")

        anlagen_templates, allgemein_templates = [], []
        for root, _, files in os.walk(vorlagen_ordner):
            for file in files:
                if file.endswith('.docx') and not file.startswith('~'):
                    path_parts = os.path.relpath(root, vorlagen_ordner).lower().split(os.sep)
                    if 'anlagen' in path_parts: anlagen_templates.append(os.path.join(root, file))
                    elif 'allgemein' in path_parts: allgemein_templates.append(os.path.join(root, file))
        
        projekt_name = str(datensaetze[0].get('projekt_name', fallback_marken.get('projekt_name', 'Unbenanntes_Projekt'))).strip()
        haupt_export_ordner = os.path.join(export_ordner, datetime.now().strftime('%Y-%m-%d_%H%M%S') + "_" + projekt_name)

        total_docs = (len(datensaetze) * len(anlagen_templates)) + len(allgemein_templates)
        current_doc = 0

        def process_and_save(template, data, is_anlage):
            nonlocal current_doc; current_doc += 1
            if worker_thread and worker_thread.isInterruptionRequested(): return

            progress_callback(current_doc, total_docs); file_callback(os.path.basename(template))
            try:
                context = data.copy()
                for key, fallback_value in fallback_marken.items():
                    primary_value = context.get(key)
                    is_empty = pd.isna(primary_value) or (isinstance(primary_value, str) and not primary_value.strip())
                    if is_empty: context[key] = fallback_value
                
                # datetime_utc dynamisch hinzufügen/überschreiben
                try:
                    # Nur aktuelles Datum im Format JJJJ-MM-DD
                    context['datetime_utc'] = datetime.now().strftime('%Y-%m-%d')
                except Exception as e:
                    context['datetime_utc'] = f"[Format-Fehler: {e}]"
                    _log(f"Ungültiges Zeitstempel-Format: '{datetime_utc_format}'. Fehler: {e}", "WARN")

                context['_bilder_ordner'] = bilder_ordner

                doc = ersetze_platzhalter_mit_docxtpl(template, context, svg_png_map, _log)
                _speichere_dokument(doc, template, data, is_anlage, haupt_export_ordner, categories, _log, is_dark_mode)
            except TemplateSyntaxError as e:
                _log(f"FEHLER in Vorlage '{os.path.basename(template)}': Ungültige Syntax. Bitte prüfen Sie die Platzhalter.", "FATAL")
                _log(f"Details: {e}", "ERROR")
            except Exception as e: _log(f"FEHLER bei '{os.path.basename(template)}': {e}", "ERROR")

        _log("\n--- Verarbeitung 'Anlagen' ---\n", "SEP")
        for eintrag in datensaetze:
            if worker_thread and worker_thread.isInterruptionRequested(): break
            for vorlage_pfad in anlagen_templates: process_and_save(vorlage_pfad, eintrag, True)
        
        if worker_thread and not worker_thread.isInterruptionRequested():
            _log("\n--- Verarbeitung 'Allgemein' ---\n", "SEP")
            if allgemein_templates:
                for vorlage_pfad in allgemein_templates: process_and_save(vorlage_pfad, datensaetze[0], False)
        
        if worker_thread and worker_thread.isInterruptionRequested():
            _log("\n" + "="*60, "SEP"); _log("Vorgang vom Benutzer abgebrochen.", "WARN")
        else:
            _log("\n" + "="*60, "SEP"); _log("Alle Aufgaben abgeschlossen.", "SUCCESS")
        
        return haupt_export_ordner
    except Exception as e: 
        _log_handler(f"FATALER FEHLER: {e}\n{traceback.format_exc()}", "FATAL", log_callback)
        return None
    finally:
        if temp_dir and os.path.exists(temp_dir):
            try: shutil.rmtree(temp_dir); _log_handler(f"Temporäres Cache-Verzeichnis gelöscht.", "INFO", log_callback)
            except Exception as e: _log_handler(f"Cache konnte nicht gelöscht werden: {e}", "ERROR", log_callback)

# ==============================================================================
# GUI-LOGIK
# ==============================================================================
# Dieser Abschnitt definiert die grafische Benutzeroberfläche mit PySide6.
# Er beinhaltet das Hauptfenster, die Tabs, alle Widgets (Knöpfe, Eingabefelder)
# und die Logik zur Interaktion mit dem Benutzer.

def resource_path(relative_path):
    """Gibt den absoluten Pfad zu einer Ressource zurück, auch im PyInstaller-EXE-Modus."""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

class PandasModel(QAbstractTableModel):
    """
    Ein benutzerdefiniertes Tabellenmodell für PySide6, das einen
    Pandas DataFrame als Datenquelle für eine QTableView verwenden kann.
    Dies ermöglicht die Anzeige der Excel-Daten direkt in der GUI.
    """
    def __init__(self, df): super().__init__(); self._df = df
    def rowCount(self, parent=None): return self._df.shape[0]
    def columnCount(self, parent=None): return self._df.shape[1]
    def data(self, index, role=Qt.DisplayRole):
        if index.isValid() and role == Qt.DisplayRole: return str(self._df.iloc[index.row(), index.column()])
    def headerData(self, col, orient, role=Qt.DisplayRole):
        if orient == Qt.Horizontal and role == Qt.DisplayRole: return self._df.columns[col]

class Worker(QThread):
    """
    Ein QThread-Worker, der die zeitintensive Dokumentenerstellung in einem
    separaten Thread ausführt. Dies verhindert, dass die GUI während des
    Prozesses "einfriert" und nicht mehr reagiert.
    
    Signale:
    - log: Sendet Log-Nachrichten an die GUI.
    - progress: Sendet den Fortschritt (aktueller Wert, Maximalwert).
    - finished: Signalisiert das Ende der Verarbeitung.
    - current_file: Sendet den Namen der aktuell bearbeiteten Datei.
    """
    log = Signal(str) # html_message
    progress = Signal(int, int)
    finished = Signal(bool, str) # success, export_path
    current_file = Signal(str)

    def __init__(self, **kwargs):
        super().__init__()
        self.params = kwargs
        self.dry_run = kwargs.get('dry_run', False)

    def run(self):
        """Startet die Verarbeitung durch Aufruf der `verarbeite_vorlagen` Funktion."""
        export_path = None
        try:
            thread_safe_params = self.params.copy()
            thread_safe_params['log_callback'] = self.log.emit
            thread_safe_params['progress_callback'] = self.progress.emit
            thread_safe_params['file_callback'] = self.current_file.emit
            thread_safe_params['worker_thread'] = self
            thread_safe_params.pop('theme', None)
            
            result = verarbeite_vorlagen(**thread_safe_params)
            
            if self.dry_run:
                self.finished.emit(result, None) # Bei Trockenlauf ist das Ergebnis ein Boolean
            elif not self.isInterruptionRequested():
                self.finished.emit(bool(result), result) # Bei echtem Lauf ist es der Pfad
            else:
                self.finished.emit(False, None)

        except Exception as e:
            error_msg = f"FATALER FEHLER im Worker-Thread: {e}\n{traceback.format_exc()}"
            _log_handler(error_msg, "FATAL", self.log.emit)
            self.finished.emit(False, None)

class MainWindow(QMainWindow):
    """
    Das Hauptfenster der Anwendung.
    Es initialisiert die Benutzeroberfläche, verwaltet die Benutzereingaben,
    lädt und speichert Einstellungen und startet den Worker-Thread.
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f'DocxTpl Automatisierung v{VERSION} - Build {BUILD_DATE}')
        self.category_widgets = {}
        self.theme_action_group = None
        self.last_export_path = None
        self.load_settings_and_init_vars()
        self.setup_ui()

    def load_settings_and_init_vars(self):
        """Lädt die zuletzt verwendeten Einstellungen (Pfade, Optionen) aus einer JSON-Datei."""
        last = self.load_settings()
        self.paths = {
            'excel_path': last.get('excel_path', ''),
            'vorlagen_ordner': last.get('vorlagen_ordner', ''),
            'bilder_ordner': last.get('bilder_ordner', ''),
            'export_ordner': last.get('export_ordner', '')
        }
        self.settings = {
            'header_row': last.get('header_row', 3),
            'svg_scale': last.get('svg_scale', 3),
            'png_compression': last.get('png_compression', -1),
            'theme': last.get('theme', 'Light'),
            'datetime_utc_format': last.get('datetime_utc_format', '%Y-%m-%d %H:%M:%S UTC')
        }
        self.categories = last.get('categories', {'b_': 'Beschilderung', 'ba_': 'Betriebsanweisung'})

    def setup_ui(self):
        """Erstellt und arrangiert alle UI-Elemente im Hauptfenster."""
        self.resize(800, 700)
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        tab_widget = QTabWidget()
        main_layout.addWidget(tab_widget)

        # Tab 1: Hauptsteuerung
        controls_tab = QWidget()
        tab_widget.addTab(controls_tab, "Hauptsteuerung")
        self.setup_controls_tab(controls_tab)

        # Tab 2: Kategorien
        categories_tab = QWidget()
        tab_widget.addTab(categories_tab, "Kategorien")
        self.setup_categories_tab(categories_tab)

        # Tab 3: Logs
        logs_tab = QWidget()
        tab_widget.addTab(logs_tab, "Logs")
        self.setup_logs_tab(logs_tab)

        self.setup_menu()
        self.apply_theme(self.settings.get('theme', 'Light'))
        
        for key in self.paths:
            self.validate_path(getattr(self, f"{key}_edit"))

        self.show_excel_data()
        
        # Zeige Versions-Info in den Logs
        try:
            import PySide6
            pyside_version = PySide6.__version__
        except ImportError:
            pyside_version = "Nicht verfügbar"
            
        self.log_text.append(f'<div style="color: #666; text-align: center; padding: 10px; border: 1px solid #ccc; margin: 10px 0; background-color: #f9f9f9;">'
                            f'<strong>DocxTpl Automatisierung v{VERSION}</strong><br>'
                            f'Build: {BUILD_DATE}<br>'
                            f'Python: {sys.version.split()[0]} | PySide6: {pyside_version}</div>')

    def setup_controls_tab(self, tab):
        """Erstellt den Inhalt des 'Hauptsteuerung'-Tabs."""
        layout = QVBoxLayout(tab)
        path_map = {'excel_path': 'Excel-Datei', 'vorlagen_ordner': 'Vorlagen-Ordner',
                    'bilder_ordner': 'Bilder-Ordner (optional)', 'export_ordner': 'Export-Ordner'}
        for key, name in path_map.items():
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{name}:"))
            path_edit = QLineEdit(self.paths.get(key, ''))
            path_edit.editingFinished.connect(lambda k=key, p=path_edit: self.on_path_change(k, p.text()))
            path_edit.textChanged.connect(lambda text, widget=path_edit: self.validate_path(widget))
            setattr(self, f"{key}_edit", path_edit)
            row.addWidget(path_edit)
            browse_btn = QPushButton("...")
            browse_btn.clicked.connect(lambda c, k=key: self.browse(k))
            row.addWidget(browse_btn)
            # Excel-Icon-Button nur beim Excel-Pfad
            if key == 'excel_path':
                excel_icon_btn = QPushButton()
                excel_icon_btn.setIcon(QIcon('Pictures/excel_icon.png'))  # Passe den Pfad ggf. an
                excel_icon_btn.setToolTip("Excel-Datei öffnen")
                excel_icon_btn.setFixedWidth(32)
                excel_icon_btn.clicked.connect(self.open_excel_file)
                row.addWidget(excel_icon_btn)
            layout.addLayout(row)

        settings_layout = QHBoxLayout()
        setting_map = {'header_row': ('Header-Zeile', (1, 100), 3), 'svg_scale': ('SVG-Skala', (1, 10), 3),
                       'png_compression': ('PNG-Komp. (0=Max)', (-1, 100), -1)}
        for key, (name, r, default) in setting_map.items():
            settings_layout.addWidget(QLabel(name))
            spin = QSpinBox()
            spin.setRange(*r)
            spin.setValue(self.settings.get(key, default))
            spin.valueChanged.connect(self.save_all_settings)
            setattr(self, f"{key}_spin", spin)
            settings_layout.addWidget(spin)
        layout.addLayout(settings_layout)

        # Layout für UTC-Format
        utc_format_layout = QHBoxLayout()
        utc_format_layout.addWidget(QLabel("Zeitstempel-Format (datetime_utc):"))
        self.datetime_utc_format_edit = QLineEdit(self.settings.get('datetime_utc_format'))
        self.datetime_utc_format_edit.editingFinished.connect(self.save_all_settings)
        utc_format_layout.addWidget(self.datetime_utc_format_edit)
        layout.addLayout(utc_format_layout)

        # Versions-Info
        version_layout = QHBoxLayout()
        version_layout.addWidget(QLabel(f"Version: {VERSION}"))
        version_layout.addWidget(QLabel(f"Build: {BUILD_DATE}"))
        version_layout.addStretch()
        layout.addLayout(version_layout)

        layout.addWidget(QFrame(frameShape=QFrame.HLine))
        self.table = QTableView()
        layout.addWidget(self.table)
        self.progress = QProgressBar()
        layout.addWidget(self.progress)
        self.current_file_label = QLabel("Bereit zum Starten...")
        self.current_file_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.current_file_label)
        
        self.open_export_folder_btn = QPushButton("Export-Ordner öffnen")
        self.open_export_folder_btn.clicked.connect(self.open_export_folder)
        self.open_export_folder_btn.setVisible(False)
        layout.addWidget(self.open_export_folder_btn)
        
        btn_row = QHBoxLayout()
        self.dry_run_btn = QPushButton("Konfiguration prüfen")
        self.start_btn = QPushButton("Start")
        self.close_btn = QPushButton("Schließen")

        self.dry_run_btn.clicked.connect(self.start_dry_run)
        self.start_btn.clicked.connect(self.start)
        self.close_btn.clicked.connect(self.handle_close_or_cancel)

        btn_row.addWidget(self.dry_run_btn)
        btn_row.addWidget(self.start_btn)
        btn_row.addWidget(self.close_btn)
        layout.addLayout(btn_row)

    def setup_categories_tab(self, tab):
        """Erstellt den Inhalt des 'Kategorien'-Tabs für die Verwaltung der Ausgabeordner."""
        layout = QVBoxLayout(tab)
        
        # Versions-Info
        version_info = QLabel(f"Version {VERSION} - Build {BUILD_DATE}")
        version_info.setStyleSheet("color: #666; font-size: 10px; padding: 5px;")
        version_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(version_info)
        
        group_box = QGroupBox("Dokumentkategorien")
        layout.addWidget(group_box)
        group_layout = QVBoxLayout(group_box)
        
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        self.categories_ui_layout = QVBoxLayout(scroll_content)
        scroll.setWidget(scroll_content)

        for prefix, name in self.categories.items():
            self.add_category_widget(prefix, name)
            
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("+"); add_btn.setFixedWidth(30); add_btn.setToolTip("Neue Kategorie hinzufügen")
        add_btn.clicked.connect(self.add_category)
        remove_btn = QPushButton("-"); remove_btn.setFixedWidth(30); remove_btn.setToolTip("Ausgewählte Kategorie entfernen")
        remove_btn.clicked.connect(self.remove_category)
        btn_layout.addWidget(add_btn); btn_layout.addWidget(remove_btn); btn_layout.addStretch()
        
        group_layout.addLayout(btn_layout)
        group_layout.addWidget(scroll)

        save_btn = QPushButton("Kategorien speichern")
        save_btn.clicked.connect(self.save_categories)
        layout.addWidget(save_btn)
        layout.addStretch()

    def setup_logs_tab(self, tab):
        """Erstellt den Inhalt des 'Logs'-Tabs mit Filter-Dropdown."""
        layout = QVBoxLayout(tab)
        
        # Versions-Info (immer sichtbar)
        version_info = QLabel(f"DocxTpl Automatisierung v{VERSION} - Build {BUILD_DATE}")
        version_info.setStyleSheet("color: #666; font-size: 10px; padding: 5px;")
        version_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(version_info)
        
        # Filter-Dropdown
        self.log_filter_combo = QComboBox()
        self.log_filter_combo.addItems([
            "ALLE", "INFO", "WARN", "ERROR", "SUCCESS", "FATAL", "SEP"
        ])
        self.log_filter_combo.currentTextChanged.connect(self.apply_log_filter)
        layout.addWidget(self.log_filter_combo)
        # Log-Textfeld
        self.log_text = QTextEdit(); self.log_text.setReadOnly(True)
        self.log_text.setContextMenuPolicy(Qt.CustomContextMenu)
        self.log_text.customContextMenuRequested.connect(self.show_log_context_menu)
        self.log_lines = []  # Speichert alle Log-Zeilen als (level, html) Tupel
        layout.addWidget(self.log_text)

    def setup_menu(self):
        """Erstellt die Menüleiste für die Anwendung, inkl. Theme-Auswahl."""
        menu_bar = self.menuBar()
        view_menu = menu_bar.addMenu("Ansicht")
        theme_menu = view_menu.addMenu("Theme")

        self.theme_action_group = QActionGroup(self)
        self.theme_action_group.setExclusive(True)
        
        actions = {
            "Light": QAction("Hell (Standard)", self, checkable=True),
            "Dark": QAction("Dark Mode", self, checkable=True),
            "Girly": QAction("Girly Mode", self, checkable=True)
        }
        
        for name, action in actions.items():
            action.setData(name)
            self.theme_action_group.addAction(action)
            theme_menu.addAction(action)

        self.theme_action_group.triggered.connect(self.on_theme_selected)

    def on_theme_selected(self, action):
        """Wird aufgerufen, wenn ein Theme aus dem Menü ausgewählt wird."""
        theme_name = action.data()
        self.apply_theme(theme_name)
        self.settings['theme'] = theme_name
        self.save_all_settings()

    def apply_theme(self, theme_name):
        """Wendet das ausgewählte Stylesheet an und aktualisiert die Menü-Auswahl. Setzt im Dark Mode den Log-Text auf weiß."""
        stylesheet = self.get_stylesheet(theme_name)
        self.setStyleSheet(stylesheet) # Apply to main window
        for action in self.theme_action_group.actions():
            if action.data() == theme_name:
                action.setChecked(True)
                break
        # Im Dark Mode: Log-Text maximal hell und kontrastreich
        self.is_dark_mode = (theme_name == 'Dark')
        if self.is_dark_mode:
            self.log_text.setStyleSheet("color: #FFF; background-color: #181818; font-weight: 500;")
        else:
            self.log_text.setStyleSheet("")

    def get_stylesheet(self, theme_name):
        """Gibt den QSS-Stylesheet-String für das angeforderte Theme zurück."""
        theme_map = {
            'Dark': 'dark.qss',
            'Girly': 'girly.qss'
        }
        qss_file = theme_map.get(theme_name)
        if qss_file:
            qss_path = resource_path(qss_file)
            if os.path.exists(qss_path):
                try:
                    with open(qss_path, 'r') as f:
                        return f.read()
                except IOError:
                    return "" # Fallback to default
        return ""  # Light/Default Theme

    def on_path_change(self, key, value):
        """Wird aufgerufen, wenn ein Pfad geändert wird, und speichert die Einstellung."""
        self.paths[key] = value
        self.save_all_settings()
        
        # Aktualisiere Excel-Datenanzeige wenn sich der Excel-Pfad ändert
        if key == 'excel_path':
            self.show_excel_data()

    def browse(self, key):
        """Öffnet einen Datei- oder Ordnerdialog und aktualisiert das entsprechende Eingabefeld."""
        path_edit = getattr(self, f"{key}_edit")
        dialog_title = "Ordner auswählen" if 'ordner' in key else "Excel-Datei auswählen"
        start_path = path_edit.text()
        if 'ordner' in key:
            path = QFileDialog.getExistingDirectory(self, dialog_title, start_path)
        else:
            path, _ = QFileDialog.getOpenFileName(self, dialog_title, start_path, "*.xlsx *.xls")
        if path:
            path_edit.setText(path)
            self.on_path_change(key, path)
            self.validate_path(path_edit)

    def validate_path(self, line_edit_widget):
        """Prüft, ob der Pfad in einem QLineEdit existiert und färbt es entsprechend."""
        path = line_edit_widget.text()
        is_optional_empty = 'bilder_ordner' in self.paths and line_edit_widget == self.bilder_ordner_edit and not path
        
        valid_style = "background-color: #d4edda; color: #155724;" # Greenish
        invalid_style = "background-color: #f8d7da; color: #721c24;" # Reddish

        if os.path.exists(path) or is_optional_empty:
            line_edit_widget.setStyleSheet(valid_style)
        else:
            line_edit_widget.setStyleSheet(invalid_style)

    def add_category_widget(self, prefix, name):
        """Fügt der UI eine neue Zeile zur Eingabe einer Kategorie hinzu."""
        row_layout = QHBoxLayout()
        checkbox = QCheckBox()
        prefix_edit = QLineEdit(prefix)
        name_edit = QLineEdit(name)
        
        row_layout.addWidget(checkbox)
        row_layout.addWidget(QLabel("Präfix:"))
        row_layout.addWidget(prefix_edit)
        row_layout.addWidget(QLabel("Ordnername:"))
        row_layout.addWidget(name_edit)
        
        self.categories_ui_layout.addLayout(row_layout)
        self.category_widgets[row_layout] = (checkbox, prefix_edit, name_edit)

    def add_category(self):
        """Event-Handler, um eine neue, leere Kategorie-Zeile hinzuzufügen."""
        self.add_category_widget("", "")

    def remove_category(self):
        """Event-Handler, um alle ausgewählten Kategorie-Zeilen zu entfernen."""
        to_remove = [l for l, (c, _, _) in self.category_widgets.items() if c.isChecked()]
        for layout in to_remove:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget(): child.widget().deleteLater()
            self.categories_ui_layout.removeItem(layout)
            layout.deleteLater()
            del self.category_widgets[layout]

    def save_categories(self):
        """Sammelt die Daten aus den Kategorie-Eingabefeldern und speichert sie."""
        self.categories.clear()
        # Werte bestehen aus (checkbox, prefix_edit, name_edit)
        for checkbox, prefix_edit, name_edit in self.category_widgets.values():
            prefix = prefix_edit.text().strip()
            name = name_edit.text().strip()
            if prefix and name:
                self.categories[prefix] = name
        self.save_all_settings()
        QMessageBox.information(self, "Gespeichert", "Kategorien aktualisiert.")

    def save_all_settings(self):
        """Sammelt alle aktuellen Einstellungen und speichert sie in 'settings.json'."""
        for key in self.settings:
            if hasattr(self, f"{key}_spin"):
                self.settings[key] = getattr(self, f"{key}_spin").value()
        
        if hasattr(self, 'datetime_utc_format_edit'):
            self.settings['datetime_utc_format'] = self.datetime_utc_format_edit.text()

        try:
            with open('settings.json', 'w') as f:
                json.dump({**self.paths, **self.settings, 'categories': self.categories}, f, indent=4)
        except IOError as e: 
            self.log_text.append(f"Speicherfehler: {e}")
            _log_handler(f"Speicherfehler: {e}", "ERROR", self.append_html_log)

    def load_settings(self):
        """Lädt Einstellungen aus 'settings.json', falls die Datei existiert."""
        try:
            if os.path.exists('settings.json'):
                with open('settings.json', 'r') as f: return json.load(f)
        except (json.JSONDecodeError, IOError): pass
        return {}
        
    def start_dry_run(self):
        """Startet den Trockenlauf-Prozess."""
        if not all(self.paths.get(k) for k in ['excel_path', 'vorlagen_ordner']):
            QMessageBox.warning(self, "Fehlende Pfade", "Bitte Excel- und Vorlagen-Pfad für den Trockenlauf angeben!"); return
        self.save_all_settings()
        self.log_text.clear()
        
        worker_params = {**self.paths, **self.settings, 'categories': self.categories, 'dry_run': True, 'is_dark_mode': self.is_dark_mode}

        self.worker = Worker(**worker_params)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.log.connect(self.append_html_log)
        
        # Für den Trockenlauf brauchen wir diese Signale nicht, aber der Worker erwartet sie
        self.worker.current_file.connect(lambda: None)
        self.worker.progress.connect(lambda: None)
        
        self.worker.start()
        self.set_ui_running_state(True)
    
    def start(self, dry_run=False):
        """
        Startet den Dokumentenerstellungsprozess.
        - Prüft, ob alle notwendigen Pfade angegeben sind.
        - Speichert die aktuellen Einstellungen.
        - Erstellt und startet den Worker-Thread mit allen notwendigen Parametern.
        - Deaktiviert den 'Start'-Knopf, um doppelte Ausführungen zu verhindern.
        """
        if not all(self.paths.get(k) for k in ['excel_path', 'vorlagen_ordner', 'export_ordner']):
            QMessageBox.warning(self, "Fehlende Pfade", "Bitte Excel-, Vorlagen- und Export-Pfad angeben!"); return
        self.save_all_settings()
        self.log_text.clear()
        
        worker_params = {**self.paths, **self.settings, 'categories': self.categories, 'dry_run': False, 'is_dark_mode': self.is_dark_mode}

        self.worker = Worker(**worker_params)
        
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.log.connect(self.append_html_log)
        self.worker.current_file.connect(self.update_current_file_label)
        self.worker.progress.connect(self.update_progress_bar)
        
        self.worker.start()
        self.set_ui_running_state(True)

    def set_ui_running_state(self, is_running):
        """Aktiviert/Deaktiviert UI-Elemente, während der Worker läuft."""
        self.start_btn.setEnabled(not is_running)
        self.dry_run_btn.setEnabled(not is_running)
        self.close_btn.setText("Abbrechen" if is_running else "Schließen")
        if not is_running:
            self.current_file_label.setText("Bereit zum Starten...")
        self.open_export_folder_btn.setVisible(False)

    def handle_close_or_cancel(self):
        """Schließt die App oder bricht den Worker ab."""
        if hasattr(self, 'worker') and self.worker.isRunning():
            reply = QMessageBox.question(self, "Abbrechen?", 
                                         "Möchten Sie den aktuellen Vorgang wirklich abbrechen?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.worker.requestInterruption()
        else:
            self.close()

    def on_worker_finished(self, success, export_path):
        """Wird aufgerufen, wenn der Worker-Thread seine Arbeit beendet hat."""
        self.set_ui_running_state(False)
        self.last_export_path = export_path
        
        is_dry_run = self.worker.dry_run
        was_cancelled = self.worker.isInterruptionRequested()

        if was_cancelled:
            self.current_file_label.setText("Vorgang abgebrochen.")
            QMessageBox.warning(self, "Abgebrochen", "Der Vorgang wurde vom Benutzer abgebrochen.")
        elif is_dry_run:
            if success:
                self.current_file_label.setText("Trockenlauf erfolgreich!")
                QMessageBox.information(self, "Trockenlauf", "Konfigurationsprüfung erfolgreich abgeschlossen.")
            else:
                self.current_file_label.setText("Trockenlauf fehlgeschlagen.")
                QMessageBox.warning(self, "Trockenlauf", "Konfigurationsprüfung hat Fehler gefunden. Bitte Logs prüfen.")
        elif success and self.last_export_path:
            self.current_file_label.setText("Verarbeitung abgeschlossen!")
            self.open_export_folder_btn.setVisible(True)
            QMessageBox.information(self, "Fertig", "Alle Dokumente wurden erfolgreich erstellt.")
        else:
            self.current_file_label.setText("Verarbeitung mit Fehlern abgeschlossen.")
            QMessageBox.warning(self, "Fehler", "Die Verarbeitung wurde mit Fehlern abgeschlossen. Bitte Logs prüfen.")

    def update_current_file_label(self, filename):
        """Aktualisiert das Label, das den Namen der aktuell verarbeiteten Datei anzeigt."""
        self.current_file_label.setText(f"Verarbeite: {filename}...")

    def update_progress_bar(self, current, total):
        """Aktualisiert die Fortschrittsanzeige."""
        if total > 0:
            self.progress.setValue(int(current / total * 100))
        else:
            self.progress.setValue(0)

    def append_html_log(self, html):
        """Ein Slot, der HTML-formatierten Text sicher an das Log-Fenster anhängt und für Filter speichert."""
        import re
        # Robust: Suche nach [ LEVEL ] mit beliebigen Leerzeichen, unabhängig von HTML-Tags
        m = re.search(r'\[\s*([A-Z]+)\s*\]', html)
        level = m.group(1) if m else "INFO"
        self.log_lines.append((level, html))
        self.apply_log_filter()

    def apply_log_filter(self):
        """Zeigt nur die Log-Zeilen an, die dem gewählten Filter entsprechen."""
        filter_level = self.log_filter_combo.currentText()
        self.log_text.clear()
        for level, html in self.log_lines:
            if filter_level == "ALLE" or level == filter_level:
                self.log_text.moveCursor(QTextCursor.End)
                self.log_text.insertHtml(html)
        self.log_text.ensureCursorVisible()

    def show_log_context_menu(self, position):
        """Zeigt ein Kontextmenü für Log-Einträge mit Optionen zum Öffnen von Ordnern/Dateien/Textmarken."""
        cursor = self.log_text.cursorForPosition(position)
        cursor.select(QTextCursor.WordUnderCursor)
        selected_text = cursor.selectedText()
        
        # Suche nach Pfaden in der aktuellen Zeile
        line_cursor = self.log_text.cursorForPosition(position)
        line_cursor.movePosition(QTextCursor.StartOfLine)
        line_cursor.movePosition(QTextCursor.EndOfLine, QTextCursor.KeepAnchor)
        line_text = line_cursor.selectedText()
        
        menu = QMenu(self)
        
        # Extrahiere mögliche Pfade aus der Zeile
        paths = self.extract_paths_from_log_line(line_text)
        
        if paths:
            # Füge Menüpunkte für jeden gefundenen Pfad hinzu
            for path_type, path, display_name in paths:
                if path_type == "file":
                    action = menu.addAction(f"📄 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_file(p))
                elif path_type == "folder":
                    action = menu.addAction(f"📁 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_folder(p))
                elif path_type == "template":
                    action = menu.addAction(f"📝 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_template_file(p))
            menu.addSeparator()

        # Prüfe auf fehlende Platzhalter/Textmarken
        fehlende_marken = self.extract_missing_placeholders_from_log_line(line_text)
        if fehlende_marken:
            action = menu.addAction("➕ Textmarke(n) im Excel hinzufügen")
            action.triggered.connect(lambda checked, marken=fehlende_marken: self.add_placeholders_to_excel(marken))
            menu.addSeparator()

        # Prüfe auf fehlende Size-Textmarken
        fehlende_size_marken = self.extract_missing_size_placeholders_from_log_line(line_text)
        if fehlende_size_marken:
            action = menu.addAction("➕ Size-Textmarke(n) im Excel hinzufügen")
            action.triggered.connect(lambda checked, marken=fehlende_size_marken: self.add_placeholders_to_excel(marken))
            menu.addSeparator()

        # Standard-Menüpunkte
        copy_action = menu.addAction("📋 Ausgewählten Text kopieren")
        copy_action.triggered.connect(self.copy_selected_text)
        
        clear_action = menu.addAction("🗑️ Log löschen")
        clear_action.triggered.connect(self.clear_log)
        
        menu.exec_(self.log_text.mapToGlobal(position))

    def extract_paths_from_log_line(self, line_text):
        """Extrahiert mögliche Datei- und Ordnerpfade aus einer Log-Zeile."""
        paths = []
        import re
        # Suche nach Dateinamen in verschiedenen Log-Formaten
        # 1. Export-Log: Dokument: 'Dateiname.docx'
        datei_matches = re.findall(r"Dokument: '([^']+)'", line_text)
        # 2. Fehler/Warnung: Vorlage 'Dateiname.docx'
        vorlage_matches = re.findall(r"Vorlage '([^']+)'", line_text)
        # 3. Export-Log: in Ordner: 'Ordnername'
        ordner_matches = re.findall(r"in Ordner: '([^']+)'", line_text)

        # Exportierte Dateien suchen (wie bisher)
        if hasattr(self, 'last_export_path') and self.last_export_path:
            base_path = self.last_export_path
            # Füge Dateien aus Export-Log hinzu
            for datei in datei_matches:
                for category_name in self.categories.values():
                    potential_path = os.path.join(base_path, category_name, datei)
                    if os.path.exists(potential_path):
                        paths.append(("file", potential_path, datei))
                        break
            # Füge Dateien aus Fehler/Warnung hinzu
            for datei in vorlage_matches:
                for category_name in self.categories.values():
                    potential_path = os.path.join(base_path, category_name, datei)
                    if os.path.exists(potential_path):
                        paths.append(("file", potential_path, datei))
                        break
            # Füge Ordner hinzu
            for ordner in ordner_matches:
                potential_path = os.path.join(base_path, ordner)
                if os.path.exists(potential_path):
                    paths.append(("folder", potential_path, ordner))

        # Vorlagen-Dateien suchen
        # Suche in self.paths['vorlagen_ordner'] und allen Unterordnern nach passenden Dateinamen
        if hasattr(self, 'paths') and 'vorlagen_ordner' in self.paths:
            vorlagen_root = self.paths['vorlagen_ordner']
            alle_vorlagen = set(datei_matches + vorlage_matches)
            for suchname in alle_vorlagen:
                for root, _, files in os.walk(vorlagen_root):
                    for file in files:
                        if file == suchname:
                            full_path = os.path.join(root, file)
                            paths.append(("template", full_path, suchname))
        return paths

    def extract_missing_placeholders_from_log_line(self, line_text):
        """Extrahiert fehlende Platzhalter/Textmarken aus einer Log-Zeile."""
        import re
        # Sucht nach: Fehlende Platzhalter in Excel: {'foo', 'bar'}
        match = re.search(r"Fehlende Platzhalter in Excel: (\{.*?\})", line_text)
        if match:
            try:
                # Sichere Auswertung des Sets
                marken = eval(match.group(1), {"__builtins__": None}, {})
                if isinstance(marken, set):
                    return sorted(marken)
            except Exception:
                pass
        return []

    def extract_missing_size_placeholders_from_log_line(self, line_text):
        """Extrahiert fehlende Size-Textmarken aus einer Log-Zeile."""
        import re
        # NEU: Erkenne auch die neue Info-Log-Zeile
        # Beispiel: Hinweis: Für Platzhalter 'ba_logo_img' wird die Standardgröße verwendet, da kein 'ba_logo_img_size' oder 'ba_logo_img_Size' in Excel/Fallback gefunden wurde.
        match_alt = re.search(r"Keine Größenangabe\s*\(([^)]+)\).*?für Platzhalter", line_text)
        match_neu = re.search(r"Hinweis: Für Platzhalter '([^']+)' wird die Standardgröße verwendet, da kein '([^']+)' oder '([^']+)' in Excel/Fallback gefunden wurde", line_text)
        if match_alt:
            inhalt = match_alt.group(1)
            self.append_html_log(f'<span style="color:#888">[DEBUG] Klammer-Inhalt: {inhalt}</span><br>')
            marken = re.findall(r"'([^']+)'", inhalt)
            if marken:
                self.append_html_log(f'<span style="color:#888">[DEBUG] Erkannte Size-Textmarken: {", ".join(marken)}</span><br>')
            else:
                self.append_html_log(f'<span style="color:#888">[DEBUG] Keine Size-Textmarken erkannt.</span><br>')
            return sorted(marken)
        elif match_neu:
            # Extrahiere die beiden _size-Varianten
            size1 = match_neu.group(2)
            size2 = match_neu.group(3)
            marken = [size1, size2]
            self.append_html_log(f'<span style="color:#888">[DEBUG] Erkannte Size-Textmarken (neu): {", ".join(marken)}</span><br>')
            return sorted(marken)
        self.append_html_log(f'<span style="color:#888">[DEBUG] Keine Size-Textmarken erkannt.</span><br>')
        return []

    def add_placeholders_to_excel(self, placeholders):
        """Fügt die angegebenen Platzhalter als neue Zeilen im zweiten Blatt der aktuellen Excel-Datei hinzu (Spalte 2=Textmarke, Spalte 3=leer)."""
        from openpyxl import load_workbook
        import os
        excel_path = self.paths.get('excel_path')
        if not excel_path or not os.path.exists(excel_path):
            QMessageBox.warning(self, "Fehler", "Excel-Datei nicht gefunden!")
            return
        try:
            wb = load_workbook(excel_path)
            if len(wb.sheetnames) < 2:
                QMessageBox.warning(self, "Fehler", "Die Excel-Datei hat kein zweites Blatt!")
                return
            ws = wb[wb.sheetnames[1]]
            # Nur eine Size-Variante pro Basis-Textmarke einfügen
            filtered = {}
            for marke in placeholders:
                if marke.lower().endswith('_size'):
                    base = marke[:-5].lower()
                    # Bevorzuge die Variante mit kleinem 's'
                    if base not in filtered or marke.endswith('_size'):
                        filtered[base] = marke
                else:
                    filtered[marke] = marke
            for marke in filtered.values():
                ws.append([None, marke, ""])
            wb.save(excel_path)
            QMessageBox.information(self, "Erfolg", f"{len(filtered)} Textmarke(n) wurden im zweiten Blatt hinzugefügt.")
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte Textmarken nicht hinzufügen: {e}")
    
    def open_file(self, file_path):
        """Öffnet eine Datei mit der Standard-Anwendung."""
        try:
            os.startfile(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte die Datei nicht öffnen: {e}")
    
    def open_folder(self, folder_path):
        """Öffnet einen Ordner im Explorer."""
        try:
            os.startfile(folder_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte den Ordner nicht öffnen: {e}")
    
    def open_template_file(self, file_path):
        """Öffnet eine Word-Vorlage mit der Standard-Anwendung."""
        try:
            os.startfile(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte die Vorlage nicht öffnen: {e}")
    
    def copy_selected_text(self):
        """Kopiert den ausgewählten Text in die Zwischenablage."""
        cursor = self.log_text.textCursor()
        if cursor.hasSelection():
            text = cursor.selectedText()
            clipboard = QApplication.clipboard()
            clipboard.setText(text)
    
    def clear_log(self):
        """Löscht den gesamten Log-Inhalt."""
        reply = QMessageBox.question(self, "Log löschen", 
                                   "Möchten Sie wirklich den gesamten Log löschen?",
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.log_text.clear()

    def show_excel_data(self):
        """
        Versucht, die Daten aus der angegebenen Excel-Datei zu laden und
        in der Tabelle (QTableView) in der GUI anzuzeigen.
        """
        path = self.paths.get('excel_path')
        header_row = self.settings.get('header_row', 3)
        if path and os.path.exists(path):
            try:
                df = pd.read_excel(path, header=header_row - 1)
                self.table.setModel(PandasModel(df))
            except Exception as e:
                self.table.setModel(None) # Clear table on error
                QMessageBox.critical(self, "Fehler beim Lesen der Excel-Datei", str(e))
    
    def open_export_folder(self):
        """Öffnet den zuletzt verwendeten Export-Ordner im Datei-Explorer."""
        if self.last_export_path and os.path.isdir(self.last_export_path):
            try:
                os.startfile(self.last_export_path)
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Konnte den Ordner nicht öffnen: {e}")

    def open_excel_file(self):
        excel_path = self.paths.get('excel_path')
        if excel_path and os.path.exists(excel_path):
            try:
                os.startfile(excel_path)
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Excel konnte nicht geöffnet werden: {e}")
        else:
            QMessageBox.warning(self, "Fehler", "Excel-Datei nicht gefunden!")

    def closeEvent(self, event):
        """Stellt sicher, dass der Worker-Thread sauber beendet wird, wenn das Fenster geschlossen wird."""
        if hasattr(self, 'worker') and self.worker.isRunning():
            self.worker.requestInterruption()
            self.worker.wait()
        event.accept()

def main():
    """Hauptfunktion, die die QApplication startet und das Hauptfenster anzeigt."""
    try:
        app = QApplication(sys.argv)
        if os.path.exists('Pictures/Logo.png'): app.setWindowIcon(QIcon('Pictures/Logo.png'))
        win = MainWindow(); win.show(); sys.exit(app.exec())
    except Exception: print(f"Kritischer Fehler:\n{traceback.format_exc()}", file=sys.stderr)

if __name__ == '__main__':
    main() 