# -*- coding: utf-8 -*-
"""Seite: Excel-Import (Migration)."""
import streamlit as st
import pandas as pd
import sys
import os
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.auth import require_admin
from utils.supabase_client import get_supabase
from utils.helpers import rows_to_list

require_admin()
sb = get_supabase()

st.title("Excel-Import")
st.caption("Bestehende Excel-Daten in Supabase migrieren.")

projekte = sb.table("projekte").select("id, name").order("name").execute()
proj_list = rows_to_list(projekte.data)
if not proj_list:
    st.warning("Bitte zuerst ein Projekt anlegen.")
    st.stop()

proj_options = {p["name"]: p["id"] for p in proj_list}
projekt_name = st.selectbox("Ziel-Projekt", options=list(proj_options.keys()))
projekt_id = proj_options[projekt_name]
header_row = st.number_input("Header-Zeile (Excel)", min_value=1, value=3)

uploaded = st.file_uploader("Excel-Datei (.xlsx)", type=["xlsx", "xls"])
if uploaded:
    try:
        df = pd.read_excel(uploaded, header=int(header_row) - 1).dropna(how="all")
        df.columns = [str(c).strip() for c in df.columns]
        df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

        # Blatt 2: Fallback-Marken
        wb = load_workbook(uploaded, read_only=True)
        fallback = {}
        if len(wb.sheetnames) > 1:
            ws = wb[wb.sheetnames[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 3 and row[1]:
                    key = str(row[1]).strip()
                    value = str(row[2]).strip() if row[2] is not None else ""
                    fallback[key] = value
        wb.close()

        # Nur Zeilen mit anlage_seriennummer
        serien_col = "anlage_seriennummer"
        if serien_col not in df.columns:
            st.error(f"Spalte '{serien_col}' nicht gefunden. Spalten: {list(df.columns)}")
        else:
            datensaetze = df[df[serien_col].notna()].to_dict("records")
            st.info(f"Gefunden: {len(datensaetze)} Datensätze, {len(fallback)} Fallback-Marken.")

            if st.button("Import starten"):
                with st.spinner("Importiere..."):
                    # Fallback-Marken
                    for k, v in fallback.items():
                        try:
                            sb.table("fallback_marken").upsert({
                                "projekt_id": projekt_id,
                                "schluessel": k,
                                "wert": v,
                            }, on_conflict="projekt_id,schluessel").execute()
                        except Exception as e:
                            st.warning(f"Fallback {k}: {e}")

                    # Anlagen + Daten
                    for i, row in enumerate(datensaetze):
                        sn = str(row.get(serien_col, "")).strip()
                        if not sn:
                            continue
                        try:
                            # Prüfen ob Anlage existiert
                            existing = sb.table("anlagen").select("id").eq("projekt_id", projekt_id).eq("seriennummer", sn).execute()
                            if existing.data:
                                anlage_id = existing.data[0]["id"]
                            else:
                                ins = sb.table("anlagen").insert({
                                    "projekt_id": projekt_id,
                                    "seriennummer": sn,
                                }).execute()
                                anlage_id = ins.data[0]["id"]
                            for key, val in row.items():
                                if key == serien_col or pd.isna(val):
                                    continue
                                sb.table("anlagen_daten").upsert({
                                    "anlage_id": anlage_id,
                                    "schluessel": str(key).strip(),
                                    "wert": str(val).strip() if val is not None else "",
                                }, on_conflict="anlage_id,schluessel").execute()
                        except Exception as e:
                            st.error(f"Fehler bei {sn}: {e}")
                            break

                st.success(f"Import abgeschlossen. {len(datensaetze)} Anlagen migriert.")
                st.rerun()

    except Exception as e:
        st.error(f"Fehler beim Lesen: {e}")
        import traceback
        st.code(traceback.format_exc())
