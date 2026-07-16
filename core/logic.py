import io
import os
import shutil
import tempfile
import threading
import time
import traceback
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from datetime import datetime

import pandas as pd
import qrcode
from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from jinja2.exceptions import TemplateSyntaxError
from openpyxl import load_workbook

from core.logging import _log_handler
from core.rules import (
    count_anlagen_emits_with_rules,
    evaluate_anlage_template_decision,
)
from core.utils import svg_to_png_file_pyside


def _norm_template_path(p):
    """Absoluter, plattformnormalisierter Pfad für Vergleiche (Windows: normcase)."""
    return os.path.normcase(os.path.normpath(os.path.abspath(p)))


def _max_word_com_workers():
    """Grenze für parallele Word-COM-Instanzen (PDF, Seitenprüfung)."""
    return max(1, min(4, (os.cpu_count() or 2)))


def _max_docgen_thread_workers():
    """Grenze paralleler docxtpl-/Kopier-Jobs in Threads."""
    return max(1, min(4, (os.cpu_count() or 2)))


def _format_rule_decision_log(decision: dict, eintrag: dict, template_path: str) -> str:
    seriennummer = str(eintrag.get('anlage_seriennummer', 'Unbekannt')).strip()
    if not decision.get("has_rule"):
        return (
            f"[Regel] Anlage '{seriennummer}' | Vorlage '{os.path.basename(template_path)}' | "
            f"keine aktive Regel ({decision.get('reason')})"
        )
    rule_name = decision.get("rule_name") or "(ohne Name)"
    configured_branch = (decision.get("configured_branch") or "-").upper()
    applied_branch = (decision.get("applied_branch") or "-").upper()
    emit_label = "ERZEUGEN" if decision.get("emit") else "ÜBERSPRINGEN"
    return (
        f"[Regel] Anlage '{seriennummer}' | Vorlage '{os.path.basename(template_path)}' | "
        f"Regel '{rule_name}' | Vorlage-Zweig={configured_branch} | "
        f"Entscheid={applied_branch} | Ergebnis={emit_label} | Grund: {decision.get('reason')}"
    )


def sammle_vorlagen_pfade(vorlagen_ordner):
    """
    Sammelt .docx-Vorlagen unter anlagen/ bzw. allgemein/ (gleiche Regeln wie die Generierung).

    Returns:
        (anlagen_templates, allgemein_templates): Listen absoluter Pfade.
    """
    anlagen_templates, allgemein_templates = [], []
    if not vorlagen_ordner or not os.path.isdir(vorlagen_ordner):
        return anlagen_templates, allgemein_templates
    for root, _, files in os.walk(vorlagen_ordner):
        for file in files:
            if file.endswith('.docx') and not file.startswith('~'):
                path_parts = os.path.relpath(root, vorlagen_ordner).lower().split(os.sep)
                if 'anlagen' in path_parts:
                    anlagen_templates.append(os.path.join(root, file))
                elif 'allgemein' in path_parts:
                    allgemein_templates.append(os.path.join(root, file))
    return anlagen_templates, allgemein_templates


def liste_textmarken_aus_docx(doc_path: str, log_callback=None):
    """
    Liefert sortierte Text-Platzhalter aus einer .docx-Vorlage (ohne _img / _qr / _link),
    wie bei der Render-Reihenfolge in ersetze_platzhalter_mit_docxtpl.
    """
    if not doc_path or not os.path.isfile(doc_path):
        return []
    try:
        doc = DocxTemplate(doc_path)
        variables = doc.get_undeclared_template_variables()
    except Exception as e:
        _log_handler(f"Textmarken aus Vorlage: {e}", "WARN", log_callback)
        return []
    image_keys = {k for k in variables if k.endswith("_img")}
    qr_keys = {k for k in variables if k.endswith(("_qr", "_link"))}
    text_keys = sorted(k for k in variables if k not in image_keys and k not in qr_keys)
    return text_keys


def filter_vorlagen_nach_auswahl(anlagen_templates, allgemein_templates, selected_template_paths):
    """
    Schränkt die Listen auf gewählte Pfade ein.

    - selected_template_paths None: keine Einschränkung (alle übergebenen).
    - sonst: Schnittmenge nach normalisiertem Pfad.
    """
    if selected_template_paths is None:
        return list(anlagen_templates), list(allgemein_templates)
    sel = {_norm_template_path(p) for p in selected_template_paths}
    anlagen_f = [p for p in anlagen_templates if _norm_template_path(p) in sel]
    allg_f = [p for p in allgemein_templates if _norm_template_path(p) in sel]
    return anlagen_f, allg_f


def lade_excel_daten(pfad, header_row=3, log_callback=None):
    """
    Lädt Daten aus einer Excel-Datei.
    - Liest das erste Tabellenblatt in einen Pandas DataFrame. Jede Zeile, die eine
      'anlage_seriennummer' enthält, wird als ein Datensatz (dictionary) behandelt.
    - Liest das zweite Tabellenblatt als Quelle für Fallback-Textmarken. Dies ist
      nützlich für allgemeine Informationen (z.B. Projektname, Datum), die für
      alle Dokumente gleich sind.

    Args:
        pfad (str): Pfad zur Excel-Datei.
        header_row (int): Die Zeilennummer, die die Spaltenüberschriften enthält.
        log_callback (callable, optional): Callback für Log-Nachrichten.

    Returns:
        tuple: Ein Tupel mit (Liste von Datensätzen, Dictionary mit Fallback-Marken).
    """
    _log_handler(f"Lade Excel-Daten aus: {pfad}", "INFO", log_callback)
    if not os.path.isfile(pfad):
        raise FileNotFoundError(f"FEHLER: Excel-Datei nicht gefunden: {pfad}")
    try:
        try:
            df = pd.read_excel(pfad, header=header_row - 1).dropna(how='all')
        except (PermissionError, OSError) as e:
            if getattr(e, 'errno', None) == 13 or isinstance(e, PermissionError):
                from core.excel_com import lade_excel_daten_via_com
                result = lade_excel_daten_via_com(pfad, header_row, log_callback)
                if result is not None:
                    return result
            raise
        df.columns = [str(col).strip() for col in df.columns]
        df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
        datensaetze = [
            row.to_dict()
            for _, row in df.iterrows()
            if 'anlage_seriennummer' in row and pd.notna(row['anlage_seriennummer'])
        ]
        _log_handler(f"Erfolgreich {len(datensaetze)} Datensätze (Zeilen) geladen.", "SUCCESS", log_callback)

        fallback_marken = {}
        wb = None
        try:
            wb = load_workbook(pfad, read_only=True)
            if len(wb.sheetnames) > 1:
                ws = wb[wb.sheetnames[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 3 and row[1]:
                        key = str(row[1]).strip()
                        value = str(row[2]).strip() if row[2] is not None else ""
                        fallback_marken[key] = value
                _log_handler(
                    f"Erfolgreich {len(fallback_marken)} Fallback-Textmarken aus dem zweiten Blatt geladen.",
                    "SUCCESS",
                    log_callback
                )
        except (PermissionError, OSError) as e:
            if getattr(e, 'errno', None) == 13 or isinstance(e, PermissionError):
                from core.excel_com import lade_excel_daten_via_com
                result = lade_excel_daten_via_com(pfad, header_row, log_callback)
                if result is not None:
                    _, fallback_marken = result
            else:
                raise
        finally:
            if wb is not None:
                wb.close()

        return datensaetze, fallback_marken
    except Exception as e:
        raise Exception(f"FEHLER beim Lesen der Excel-Datei {pfad}: {e}") from e


def ersetze_platzhalter_mit_docxtpl(doc_path, context, svg_png_map, log_callback=None):
    """
    Ersetzt Platzhalter in einem Word-Dokument (.docx) mit den Daten aus dem Kontext.
    Nutzt docxtpl, was eine Jinja2-ähnliche Syntax erlaubt.

    Verarbeitungsreihenfolge (nach User-Wunsch für bessere Nachvollziehbarkeit):
    1. Bild-Platzhalter (_img)
    2. QR-Code-Platzhalter (_qr, _link)
    3. Alle anderen Text-Platzhalter

    Spezielle Logik:
    - Bilder: Platzhalter mit Suffix '_img' werden durch Bilder ersetzt. Die Bildgröße
      kann über einen zusätzlichen Platzhalter mit Suffix '_img_size' gesteuert werden.
    - QR-Codes: Platzhalter mit Suffix '_qr' oder '_link' werden in QR-Codes umgewandelt.
      Auch hier kann die Größe mit einem '_qr_size' oder '_link_size' Suffix angepasst werden.
    - SVG-Unterstützung: Verwendet die vorab konvertierten PNGs aus der svg_png_map.

    Args:
        doc_path (str): Pfad zur Word-Vorlagendatei.
        context (dict): Dictionary mit den Daten zum Ersetzen.
        svg_png_map (dict): Mapping von SVG-Pfaden zu temporären PNG-Pfaden.
        log_callback (callable, optional): Callback für Log-Nachrichten.

    Returns:
        DocxTemplate: Das bearbeitete Dokumentenobjekt.
    """
    doc = DocxTemplate(doc_path)
    try:
        undeclared_variables = doc.get_undeclared_template_variables()
    except Exception:
        undeclared_variables = set(context.keys())

    render_context = context.copy()
    bilder_ordner = context.get('_bilder_ordner')

    # Platzhalter nach Typ sortieren, um eine definierte Verarbeitungsreihenfolge zu gewährleisten.
    image_keys = sorted([k for k in undeclared_variables if k.endswith('_img')])
    qr_keys = sorted([k for k in undeclared_variables if k.endswith(('_qr', '_link'))])
    text_keys = sorted([k for k in undeclared_variables if k not in image_keys and k not in qr_keys])

    # Gewünschte Verarbeitungsreihenfolge: Bilder -> QR-Codes -> Text.
    processing_order = image_keys + qr_keys + text_keys

    for key in processing_order:
        value = context.get(key)
        if not isinstance(value, str) or not value.strip():
            continue

        # 1. Bild-Platzhalter verarbeiten
        if key in image_keys and bilder_ordner:
            width_cm = 15
            size_val = context.get(key + '_size', context.get(key + '_Size'))
            if pd.notna(size_val):
                try:
                    width_cm = float(size_val)
                except (ValueError, TypeError):
                    pass
            bild_pfad = os.path.join(bilder_ordner, value)
            if os.path.exists(bild_pfad):
                try:
                    render_context[key] = InlineImage(doc, svg_png_map.get(bild_pfad, bild_pfad), width=Cm(width_cm))
                except Exception as e:
                    render_context[key] = "[Bild-Fehler]"
                    _log_handler(f"Bild {value} Fehler: {e}", "ERROR", log_callback)
            else:
                render_context[key] = "[Bild nicht gefunden]"
                _log_handler(f"Bild {value} nicht gefunden", "WARN", log_callback)

        # 2. QR-Code-Platzhalter verarbeiten
        elif key in qr_keys:
            width_cm = 4
            size_val = context.get(key + '_size', context.get(key + '_Size'))
            if pd.notna(size_val):
                try:
                    width_cm = float(size_val)
                except (ValueError, TypeError):
                    pass
            try:
                qr_img = qrcode.make(value)
                img_bytes = io.BytesIO()
                qr_img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                render_context[key] = InlineImage(doc, img_bytes, width=Cm(width_cm))
            except Exception as e:
                render_context[key] = "[QR-Fehler]"
                _log_handler(f"QR-Code für '{value}' Fehler: {e}", "ERROR", log_callback)

    # 3. Alle Platzhalter (inkl. der vorbereiteten Bilder/QRs) von docxtpl rendern lassen.
    doc.render(render_context, autoescape=True)
    return doc


def _speichere_dokument(
    doc,
    vorlage_pfad,
    eintrag,
    is_anlage,
    haupt_export_ordner,
    categories,
    log_callback,
    is_dark_mode=False,
    seitenpruefung_pending=None,
    seitenpruefung_lock=None,
):
    """
    Speichert ein generiertes Word-Dokument im richtigen Ordner und mit dem korrekten Namen.

    - Kategorisierung: Der Dateiname der Vorlage wird geprüft. Beginnt er mit einem
      der in der GUI definierten Präfixe (z.B. 'B_'), wird das Dokument in den
      zugehörigen Kategorie-Ordner (z.B. 'Beschilderung') gespeichert.
    - Benennung:
        - Anlagenspezifische Dokumente: '{seriennummer}_{vorlagenname}.docx'
        - Allgemeine Dokumente: '{vorlagenname}.docx'
    - Optional: seitenpruefung_pending sammelt (Vorlage, Ziel, Dateiname) für batched Seitenprüfung.
    """
    vorlage_name = os.path.basename(vorlage_pfad)
    cleaned_name = vorlage_name
    export_kategorie_ordner = haupt_export_ordner
    for prefix, cat_name in categories.items():
        if vorlage_name.lower().startswith(prefix.lower()):
            cleaned_name = vorlage_name[len(prefix):]
            export_kategorie_ordner = os.path.join(haupt_export_ordner, cat_name)
            break
    os.makedirs(export_kategorie_ordner, exist_ok=True)
    seriennummer = str(eintrag.get('anlage_seriennummer', 'Unbekannt')).strip()
    ziel_name = f"{seriennummer}_{cleaned_name}" if is_anlage else cleaned_name

    # Speichere das Dokument
    ziel_pfad = os.path.join(export_kategorie_ordner, ziel_name)
    doc.save(ziel_pfad)

    if is_anlage:
        _log_handler(
            f"-> Speichere anlagenspez. Dokument: '{ziel_name}' für Anlage: '{seriennummer}' in Ordner: '{os.path.basename(export_kategorie_ordner)}'",
            "SUCCESS",
            log_callback,
            is_dark_mode
        )
    else:
        _log_handler(
            f"-> Speichere allgemeines Dokument: '{ziel_name}' in Ordner: '{os.path.basename(export_kategorie_ordner)}'",
            "SUCCESS",
            log_callback,
            is_dark_mode
        )

    if seitenpruefung_pending is not None:
        entry = (vorlage_pfad, ziel_pfad, ziel_name)
        if seitenpruefung_lock is not None:
            with seitenpruefung_lock:
                seitenpruefung_pending.append(entry)
        else:
            seitenpruefung_pending.append(entry)


def _lageplan_ziel_pfad(vorlage_pfad, eintrag, haupt_export_ordner, categories):
    """Gleiche Ordner- und Dateinamen-Logik wie _speichere_dokument (nur Anlagen)."""
    vorlage_name = os.path.basename(vorlage_pfad)
    cleaned_name = vorlage_name
    export_kategorie_ordner = haupt_export_ordner
    for prefix, cat_name in categories.items():
        if vorlage_name.lower().startswith(prefix.lower()):
            cleaned_name = vorlage_name[len(prefix):]
            export_kategorie_ordner = os.path.join(haupt_export_ordner, cat_name)
            break
    seriennummer = str(eintrag.get('anlage_seriennummer', 'Unbekannt')).strip()
    ziel_name = f"{seriennummer}_{cleaned_name}"
    return export_kategorie_ordner, ziel_name


def _lageplan_quell_unterordner_liste(categories):
    """Suchreihenfolge: Kategorie zu rl_, dann Lageplan, Plan."""
    seen = set()
    out = []
    for pref, name in categories.items():
        if str(pref).lower().startswith('rl') and name:
            if name not in seen:
                out.append(name)
                seen.add(name)
    for fb in ('Lageplan', 'Plan'):
        if fb not in seen:
            out.append(fb)
            seen.add(fb)
    return out


def find_lageplan_source_file(previous_export_root, categories, vorlage_pfad, eintrag):
    """Pfad zur .docx im letzten Export, falls vorhanden; sonst None."""
    if not previous_export_root or not os.path.isdir(previous_export_root):
        return None
    vorlage_name = os.path.basename(vorlage_pfad)
    if not vorlage_name.lower().startswith('rl_'):
        return None
    _, ziel_name = _lageplan_ziel_pfad(vorlage_pfad, eintrag, previous_export_root, categories)
    for sub in _lageplan_quell_unterordner_liste(categories):
        candidate = os.path.normpath(os.path.join(previous_export_root, sub, ziel_name))
        if os.path.isfile(candidate):
            return candidate
    return None


def _normalize_windows_path_for_word_com(path):
    """Normalisiert Pfade für Word-COM (Backslashes, absolut, ohne URL-Encoding)."""
    normalized = os.path.normpath(os.path.abspath(path))
    return normalized.replace("/", "\\")


# Word VBA: wdStatisticPages
_WD_STATISTIC_PAGES = 2


def _word_seitenanzahl_in_docx(docx_path):
    """
    Liefert die von Word berechnete Seitenzahl der .docx, oder None bei Fehler
    (nicht Windows, kein Word/pywin32, Datei fehlt, COM).
    """
    if os.name != "nt" or not docx_path or not os.path.isfile(docx_path):
        return None
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return None
    path = _normalize_windows_path_for_word_com(docx_path)
    word = None
    doc = None
    n_out = None
    try:
        pythoncom.CoInitialize()
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        doc = word.Documents.Open(path, ReadOnly=True, AddToRecentFiles=False)
        n_raw = int(doc.Content.ComputeStatistics(_WD_STATISTIC_PAGES))
        n_out = n_raw if n_raw > 0 else None
    except Exception:
        n_out = None
    finally:
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
    return n_out


def _seitenpruefung_zeile_sequenziell(
    n_v, vorlage_basename, ausgabe_pfad, kurzname, log_callback, is_dark_mode
):
    """
    Eine Seitenvergleichszeile (Hilfslogik, Fallback ohne Multiprocessing).
    n_v: vorher ermittelte Seitenzahl der Vorlage oder None.
    """
    n_a = _word_seitenanzahl_in_docx(ausgabe_pfad)
    if n_v is None or n_a is None or n_v < 1 or n_a < 1:
        return None
    if n_a > n_v:
        msg = (
            f"Seitenprüfung: „{kurzname}“ hat {n_a} Seite(n), die Vorlage "
            f"„{vorlage_basename}“ hat {n_v} — bitte Inhalt/Layout prüfen."
        )
        _log_handler(msg, "WARN", log_callback, is_dark_mode)
        return msg
    return None


def _seiten_row_mp_worker(args):
    """
    Worker für ProcessPool: (n_v, ausgabe_pfad, kurzname, vorlage_basename) -> warn_msg oder None
    muss Modul-Top-Level sein (Windows-spawn).
    """
    n_v, ausgabe_pfad, kurzname, vorlage_basename = args
    if n_v is None or n_v < 1:
        return None
    n_a = _word_seitenanzahl_in_docx(ausgabe_pfad)
    if n_a is None or n_a < 1:
        return None
    if n_a > n_v:
        return (
            f"Seitenprüfung: „{kurzname}“ hat {n_a} Seite(n), die Vorlage "
            f"„{vorlage_basename}“ hat {n_v} — bitte Inhalt/Layout prüfen."
        )
    return None


def _fuehre_seitenpruefung_batched(
    pending,
    log_callback,
    is_dark_mode,
) -> list:
    """
    Führt Sammel-Seitenvergleiche aus (Vorlagen-Seitenzahlen pro Vorlage 1×, erzeugt parallel).
    pending: list[tuple(vorlage_pfad, ausgabe_pfad, kurzname_ausgabe)]
    Rückgabe: Liste der Warn-Strings; schreibt WARN ins Log.
    """
    if not pending or os.name != "nt":
        return []
    warnungen = []
    workers = _max_word_com_workers()
    v_unique = {}
    for v, a, n in pending:
        kn = _norm_template_path(v)
        if kn not in v_unique:
            v_unique[kn] = v
    u_paths = list(v_unique.values())
    v_to_n = {}
    try:
        with ProcessPoolExecutor(max_workers=workers) as ex:
            counts = list(ex.map(_word_seitenanzahl_in_docx, u_paths))
        v_to_n = {_norm_template_path(p): c for p, c in zip(u_paths, counts)}
    except Exception:
        for p in u_paths:
            v_to_n[_norm_template_path(p)] = _word_seitenanzahl_in_docx(p)
    try:
        row_args = [
            (v_to_n.get(_norm_template_path(v), None), a, n, os.path.basename(v))
            for (v, a, n) in pending
        ]
        with ProcessPoolExecutor(max_workers=workers) as ex:
            for msg in ex.map(_seiten_row_mp_worker, row_args):
                if msg:
                    _log_handler(msg, "WARN", log_callback, is_dark_mode)
                    warnungen.append(msg)
    except Exception:
        cache = {_norm_template_path(p): n for p, n in v_to_n.items()}
        for v, a, name in pending:
            nv = cache.get(_norm_template_path(v))
            msg = _seitenpruefung_zeile_sequenziell(
                nv, os.path.basename(v), a, name, log_callback, is_dark_mode
            )
            if msg:
                warnungen.append(msg)
    return warnungen


def _docx_to_pdf_word_com_single(docx_path):
    """
    Konvertiert eine .docx zu .pdf im gleichen Ordner (Word-COM, Windows).
    Rückgabe: (erfolg: bool, fehler oder None).
    """
    try:
        import pythoncom
        import pywintypes
        import win32com.client
    except ImportError as e:
        return False, e

    wd_export_format_pdf = 17
    wd_export_optimize_for_print = 0
    wd_export_all_document = 0
    wd_export_document_content = 0
    wd_export_create_no_bookmarks = 0
    transient_hresults = {-2147418111, -2147417848}

    docx_path = _normalize_windows_path_for_word_com(docx_path)
    pdf_path = _normalize_windows_path_for_word_com(os.path.splitext(docx_path)[0] + ".pdf")

    if not os.path.isfile(docx_path):
        return False, FileNotFoundError(f"Datei nicht gefunden: {docx_path}")

    last_error = None
    for attempt in range(1, 4):
        word = None
        doc = None
        try:
            pythoncom.CoInitialize()
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            doc = word.Documents.Open(docx_path, ReadOnly=True, AddToRecentFiles=False)
            doc.ExportAsFixedFormat(
                OutputFileName=pdf_path,
                ExportFormat=wd_export_format_pdf,
                OpenAfterExport=False,
                OptimizeFor=wd_export_optimize_for_print,
                Range=wd_export_all_document,
                From=1,
                To=1,
                Item=wd_export_document_content,
                IncludeDocProps=True,
                KeepIRM=True,
                CreateBookmarks=wd_export_create_no_bookmarks,
                DocStructureTags=True,
                BitmapMissingFonts=True,
                UseISO19005_1=False
            )
            return True, None
        except pywintypes.com_error as e:
            last_error = e
            hresult = e.args[0] if e.args else None
            if hresult in transient_hresults and attempt < 3:
                import time
                time.sleep(0.8 * attempt)
                continue
            return False, e
        except Exception as e:
            last_error = e
            return False, e
        finally:
            if doc is not None:
                try:
                    doc.Close(False)
                except Exception:
                    pass
            if word is not None:
                try:
                    word.Quit()
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    return False, last_error


def _docx_to_pdf_com_mp_wrapper(docx_path: str) -> tuple:
    """Für ProcessPool: Rückgabe (docx_path, ok, fehlerstring). Muss top-level liegen."""
    try:
        ok, err = _docx_to_pdf_word_com_single(docx_path)
        s = repr(err) if err is not None and not ok else None
        return (docx_path, ok, s)
    except Exception as e:
        return (docx_path, False, repr(e))


def konvertiere_ein_docx_zu_pdf(docx_path, log_callback=None):
    """Eine .docx-Datei per Word-COM in PDF umwandeln; loggt Ergebnis."""
    ok, err = _docx_to_pdf_word_com_single(docx_path)
    if isinstance(err, ImportError):
        _log_handler(
            "PDF-Export: pywin32 nicht verfügbar. Bitte in der aktiven Umgebung installieren: "
            "pip install pywin32 && pywin32_postinstall.py -install",
            "WARN",
            log_callback,
        )
        return False
    if ok:
        _log_handler(
            f"PDF erstellt: {os.path.basename(os.path.splitext(docx_path)[0] + '.pdf')}",
            "SUCCESS",
            log_callback,
        )
        return True
    _log_handler(f"PDF fehlgeschlagen für '{os.path.basename(docx_path)}': {err}", "ERROR", log_callback)
    return False


def _convert_docx_to_pdf_in_folder(ordner, log_callback, progress_callback=None, file_callback=None):
    """Konvertiert alle .docx-Dateien im Ordner (rekursiv) in PDFs über Word-COM (Windows). Mehrere parallel."""
    result = {
        "docx_count": 0,
        "success_count": 0,
        "failure_count": 0,
        "error": None,
    }
    try:
        import pywintypes  # noqa: F401 — Prüfung, ob pywin32 nutzbar ist
        import win32com.client  # noqa: F401
    except ImportError:
        log_callback(
            "PDF-Export: pywin32 nicht verfügbar. Bitte in der aktiven Umgebung installieren: "
            "pip install pywin32 && pywin32_postinstall.py -install",
            "WARN"
        )
        result["error"] = "missing_pywin32"
        return result

    docx_paths = []
    for root, _, files in os.walk(ordner):
        for name in files:
            if not name.lower().endswith('.docx') or name.startswith('~'):
                continue
            docx_paths.append(os.path.join(root, name))

    result["docx_count"] = len(docx_paths)
    if not docx_paths:
        if progress_callback:
            progress_callback(0, 0)
        log_callback("Keine Word-Dokumente zur PDF-Konvertierung gefunden.", "INFO")
        return result

    def _verarbeite_einzel(docx_path: str, name: str):
        if file_callback:
            file_callback(name)
        t0 = time.perf_counter()
        try:
            ok, error = _docx_to_pdf_word_com_single(docx_path)
            dt_s = time.perf_counter() - t0
            if ok:
                result["success_count"] += 1
                log_callback(f"  -> PDF erstellt: {name}", "SUCCESS")
                log_callback(f"  -> Laufzeit PDF-Konvertierung '{name}': {dt_s:.2f}s", "INFO")
            else:
                result["failure_count"] += 1
                log_callback(f"  -> PDF fehlgeschlagen für '{name}': {error}", "ERROR")
                log_callback(f"  -> Laufzeit PDF-Konvertierung '{name}': {dt_s:.2f}s", "WARN")
        except Exception as e:
            dt_s = time.perf_counter() - t0
            result["failure_count"] += 1
            log_callback(f"  -> PDF fehlgeschlagen für '{name}': {e}", "ERROR")
            log_callback(f"  -> Laufzeit PDF-Konvertierung '{name}': {dt_s:.2f}s", "WARN")

    use_parallel = os.name == "nt" and len(docx_paths) > 1
    done = 0
    ntot = result["docx_count"]
    if use_parallel:
        try:
            log_callback(
                f"  → PDF-Konvertierung: {ntot} Datei(en), bis zu {_max_word_com_workers()} parallele Word-Aufgaben",
                "INFO",
            )
            with ProcessPoolExecutor(max_workers=_max_word_com_workers()) as ex:
                future_map = {ex.submit(_docx_to_pdf_com_mp_wrapper, p): p for p in docx_paths}
                started_at = {p: time.perf_counter() for p in docx_paths}
                for fut in as_completed(future_map):
                    docx_path, ok, serr = fut.result()
                    name = os.path.basename(docx_path)
                    dt_s = time.perf_counter() - started_at.get(docx_path, time.perf_counter())
                    if file_callback:
                        file_callback(name)
                    if ok:
                        result["success_count"] += 1
                        log_callback(f"  -> PDF erstellt: {name}", "SUCCESS")
                        log_callback(f"  -> Laufzeit PDF-Konvertierung '{name}': {dt_s:.2f}s", "INFO")
                    else:
                        result["failure_count"] += 1
                        log_callback(f"  -> PDF fehlgeschlagen für '{name}': {serr}", "ERROR")
                        log_callback(f"  -> Laufzeit PDF-Konvertierung '{name}': {dt_s:.2f}s", "WARN")
                    done += 1
                    if progress_callback:
                        progress_callback(done, ntot)
        except Exception as e:
            log_callback(f"  → Parallele PDF-Ausführung fehlgeschlagen ({e}), fahre nacheinander fort.", "WARN")
            result["success_count"] = 0
            result["failure_count"] = 0
            done = 0
            for index, docx_path in enumerate(docx_paths, start=1):
                name = os.path.basename(docx_path)
                _verarbeite_einzel(docx_path, name)
                if progress_callback:
                    progress_callback(index, ntot)
    else:
        for index, docx_path in enumerate(docx_paths, start=1):
            name = os.path.basename(docx_path)
            _verarbeite_einzel(docx_path, name)
            if progress_callback:
                progress_callback(index, ntot)

    if result["success_count"]:
        log_callback(f"PDF-Erstellung abgeschlossen: {result['success_count']} Datei(en).", "SUCCESS")
    if result["failure_count"]:
        log_callback(
            f"PDF-Erstellung mit Fehlern: {result['failure_count']} von {result['docx_count']} Datei(en) fehlgeschlagen.",
            "WARN"
        )
    return result


def verarbeite_vorlagen_trockenlauf(
    vorlagen_ordner,
    excel_path,
    log_callback,
    header_row,
    bilder_ordner=None,
    selected_template_paths=None,
    rules_enabled=False,
    signage_rules=None,
    categories=None,
    reuse_lageplan_from_last_export=False,
    previous_export_root=None,
    export_as_pdf=False,
):
    """
    Führt eine "Trockenlauf"-Validierung durch, ohne Dokumente zu erstellen.
    Prüft, ob alle referenzierten Bilder und Platzhalter vorhanden sind.

    Nur Vorlagen unter anlagen/ und allgemein/ (wie bei der Generierung).
    Mit selected_template_paths wird auf eine Teilmenge eingeschränkt.
    """
    def _log(msg, level="INFO"):
        _log_handler(msg, level, log_callback)

    _log("=" * 60 + "\n", "SEP")
    _log("Starte Trockenlauf (Konfigurationsprüfung)...", "INFO")
    _log("=" * 60 + "\n", "SEP")

    is_ok = True

    try:
        # 1. Excel-Daten laden
        datensaetze, fallback_marken = lade_excel_daten(excel_path, header_row, log_callback)
        if not datensaetze:
            _log("Keine Datensätze in Excel gefunden. Abbruch.", "WARN")
            return False

        excel_columns = set(datensaetze[0].keys()) | set(fallback_marken.keys())
        _log(f"Gefundene Spalten/Marken in Excel: {len(excel_columns)}", "INFO")
        _log(f"Gefundene Fallback-Textmarken (Blatt 2): {sorted(list(fallback_marken.keys()))}", "INFO")

        # 2. Vorlagen sammeln (nur Anlagen/Allgemein, optional gefiltert)
        anlagen_raw, allg_raw = sammle_vorlagen_pfade(vorlagen_ordner)
        anlagen_t, allg_t = filter_vorlagen_nach_auswahl(anlagen_raw, allg_raw, selected_template_paths)
        all_templates = anlagen_t + allg_t
        _log(
            f"Gefundene Word-Vorlagen (Anlagen/Allgemein): {len(all_templates)} "
            f"(Anlagen: {len(anlagen_t)}, Allgemein: {len(allg_t)})",
            "INFO",
        )
        if export_as_pdf:
            _log(
                "PDF-Erzeugung ist aktiviert: Nach der DOCX-Erstellung würden die erzeugten Dokumente "
                "zusätzlich in PDF umgewandelt.",
                "INFO",
            )
        if not all_templates:
            _log("Keine Vorlagen für den Trockenlauf ausgewählt oder vorhanden. Abbruch.", "WARN")
            return False

        # 3. Vorlagen und Bilder validieren
        _log("\n--- Validierung der Vorlagen ---\n", "SEP")
        all_template_vars = set()
        # 'datetime_utc' wird automatisch hinzugefügt und ist daher kein Fehler, wenn es in Excel fehlt.
        valid_placeholders = excel_columns | {'datetime_utc'}

        for template_path in all_templates:
            try:
                doc = DocxTemplate(template_path)
                template_vars = doc.get_undeclared_template_variables()
                all_template_vars.update(template_vars)

                missing_vars = template_vars - valid_placeholders
                if missing_vars:
                    is_ok = False
                    _log(
                        f"Vorlage '{os.path.basename(template_path)}': Fehlende Platzhalter in Excel: {missing_vars}",
                        "ERROR"
                    )
                else:
                    _log(f"Vorlage '{os.path.basename(template_path)}': OK", "SUCCESS")

                # NEU: Für _img, _qr, _link Platzhalter prüfe nur in Excel/Fallback auf _size
                img_qr_keys = {k for k in template_vars if k.endswith(('_img', '_qr', '_link'))}
                if img_qr_keys:
                    _log(f"Platzhalter-Details für '{os.path.basename(template_path)}':", "INFO")
                for key in img_qr_keys:
                    size_key_lower = f"{key}_size"
                    size_key_camel = f"{key}_Size"
                    # Wert und Größe suchen (Excel/Fallback)
                    value = None
                    size = None
                    # Suche Wert in allen Datensätzen und Fallback
                    for eintrag in datensaetze + [fallback_marken]:
                        if key in eintrag and eintrag[key]:
                            value = eintrag[key]
                        if size_key_lower in eintrag and eintrag[size_key_lower]:
                            size = eintrag[size_key_lower]
                        elif size_key_camel in eintrag and eintrag[size_key_camel]:
                            size = eintrag[size_key_camel]
                    # Standardgrößen
                    if key.endswith('_img'):
                        default_size = 15
                        typ = 'Bild'
                    else:
                        default_size = 4
                        typ = 'QR/Link'
                    used_size = size if size is not None else default_size
                    value_str = value if value is not None else '[kein Wert in Excel/Fallback]'
                    _log(f"  - {typ}: '{key}' → Wert: {value_str}, Größe: {used_size} cm", "INFO")
                    if size is None:
                        _log(
                            f"Hinweis: Für Platzhalter '{key}' wird die Standardgröße verwendet, da kein '{size_key_lower}' oder '{size_key_camel}' in Excel/Fallback gefunden wurde.",
                            "INFO"
                        )

            except Exception as e:
                is_ok = False
                _log(
                    f"Vorlage '{os.path.basename(template_path)}': Konnte nicht gelesen werden. Fehler: {e}",
                    "FATAL"
                )

        _log("\n--- Validierung der Bilder (aus Excel-Daten) ---\n", "SEP")
        if bilder_ordner:
            image_references = set()
            for eintrag in datensaetze + [fallback_marken]:
                for key, value in eintrag.items():
                    if isinstance(value, str) and (value.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.svg'))):
                        image_references.add(value)

            _log(f"Gefundene Bild-Referenzen in Excel: {len(image_references)}", "INFO")
            for img_name in image_references:
                img_path = os.path.join(bilder_ordner, img_name)
                if not os.path.exists(img_path):
                    is_ok = False
                    _log(f"Bild '<strong>{img_name}</strong>' nicht gefunden", "WARN")
            if is_ok and image_references:
                _log("Alle referenzierten Bilder wurden gefunden.", "SUCCESS")

        else:
            _log("Kein Bilder-Ordner angegeben, Prüfung übersprungen.", "WARN")

        if rules_enabled and signage_rules:
            naive_anlagen = len(datensaetze) * len(anlagen_t)
            eff_anlagen = count_anlagen_emits_with_rules(
                datensaetze,
                anlagen_t,
                vorlagen_ordner,
                fallback_marken,
                bilder_ordner,
                True,
                signage_rules,
            )
            _log("\n--- Regeln (optionale Schilder) ---\n", "SEP")
            _log(
                f"Aktiv: {eff_anlagen} von {naive_anlagen} möglichen Anlagen-Dokumenten würden erzeugt "
                f"({naive_anlagen - eff_anlagen} übersprungen).",
                "INFO",
            )
            _log("\n--- Regel-Entscheidungen (Vorschau) ---\n", "SEP")
            for eintrag in datensaetze:
                for tp in anlagen_t:
                    decision = evaluate_anlage_template_decision(
                        vorlagen_ordner,
                        tp,
                        eintrag,
                        fallback_marken,
                        bilder_ordner,
                        True,
                        signage_rules,
                    )
                    if not decision.get("has_rule"):
                        continue
                    lvl = "INFO" if decision.get("emit") else "WARN"
                    _log(_format_rule_decision_log(decision, eintrag, tp), lvl)

        if reuse_lageplan_from_last_export:
            cats = categories or {}
            copy_n = 0
            gen_n = 0
            prev_ok = bool(previous_export_root and os.path.isdir(previous_export_root))
            for eintrag in datensaetze:
                for tp in anlagen_t:
                    decision = evaluate_anlage_template_decision(
                        vorlagen_ordner,
                        tp,
                        eintrag,
                        fallback_marken,
                        bilder_ordner,
                        rules_enabled,
                        signage_rules,
                    )
                    if not decision.get("emit"):
                        continue
                    if not os.path.basename(tp).lower().startswith('rl_'):
                        continue
                    if prev_ok and find_lageplan_source_file(previous_export_root, cats, tp, eintrag):
                        copy_n += 1
                    else:
                        gen_n += 1
            _log("\n--- Lageplan-Übernahme (Vorschau) ---\n", "SEP")
            if not prev_ok:
                _log(
                    "Kein gültiger letzter Export-Ordner: alle rl_-Lagepläne würden wie gewohnt neu erstellt.",
                    "INFO",
                )
            _log(
                f"rl_-Vorlagen: {copy_n} würden aus dem letzten Export kopiert, "
                f"{gen_n} neu aus Vorlage erzeugt.",
                "INFO",
            )

        _log("\n" + "=" * 60, "SEP")
        if is_ok:
            _log("Trockenlauf erfolgreich! Konfiguration scheint gültig.", "SUCCESS")
        else:
            _log("Trockenlauf mit Fehlern beendet. Bitte prüfen Sie die obigen Meldungen.", "ERROR")

        return is_ok

    except Exception as e:
        _log_handler(f"FATALER FEHLER im Trockenlauf: {e}\n{traceback.format_exc()}", "FATAL", log_callback)
        return False


def _fuehre_dokument_gen_jobs(
    jobs,
    parallel_doc_generation: bool,
    progress_callback,
    file_callback,
    overall_total_steps,
    worker_thread,
    svg_png_map,
    haupt_export_ordner,
    categories,
    datetime_utc_format,
    bilder_ordner,
    fallback_marken,
    seitenpruefung_pending,
    seitenpruefung_lock,
    is_dark_mode,
    _log,
):
    """
    Führt die gesammelten Anlagen-/Allgemein-Jobs (docxtpl oder Lageplan-Kopie) aus.
    parallel_doc_generation: mehrere ThreadPool-Worker; sonst nacheinander.
    """
    completed_steps = 0
    completed_lock = threading.Lock()

    def _mark_step_done():
        nonlocal completed_steps
        with completed_lock:
            completed_steps += 1
            progress_callback(completed_steps, overall_total_steps)

    def _eintrag_kopie(eintrag):
        if eintrag is None:
            return {}
        if isinstance(eintrag, dict):
            return {k: eintrag[k] for k in eintrag}
        try:
            return dict(eintrag)
        except Exception:
            return eintrag

    def _arbeite_lageplan(job):
        if worker_thread and worker_thread.isInterruptionRequested():
            return
        t0 = time.perf_counter()
        if not job.get("src") or not os.path.isfile(job["src"]):
            _log("Lageplan-Quelldatei fehlt, neu aus Vorlage.", "WARN")
            rj = {
                "type": "render",
                "ix": job["ix"],
                "template": job["vorlage_pfad"],
                "data": _eintrag_kopie(job.get("eintrag")),
                "is_anlage": True,
            }
            _arbeite_render(rj)
            return
        try:
            os.makedirs(job["export_kat"], exist_ok=True)
            ziel = os.path.join(job["export_kat"], job["ziel_name"])
            shutil.copy2(job["src"], ziel)
        except OSError as e:
            _log(
                f"Lageplan kopieren fehlgeschlagen ({e}), neu aus Vorlage.",
                "WARN",
            )
            rj = {
                "type": "render",
                "ix": job["ix"],
                "template": job["vorlage_pfad"],
                "data": _eintrag_kopie(job.get("eintrag")),
                "is_anlage": True,
            }
            _arbeite_render(rj)
            return
        dt_s = time.perf_counter() - t0
        try:
            file_callback(job["vn"])
            _log(
                f"Lageplan übernommen (letzter Export): '{job['ziel_name']}'",
                "SUCCESS",
            )
            _log(f"Laufzeit Dokument-Job '{job['ziel_name']}' (Lageplan-Kopie): {dt_s:.2f}s", "INFO")
            ziel_lage = os.path.join(job["export_kat"], job["ziel_name"])
            if seitenpruefung_lock is not None:
                with seitenpruefung_lock:
                    seitenpruefung_pending.append((job["vorlage_pfad"], ziel_lage, job["ziel_name"]))
            else:
                seitenpruefung_pending.append((job["vorlage_pfad"], ziel_lage, job["ziel_name"]))
        finally:
            _mark_step_done()

    def _arbeite_render(job):
        if worker_thread and worker_thread.isInterruptionRequested():
            return
        t0 = time.perf_counter()
        file_callback(os.path.basename(job["template"]))
        try:
            data = _eintrag_kopie(job.get("data"))
            context = data.copy()
            for key, fallback_value in fallback_marken.items():
                primary_value = context.get(key)
                is_empty = pd.isna(primary_value) or (isinstance(primary_value, str) and not primary_value.strip())
                if is_empty:
                    context[key] = fallback_value
            try:
                context["datetime_utc"] = datetime.now().strftime("%Y-%m-%d")
            except Exception as e:
                context["datetime_utc"] = f"[Format-Fehler: {e}]"
                _log(f"Ungültiges Zeitstempel-Format: '{datetime_utc_format}'. Fehler: {e}", "WARN")
            context["_bilder_ordner"] = bilder_ordner
            doc = ersetze_platzhalter_mit_docxtpl(job["template"], context, svg_png_map, _log)
            _speichere_dokument(
                doc,
                job["template"],
                data,
                job["is_anlage"],
                haupt_export_ordner,
                categories,
                _log,
                is_dark_mode,
                seitenpruefung_pending=seitenpruefung_pending,
                seitenpruefung_lock=seitenpruefung_lock,
            )
            dt_s = time.perf_counter() - t0
            _log(
                f"Laufzeit Dokument-Job '{os.path.basename(job['template'])}': {dt_s:.2f}s",
                "INFO",
            )
        except TemplateSyntaxError as e:
            dt_s = time.perf_counter() - t0
            _log(
                f"FEHLER in Vorlage '{os.path.basename(job['template'])}': Ungültige Syntax. Bitte prüfen Sie die Platzhalter.",
                "FATAL",
            )
            _log(f"Details: {e}", "ERROR")
            _log(
                f"Laufzeit Dokument-Job '{os.path.basename(job['template'])}' bis Fehler: {dt_s:.2f}s",
                "WARN",
            )
        except Exception as e:
            dt_s = time.perf_counter() - t0
            _log(f"FEHLER bei '{os.path.basename(job['template'])}': {e}", "ERROR")
            _log(
                f"Laufzeit Dokument-Job '{os.path.basename(job['template'])}' bis Fehler: {dt_s:.2f}s",
                "WARN",
            )
        finally:
            _mark_step_done()

    def _arbeite(job):
        t = job.get("type")
        if t == "lageplan_copy":
            _arbeite_lageplan(job)
        else:
            _arbeite_render(job)

    if not parallel_doc_generation or len(jobs) <= 1:
        for j in jobs:
            if worker_thread and worker_thread.isInterruptionRequested():
                break
            _arbeite(j)
    else:
        max_w = _max_docgen_thread_workers()
        with ThreadPoolExecutor(max_workers=max_w) as ex:
            _log(
                f"Dokumentenerzeugung: parallel mit bis zu {max_w} Threads (docxtpl / Lageplan).",
                "INFO",
            )
            list(ex.map(_arbeite, jobs))


def verarbeite_vorlagen_preview(
    vorlagen_ordner,
    export_ordner,
    excel_path,
    log_callback,
    progress_callback,
    file_callback,
    header_row,
    svg_scale,
    png_compression,
    categories,
    datetime_utc_format,
    bilder_ordner=None,
    worker_thread=None,
    is_dark_mode=False,
    projekt_name_override=None,
    export_as_pdf=False,
    selected_template_paths=None,
    rules_enabled=False,
    signage_rules=None,
    reuse_lageplan_from_last_export=False,
    previous_export_root=None,
    preview_template_abs=None,
    parallel_doc_generation=False,
):
    """
    Erzeugt eine Stichprobe wie im echten Lauf, aber nur mit der ersten Excel-Datenzeile
    für Anlagen-Vorlagen; Allgemein wie in Produktion (erster Datensatz + Fallback).
    """
    def _log(msg, level="INFO"):
        _log_handler(msg, level, log_callback)

    _log("=" * 60 + "\n", "SEP")
    _log("Starte Dokument-Vorschau (erste Excel-Datenzeile)...", "INFO")
    _log(
        "Hinweis: Vorschau ≠ rechtlicher Prüfstand; Layout kann bei anderen Daten abweichen.",
        "INFO",
    )
    _log("=" * 60 + "\n", "SEP")
    temp_dir = None
    try:
        datensaetze, fallback_marken = lade_excel_daten(excel_path, header_row, log_callback)
        if not datensaetze:
            _log("Keine Datensätze gefunden. Abbruch.", "WARN")
            return None

        eintrag_preview = datensaetze[0]
        sn = str(eintrag_preview.get("anlage_seriennummer", "")).strip()
        _log(f"Vorschau-Datensatz: erste Zeile (anlage_seriennummer: {sn!r}).", "INFO")

        anlagen_raw, allg_raw = sammle_vorlagen_pfade(vorlagen_ordner)
        anlagen_templates, allgemein_templates = filter_vorlagen_nach_auswahl(
            anlagen_raw, allg_raw, selected_template_paths
        )
        if preview_template_abs:
            target = _norm_template_path(preview_template_abs)
            anlagen_templates = [p for p in anlagen_templates if _norm_template_path(p) == target]
            allgemein_templates = [p for p in allgemein_templates if _norm_template_path(p) == target]
            if not anlagen_templates and not allgemein_templates:
                _log(
                    f"Vorschau: Vorlage nicht in der aktuellen Auswahl oder nicht unter anlagen/ bzw. allgemein/: "
                    f"{os.path.basename(preview_template_abs)}",
                    "WARN",
                )
                return None

        anlagen_emits = count_anlagen_emits_with_rules(
            datensaetze[:1],
            anlagen_templates,
            vorlagen_ordner,
            fallback_marken,
            bilder_ordner,
            rules_enabled,
            signage_rules,
        )
        generation_docs_total = anlagen_emits + len(allgemein_templates)
        if generation_docs_total == 0:
            _log(
                "Vorschau: Keine Dokumente zu erzeugen (0 erwartete Ausgaben nach Regeln / ohne Allgemein). "
                "Es werden trotzdem Entscheidungen protokolliert.",
                "WARN",
            )
        pdf_docs_total = generation_docs_total if export_as_pdf else 0
        overall_total_steps = generation_docs_total + pdf_docs_total
        if overall_total_steps == 0:
            overall_total_steps = 1

        temp_dir = tempfile.mkdtemp(prefix="svg_cache_")
        unique_svg_paths = set()
        if bilder_ordner:
            for eintrag in datensaetze + [fallback_marken]:
                for key, value in eintrag.items():
                    if isinstance(value, str) and value.lower().endswith('.svg'):
                        full_path = os.path.join(bilder_ordner, value)
                        if os.path.exists(full_path):
                            unique_svg_paths.add(full_path)
        svg_png_map = {}
        if unique_svg_paths:
            _log(f"Konvertiere {len(unique_svg_paths)} einzigartige SVGs...", "INFO")
            for svg_path in unique_svg_paths:
                png_filename = os.path.splitext(os.path.basename(svg_path))[0] + ".png"
                temp_png_path = os.path.join(temp_dir, png_filename)
                if svg_to_png_file_pyside(svg_path, temp_png_path, _log, svg_scale, png_compression):
                    svg_png_map[svg_path] = temp_png_path
            _log("SVG-Konvertierung abgeschlossen.", "SUCCESS")

        if projekt_name_override is not None and str(projekt_name_override).strip():
            raw = str(projekt_name_override).strip()
            projekt_name = "".join(c if c not in r'\/:*?"<>|' else "_" for c in raw) or "Unbenanntes_Projekt"
        else:
            projekt_name = str(
                datensaetze[0].get('projekt_name', fallback_marken.get('projekt_name', 'Unbenanntes_Projekt'))
            ).strip()
        haupt_export_ordner = os.path.join(
            export_ordner,
            datetime.now().strftime('%Y-%m-%d_%H%M%S') + "_" + projekt_name + "_Vorschau",
        )

        seitenpruefung_pending = []
        seitenpruefung_lock = threading.Lock() if parallel_doc_generation else None
        jobs = []
        job_ix = 0
        warned_lageplan_no_export = False

        def _eintrag_shallow(eintrag):
            if isinstance(eintrag, dict):
                return {k: eintrag[k] for k in eintrag}
            try:
                return dict(eintrag)
            except Exception:
                return eintrag

        _log("\n--- Vorschau 'Anlagen' (nur erste Zeile) ---\n", "SEP")
        for vorlage_pfad in anlagen_templates:
            if worker_thread and worker_thread.isInterruptionRequested():
                break
            decision = evaluate_anlage_template_decision(
                vorlagen_ordner,
                vorlage_pfad,
                eintrag_preview,
                fallback_marken,
                bilder_ordner,
                rules_enabled,
                signage_rules,
            )
            if decision.get("has_rule"):
                detail = _format_rule_decision_log(decision, eintrag_preview, vorlage_pfad)
                if decision.get("emit"):
                    _log(detail, "INFO")
                else:
                    _log("Vorschau übersprungen: " + detail, "WARN")
            elif not decision.get("emit"):
                _log(
                    f"Vorschau übersprungen für '{os.path.basename(vorlage_pfad)}' (keine Ausgabe).",
                    "WARN",
                )
            if not decision.get("emit"):
                continue
            vn = os.path.basename(vorlage_pfad)
            if reuse_lageplan_from_last_export and vn.lower().startswith('rl_'):
                if not previous_export_root or not os.path.isdir(previous_export_root):
                    if not warned_lageplan_no_export:
                        _log(
                            "Lageplan-Übernahme: Kein gültiger letzter Export-Ordner – "
                            "rl_-Dateien werden wie gewohnt neu erstellt.",
                            "INFO",
                        )
                        warned_lageplan_no_export = True
                    job_ix += 1
                    jobs.append(
                        {
                            "type": "render",
                            "ix": job_ix,
                            "template": vorlage_pfad,
                            "data": _eintrag_shallow(eintrag_preview),
                            "is_anlage": True,
                        }
                    )
                    continue
                src = find_lageplan_source_file(
                    previous_export_root, categories, vorlage_pfad, eintrag_preview
                )
                export_kat, ziel_name = _lageplan_ziel_pfad(
                    vorlage_pfad, eintrag_preview, haupt_export_ordner, categories
                )
                if src:
                    job_ix += 1
                    jobs.append(
                        {
                            "type": "lageplan_copy",
                            "ix": job_ix,
                            "export_kat": export_kat,
                            "ziel_name": ziel_name,
                            "src": src,
                            "vorlage_pfad": vorlage_pfad,
                            "vn": vn,
                            "eintrag": _eintrag_shallow(eintrag_preview),
                        }
                    )
                    continue
                _log(
                    f"Keine vorherige Datei für '{ziel_name}' im letzten Export – "
                    "Erstellung wie gewohnt aus Vorlage.",
                    "INFO",
                )
                job_ix += 1
                jobs.append(
                    {
                        "type": "render",
                        "ix": job_ix,
                        "template": vorlage_pfad,
                        "data": _eintrag_shallow(eintrag_preview),
                        "is_anlage": True,
                    }
                )
                continue

            job_ix += 1
            jobs.append(
                {
                    "type": "render",
                    "ix": job_ix,
                    "template": vorlage_pfad,
                    "data": _eintrag_shallow(eintrag_preview),
                    "is_anlage": True,
                }
            )

        if worker_thread and not worker_thread.isInterruptionRequested():
            _log("\n--- Vorschau 'Allgemein' ---\n", "SEP")
            if allgemein_templates:
                for vorlage_pfad in allgemein_templates:
                    job_ix += 1
                    jobs.append(
                        {
                            "type": "render",
                            "ix": job_ix,
                            "template": vorlage_pfad,
                            "data": _eintrag_shallow(datensaetze[0]),
                            "is_anlage": False,
                        }
                    )

        if len(jobs) != generation_docs_total:
            _log(
                f"Hinweis: Vorschau-Jobs ({len(jobs)}) vs. gezählte erwartete Dokumente ({generation_docs_total}) - ggf. Log prüfen.",
                "WARN",
            )

        _fuehre_dokument_gen_jobs(
            jobs,
            parallel_doc_generation,
            progress_callback,
            file_callback,
            overall_total_steps,
            worker_thread,
            svg_png_map,
            haupt_export_ordner,
            categories,
            datetime_utc_format,
            bilder_ordner,
            fallback_marken,
            seitenpruefung_pending,
            seitenpruefung_lock,
            is_dark_mode,
            _log,
        )

        seitenpruefung_warnungen = _fuehre_seitenpruefung_batched(
            seitenpruefung_pending, _log, is_dark_mode
        )

        docx_erzeugt = False
        if os.path.isdir(haupt_export_ordner):
            for root, _, files in os.walk(haupt_export_ordner):
                for name in files:
                    if name.lower().endswith('.docx') and not name.startswith('~'):
                        docx_erzeugt = True
                        break
                if docx_erzeugt:
                    break

        if worker_thread and worker_thread.isInterruptionRequested():
            _log("\n" + "=" * 60, "SEP")
            _log("Vorschau vom Benutzer abgebrochen.", "WARN")
            return None

        if not docx_erzeugt:
            _log(
                "Vorschau: Keine DOCX-Dateien erzeugt (alle Anlagen-Vorlagen übersprungen oder fehlerhaft).",
                "WARN",
            )
            return None

        _log("\n" + "=" * 60, "SEP")
        if seitenpruefung_warnungen:
            _log("--- Seitenprüfung: Zusammenfassung (Vorschau) ---", "SEP")
            _log(
                f"Bei {len(seitenpruefung_warnungen)} erzeugter Datei(en) weist das Dokument "
                f"mehr Seiten auf als die zugehörige Vorlage (Details stehen im Log oben, Prüfung per Word, Windows).",
                "WARN",
            )
        _log("Vorschau-Dokumente (DOCX) erstellt.", "SUCCESS")
        if export_as_pdf and haupt_export_ordner and os.path.isdir(haupt_export_ordner):
            _log("\n--- PDF-Erstellung (Vorschau) ---\n", "SEP")

            def _pdf_progress(pdf_current, pdf_total):
                progress_callback(generation_docs_total + pdf_current, overall_total_steps)

            def _pdf_file(filename):
                file_callback(f"PDF: {filename}")

            pdf_result = _convert_docx_to_pdf_in_folder(
                haupt_export_ordner,
                _log,
                progress_callback=_pdf_progress,
                file_callback=_pdf_file,
            )
            if pdf_result.get("error") == "missing_pywin32":
                _log(
                    "PDF-Vorschau nicht möglich: Microsoft Word und pywin32 werden für den Export benötigt.",
                    "WARN",
                )
            elif pdf_result.get("failure_count"):
                _log(
                    "Vorschau-PDF: Mindestens eine Konvertierung ist fehlgeschlagen (Word installiert / Datei nicht gesperrt?).",
                    "WARN",
                )

        return haupt_export_ordner
    except Exception as e:
        _log_handler(f"FATALER FEHLER (Vorschau): {e}\n{traceback.format_exc()}", "FATAL", log_callback)
        return None
    finally:
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
                _log_handler("Temporäres Cache-Verzeichnis gelöscht.", "INFO", log_callback)
            except Exception as e:
                _log_handler(f"Cache konnte nicht gelöscht werden: {e}", "ERROR", log_callback)


def verarbeite_vorlagen(
    vorlagen_ordner,
    export_ordner,
    excel_path,
    log_callback,
    progress_callback,
    file_callback,
    header_row,
    svg_scale,
    png_compression,
    categories,
    datetime_utc_format,
    bilder_ordner=None,
    worker_thread=None,
    dry_run=False,
    is_dark_mode=False,
    projekt_name_override=None,
    export_as_pdf=False,
    selected_template_paths=None,
    rules_enabled=False,
    signage_rules=None,
    reuse_lageplan_from_last_export=False,
    previous_export_root=None,
    parallel_doc_generation=False,
):
    """
    Die zentrale Orchestrierungsfunktion, die den gesamten Prozess steuert.

    Ablauf:
    1.  Lädt die Daten aus der Excel-Datei (Hauptdaten und Fallback-Marken).
    2.  Sucht nach allen SVG-Bildern, die in den Daten referenziert werden.
    3.  Konvertiert alle gefundenen, einzigartigen SVGs in PNGs und speichert sie
        in einem temporären Ordner, um die Performance zu verbessern (jedes SVG
        wird nur einmal konvertiert).
    4.  Durchsucht den Vorlagen-Ordner und teilt die gefundenen .docx-Dateien in
        'anlagen'-spezifische und 'allgemein'e Vorlagen auf.
    5.  Erstellt einen eindeutigen Export-Ordner für den aktuellen Durchlauf,
        benannt mit Zeitstempel und Projektname.
    6.  Verarbeitet die 'Anlagen'-Vorlagen: Iteriert über jeden Datensatz aus der
        Excel-Datei und füllt für jeden Datensatz jede Vorlage aus.
    7.  Verarbeitet die 'Allgemein'-Vorlagen: Füllt diese Vorlagen einmal aus,
        wobei der erste Datensatz und die Fallback-Marken als Datenquelle dienen.
    8.  Verwendet bei der Erstellung eine Fallback-Logik: Wenn ein Wert in einem
        Haupt-Datensatz fehlt, wird der entsprechende Wert aus den Fallback-Marken
        (zweites Excel-Blatt) verwendet.
    9.  Löscht nach Abschluss das temporäre Verzeichnis mit den PNGs.
    """
    if dry_run:
        return verarbeite_vorlagen_trockenlauf(
            vorlagen_ordner,
            excel_path,
            log_callback,
            header_row,
            bilder_ordner,
            selected_template_paths=selected_template_paths,
            rules_enabled=rules_enabled,
            signage_rules=signage_rules,
            categories=categories,
            reuse_lageplan_from_last_export=reuse_lageplan_from_last_export,
            previous_export_root=previous_export_root,
            export_as_pdf=export_as_pdf,
        )

    def _log(msg, level="INFO"):
        _log_handler(msg, level, log_callback)

    _log("=" * 60 + "\n", "SEP")
    _log("Starte Dokumentenerstellung...", "INFO")
    _log("=" * 60 + "\n", "SEP")
    temp_dir = None
    try:
        datensaetze, fallback_marken = lade_excel_daten(excel_path, header_row, log_callback)
        if not datensaetze:
            _log("Keine Datensätze gefunden. Abbruch.", "WARN")
            return

        anlagen_raw, allg_raw = sammle_vorlagen_pfade(vorlagen_ordner)
        anlagen_templates, allgemein_templates = filter_vorlagen_nach_auswahl(
            anlagen_raw, allg_raw, selected_template_paths
        )
        anlagen_emits = count_anlagen_emits_with_rules(
            datensaetze,
            anlagen_templates,
            vorlagen_ordner,
            fallback_marken,
            bilder_ordner,
            rules_enabled,
            signage_rules,
        )
        generation_docs_total = anlagen_emits + len(allgemein_templates)
        if generation_docs_total == 0:
            _log(
                "Keine Vorlagen zum Erzeugen oder keine Auswahl: Abbruch (0 Dokumente).",
                "WARN",
            )
            return None
        pdf_docs_total = generation_docs_total if export_as_pdf else 0
        overall_total_steps = generation_docs_total + pdf_docs_total

        temp_dir = tempfile.mkdtemp(prefix="svg_cache_")
        unique_svg_paths = set()
        if bilder_ordner:
            for eintrag in datensaetze + [fallback_marken]:
                for key, value in eintrag.items():
                    if isinstance(value, str) and value.lower().endswith('.svg'):
                        full_path = os.path.join(bilder_ordner, value)
                        if os.path.exists(full_path):
                            unique_svg_paths.add(full_path)
        svg_png_map = {}
        if unique_svg_paths:
            _log(f"Konvertiere {len(unique_svg_paths)} einzigartige SVGs...", "INFO")
            for svg_path in unique_svg_paths:
                png_filename = os.path.splitext(os.path.basename(svg_path))[0] + ".png"
                temp_png_path = os.path.join(temp_dir, png_filename)
                if svg_to_png_file_pyside(svg_path, temp_png_path, _log, svg_scale, png_compression):
                    svg_png_map[svg_path] = temp_png_path
            _log("SVG-Konvertierung abgeschlossen.", "SUCCESS")

        if projekt_name_override is not None and str(projekt_name_override).strip():
            raw = str(projekt_name_override).strip()
            projekt_name = "".join(c if c not in r'\/:*?"<>|' else "_" for c in raw) or "Unbenanntes_Projekt"
        else:
            projekt_name = str(
                datensaetze[0].get('projekt_name', fallback_marken.get('projekt_name', 'Unbenanntes_Projekt'))
            ).strip()
        haupt_export_ordner = os.path.join(
            export_ordner,
            datetime.now().strftime('%Y-%m-%d_%H%M%S') + "_" + projekt_name
        )

        seitenpruefung_pending = []
        seitenpruefung_lock = threading.Lock() if parallel_doc_generation else None
        jobs = []
        job_ix = 0
        warned_lageplan_no_export = False

        def _eintrag_shallow(eintrag):
            if isinstance(eintrag, dict):
                return {k: eintrag[k] for k in eintrag}
            try:
                return dict(eintrag)
            except Exception:
                return eintrag

        _log("\n--- Verarbeitung 'Anlagen' ---\n", "SEP")
        for eintrag in datensaetze:
            if worker_thread and worker_thread.isInterruptionRequested():
                break
            for vorlage_pfad in anlagen_templates:
                decision = evaluate_anlage_template_decision(
                    vorlagen_ordner,
                    vorlage_pfad,
                    eintrag,
                    fallback_marken,
                    bilder_ordner,
                    rules_enabled,
                    signage_rules,
                )
                if decision.get("has_rule"):
                    lvl = "INFO" if decision.get("emit") else "WARN"
                    _log(_format_rule_decision_log(decision, eintrag, vorlage_pfad), lvl)
                if not decision.get("emit"):
                    continue
                vn = os.path.basename(vorlage_pfad)
                if reuse_lageplan_from_last_export and vn.lower().startswith('rl_'):
                    if not previous_export_root or not os.path.isdir(previous_export_root):
                        if not warned_lageplan_no_export:
                            _log(
                                "Lageplan-Übernahme: Kein gültiger letzter Export-Ordner – "
                                "rl_-Dateien werden wie gewohnt neu erstellt.",
                                "INFO",
                            )
                            warned_lageplan_no_export = True
                        job_ix += 1
                        jobs.append(
                            {
                                "type": "render",
                                "ix": job_ix,
                                "template": vorlage_pfad,
                                "data": _eintrag_shallow(eintrag),
                                "is_anlage": True,
                            }
                        )
                        continue
                    src = find_lageplan_source_file(
                        previous_export_root, categories, vorlage_pfad, eintrag
                    )
                    export_kat, ziel_name = _lageplan_ziel_pfad(
                        vorlage_pfad, eintrag, haupt_export_ordner, categories
                    )
                    if src:
                        job_ix += 1
                        jobs.append(
                            {
                                "type": "lageplan_copy",
                                "ix": job_ix,
                                "export_kat": export_kat,
                                "ziel_name": ziel_name,
                                "src": src,
                                "vorlage_pfad": vorlage_pfad,
                                "vn": vn,
                                "eintrag": _eintrag_shallow(eintrag),
                            }
                        )
                        continue
                    _log(
                        f"Keine vorherige Datei für '{ziel_name}' im letzten Export – "
                        "Erstellung wie gewohnt aus Vorlage.",
                        "INFO",
                    )
                    job_ix += 1
                    jobs.append(
                        {
                            "type": "render",
                            "ix": job_ix,
                            "template": vorlage_pfad,
                            "data": _eintrag_shallow(eintrag),
                            "is_anlage": True,
                        }
                    )
                    continue

                job_ix += 1
                jobs.append(
                    {
                        "type": "render",
                        "ix": job_ix,
                        "template": vorlage_pfad,
                        "data": _eintrag_shallow(eintrag),
                        "is_anlage": True,
                    }
                )

        if worker_thread and not worker_thread.isInterruptionRequested():
            _log("\n--- Verarbeitung 'Allgemein' ---\n", "SEP")
            if allgemein_templates:
                for vorlage_pfad in allgemein_templates:
                    job_ix += 1
                    jobs.append(
                        {
                            "type": "render",
                            "ix": job_ix,
                            "template": vorlage_pfad,
                            "data": _eintrag_shallow(datensaetze[0]),
                            "is_anlage": False,
                        }
                    )

        if len(jobs) != generation_docs_total:
            _log(
                f"Hinweis: Sammel-Jobs ({len(jobs)}) vs. gezählte erwartete Dokumente ({generation_docs_total}) — ggf. Log prüfen.",
                "WARN",
            )

        _fuehre_dokument_gen_jobs(
            jobs,
            parallel_doc_generation,
            progress_callback,
            file_callback,
            overall_total_steps,
            worker_thread,
            svg_png_map,
            haupt_export_ordner,
            categories,
            datetime_utc_format,
            bilder_ordner,
            fallback_marken,
            seitenpruefung_pending,
            seitenpruefung_lock,
            is_dark_mode,
            _log,
        )

        seitenpruefung_warnungen = _fuehre_seitenpruefung_batched(
            seitenpruefung_pending, _log, is_dark_mode
        )

        if worker_thread and worker_thread.isInterruptionRequested():
            _log("\n" + "=" * 60, "SEP")
            _log("Vorgang vom Benutzer abgebrochen.", "WARN")
        else:
            _log("\n" + "=" * 60, "SEP")
            if seitenpruefung_warnungen:
                _log("--- Seitenprüfung: Zusammenfassung ---", "SEP")
                _log(
                    f"Bei {len(seitenpruefung_warnungen)} erzeugter Datei(en) weist das Dokument "
                    f"mehr Seiten auf als die zugehörige Vorlage (Details stehen im Log oben, Prüfung per Word, Windows).",
                    "WARN",
                )
            _log("Alle Aufgaben abgeschlossen.", "SUCCESS")
            if export_as_pdf and haupt_export_ordner and os.path.isdir(haupt_export_ordner):
                _log("\n--- PDF-Erstellung ---\n", "SEP")
                def _pdf_progress(pdf_current, pdf_total):
                    # PDF-Schritte direkt hinter die Dokumentenerzeugung hängen.
                    progress_callback(generation_docs_total + pdf_current, overall_total_steps)

                def _pdf_file(filename):
                    file_callback(f"PDF: {filename}")

                _convert_docx_to_pdf_in_folder(
                    haupt_export_ordner,
                    _log,
                    progress_callback=_pdf_progress,
                    file_callback=_pdf_file,
                )

        return haupt_export_ordner
    except Exception as e:
        _log_handler(f"FATALER FEHLER: {e}\n{traceback.format_exc()}", "FATAL", log_callback)
        return None
    finally:
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
                _log_handler("Temporäres Cache-Verzeichnis gelöscht.", "INFO", log_callback)
            except Exception as e:
                _log_handler(f"Cache konnte nicht gelöscht werden: {e}", "ERROR", log_callback)
