"""
Bild- und QR-/Link-Größen: Übersicht aus Vorlagen + Master-Excel, Schreiben nach Blatt 2 (Fallback).
"""
from __future__ import annotations

import os
from collections import defaultdict
from typing import Any

import pandas as pd
from docxtpl import DocxTemplate
from openpyxl import load_workbook

from core.logging import _log_handler
from core.logic import (
    filter_vorlagen_nach_auswahl,
    lade_excel_daten,
    sammle_vorlagen_pfade,
)
from core.rules import merge_row_context


def liste_media_basis_platzhalter_aus_docx(doc_path: str, log_callback=None) -> list[str]:
    """Alle docxtpl-Variablen mit Suffix _img, _qr, _link (Reihenfolge: Bilder, dann QR/Link)."""
    if not doc_path or not os.path.isfile(doc_path):
        return []
    try:
        doc = DocxTemplate(doc_path)
        variables = doc.get_undeclared_template_variables()
    except Exception as e:
        _log_handler(f"Media-Platzhalter aus Vorlage: {e}", "WARN", log_callback)
        return []
    img = sorted(k for k in variables if k.endswith("_img"))
    qr = sorted(k for k in variables if k.endswith(("_qr", "_link")))
    return img + qr


def _fmt_ref(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def _default_cm_for_key(base_key: str) -> float:
    return 15.0 if base_key.endswith("_img") else 4.0


def _kind_label(base_key: str) -> str:
    return "Bild" if base_key.endswith("_img") else "QR-Link"


def _size_from_row(d: dict, base_key: str):
    sl = f"{base_key}_size"
    sc = f"{base_key}_Size"
    for k in (sl, sc):
        if k not in d:
            continue
        v = d[k]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return v
    return None


def _parse_cm(val) -> float | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def canonical_size_key(base_key: str) -> str:
    """Einheitlich …_size (kleines s) für Excel-Spalte B."""
    return f"{base_key}_size"


def resolve_size_for_base_key(
    base_key: str,
    datensaetze: list[dict],
    fallback_marken: dict,
) -> tuple[float, str, bool]:
    """
    Rückgabe: (Anzeige-cm, Quellen-Label, variable_sheet1)
    """
    default_cm = _default_cm_for_key(base_key)
    size_lower = canonical_size_key(base_key)
    size_camel = f"{base_key}_Size"

    for fk in (size_lower, size_camel):
        if fk in fallback_marken:
            cm = _parse_cm(fallback_marken[fk])
            if cm is not None:
                return cm, "Fallback (Blatt 2)", False

    cms: list[float] = []
    for d in datensaetze:
        v = _size_from_row(d, base_key)
        cm = _parse_cm(v)
        if cm is not None:
            cms.append(cm)
    if not cms:
        return default_cm, "Standard (nicht gesetzt)", False
    unique = sorted(set(cms))
    if len(unique) == 1:
        return unique[0], "Blatt 1 (einheitlich)", False
    return default_cm, "variabel – bitte in Excel prüfen", True


def resolve_reference_for_base_key(
    base_key: str,
    datensaetze: list[dict],
    fallback_marken: dict,
) -> str:
    """Wie erste Anlagenzeile + Fallback (merge_row_context)."""
    if not datensaetze:
        return _fmt_ref(fallback_marken.get(base_key))
    ctx = merge_row_context(datensaetze[0], fallback_marken)
    return _fmt_ref(ctx.get(base_key))


def sammle_media_groessen_uebersicht(
    vorlagen_ordner: str,
    excel_path: str,
    header_row: int,
    selected_template_paths: list[str] | None,
    log_callback=None,
) -> list[dict[str, Any]]:
    """
    Zeilen für die GUI: template_rel, base_key, kind, reference, size_key, cm_value,
    cm_source, cm_variable, template_abs
    """
    if not excel_path or not os.path.isfile(excel_path):
        return []
    if not vorlagen_ordner or not os.path.isdir(vorlagen_ordner):
        return []

    datensaetze, fallback_marken = lade_excel_daten(excel_path, header_row, log_callback)
    anlagen_raw, allg_raw = sammle_vorlagen_pfade(vorlagen_ordner)
    anlagen_t, allg_t = filter_vorlagen_nach_auswahl(anlagen_raw, allg_raw, selected_template_paths)
    templates = sorted(anlagen_t + allg_t, key=lambda p: (os.path.basename(p).lower(), p.lower()))

    rows: list[dict[str, Any]] = []
    for abs_p in templates:
        rel = os.path.normpath(os.path.relpath(abs_p, vorlagen_ordner))
        for base_key in liste_media_basis_platzhalter_aus_docx(abs_p, log_callback):
            ref = resolve_reference_for_base_key(base_key, datensaetze, fallback_marken)
            cm, src, variable = resolve_size_for_base_key(base_key, datensaetze, fallback_marken)
            rows.append(
                {
                    "template_rel": rel,
                    "template_abs": abs_p,
                    "base_key": base_key,
                    "kind": _kind_label(base_key),
                    "reference": ref,
                    "size_key": canonical_size_key(base_key),
                    "cm_value": cm,
                    "cm_source": src,
                    "cm_variable": variable,
                }
            )
    return rows


def gruppier_nach_basis_platzhalter(detail_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Eine Zeile pro Basis-Platzhalter (z. B. ba_logo_img): alle Vorlagen, die ihn nutzen, in template_rels.
    Gleiche Excel-Textmarke → eine Größenänderung wirkt überall.
    """
    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in detail_rows:
        buckets[str(r.get("base_key", ""))].append(r)

    out: list[dict[str, Any]] = []
    for base_key in sorted(buckets.keys(), key=lambda x: x.lower()):
        if not base_key:
            continue
        grp = buckets[base_key]
        templates = sorted({g["template_rel"] for g in grp}, key=lambda x: x.lower())
        variables = any(g.get("cm_variable") for g in grp)
        cms = [float(g["cm_value"]) for g in grp]
        uniq_cm = sorted({round(c, 4) for c in cms})
        cm_conflict = len(uniq_cm) > 1

        non_empty_refs = {(g.get("reference") or "").strip() for g in grp}
        non_empty_refs.discard("")
        ref_conflict = len(non_empty_refs) > 1

        first = grp[0]
        effective_cm = uniq_cm[0] if len(uniq_cm) == 1 else float(first["cm_value"])

        src = first.get("cm_source", "")
        if cm_conflict:
            src = "variabel – bitte in Excel prüfen"

        out.append(
            {
                "base_key": base_key,
                "kind": first.get("kind", ""),
                "reference": first.get("reference", ""),
                "reference_conflict": ref_conflict,
                "size_key": canonical_size_key(base_key),
                "cm_value": effective_cm,
                "cm_source": src,
                "cm_variable": variables or cm_conflict or ref_conflict,
                "template_rels": templates,
                "templates_compact": "; ".join(templates[:12]) + (" …" if len(templates) > 12 else ""),
                "templates_tooltip": "\n".join(templates),
            }
        )
    return out


def _normalize_size_key_for_excel(size_key: str) -> str:
    sk = (size_key or "").strip()
    if not sk:
        return ""
    if sk.lower().endswith("_size"):
        return sk[:-5] + "_size"
    return f"{sk}_size"


def schreibe_groessen_nach_excel_blatt2(excel_path: str, updates: list[tuple[str, float]]) -> None:
    """
    Schreibt bzw. aktualisiert Zeilen auf Blatt 2: Spalte B = Textmarke (*_size), Spalte C = Wert.
    updates: (size_key, cm) – size_key wird als …_size normalisiert.
    """
    wb = load_workbook(excel_path)
    try:
        if len(wb.sheetnames) < 2:
            raise ValueError("Die Excel-Datei hat kein zweites Blatt (Fallback).")
        ws = wb[wb.sheetnames[1]]
        key_to_row: dict[str, int] = {}
        for r_idx in range(2, ws.max_row + 1):
            cell_b = ws.cell(row=r_idx, column=2).value
            if cell_b is None:
                continue
            key_to_row[str(cell_b).strip().lower()] = r_idx

        for size_key, cm_val in updates:
            canon = _normalize_size_key_for_excel(size_key)
            if not canon:
                continue
            lk = canon.lower()
            val_str = str(float(cm_val))
            if lk in key_to_row:
                ws.cell(row=key_to_row[lk], column=3, value=val_str)
            else:
                ws.append([None, canon, val_str])
                key_to_row[lk] = ws.max_row
        wb.save(excel_path)
    finally:
        wb.close()
