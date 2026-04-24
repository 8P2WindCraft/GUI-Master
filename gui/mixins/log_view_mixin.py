"""Haupt-Log-Tab, Filter, Kontextmenü, Excel-Vorschau in der Tabelle."""
import os
import re

import pandas as pd
from PySide6.QtGui import QTextCursor
from PySide6.QtWidgets import QApplication, QMenu, QMessageBox

from ..models import PandasModel


class LogViewMixin:
    """HTML-Log, Platzhalter aus Log ins Excel, Tabellen-Vorschau der Excel-Datei."""

    def append_html_log(self, html):
        """Ein Slot, der HTML-formatierten Text sicher an das Log-Fenster anhängt und für Filter speichert."""
        # Robust: Suche nach [ LEVEL ] mit beliebigen Leerzeichen, unabhängig von HTML-Tags
        m = re.search(r'\[\s*([A-Z]+)\s*\]', html)
        level = m.group(1) if m else "INFO"
        self.log_lines.append((level, html))
        self.apply_log_filter()

    def apply_log_filter(self):
        """Zeigt nur die Log-Zeilen an, die dem gewählten Filter entsprechen."""
        filter_level = self.log_filter_combo.currentText()
        self.log_text.clear()
        for level, html in self.log_lines:
            if filter_level == "ALLE" or level == filter_level:
                self.log_text.moveCursor(QTextCursor.End)
                self.log_text.insertHtml(html)
        self.log_text.ensureCursorVisible()

    def show_log_context_menu(self, position):
        """Zeigt ein Kontextmenü für Log-Einträge mit Optionen zum Öffnen von Ordnern/Dateien/Textmarken."""
        cursor = self.log_text.cursorForPosition(position)
        cursor.select(QTextCursor.WordUnderCursor)
        selected_text = cursor.selectedText()

        # Suche nach Pfaden in der aktuellen Zeile
        line_cursor = self.log_text.cursorForPosition(position)
        line_cursor.movePosition(QTextCursor.StartOfLine)
        line_cursor.movePosition(QTextCursor.EndOfLine, QTextCursor.KeepAnchor)
        line_text = line_cursor.selectedText()

        menu = QMenu(self)

        # Extrahiere mögliche Pfade aus der Zeile
        paths = self.extract_paths_from_log_line(line_text)

        if paths:
            # Füge Menüpunkte für jeden gefundenen Pfad hinzu
            for path_type, path, display_name in paths:
                if path_type == "file":
                    action = menu.addAction(f"📄 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_file(p))
                elif path_type == "folder":
                    action = menu.addAction(f"📁 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_folder(p))
                elif path_type == "template":
                    action = menu.addAction(f"📝 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_template_file(p))
            menu.addSeparator()

        # Prüfe auf fehlende Platzhalter/Textmarken
        fehlende_marken = self.extract_missing_placeholders_from_log_line(line_text)
        if fehlende_marken:
            action = menu.addAction("➕ Textmarke(n) im Excel hinzufügen")
            action.triggered.connect(lambda checked, marken=fehlende_marken: self.add_placeholders_to_excel(marken))
            menu.addSeparator()

        # Prüfe auf fehlende Size-Textmarken
        fehlende_size_marken = self.extract_missing_size_placeholders_from_log_line(line_text)
        if fehlende_size_marken:
            action = menu.addAction("➕ Size-Textmarke(n) im Excel hinzufügen")
            action.triggered.connect(lambda checked, marken=fehlende_size_marken: self.add_placeholders_to_excel(marken))
            menu.addSeparator()

        # Standard-Menüpunkte
        copy_action = menu.addAction("📋 Ausgewählten Text kopieren")
        copy_action.triggered.connect(self.copy_selected_text)

        clear_action = menu.addAction("🗑️ Log löschen")
        clear_action.triggered.connect(self.clear_log)

        menu.exec_(self.log_text.mapToGlobal(position))

    def extract_paths_from_log_line(self, line_text):
        """Extrahiert mögliche Datei- und Ordnerpfade aus einer Log-Zeile."""
        paths = []
        import re
        # Suche nach Dateinamen in verschiedenen Log-Formaten
        # 1. Export-Log: Dokument: 'Dateiname.docx'
        datei_matches = re.findall(r"Dokument: '([^']+)'", line_text)
        # 2. Fehler/Warnung: Vorlage 'Dateiname.docx'
        vorlage_matches = re.findall(r"Vorlage '([^']+)'", line_text)
        # 3. Export-Log: in Ordner: 'Ordnername'
        ordner_matches = re.findall(r"in Ordner: '([^']+)'", line_text)

        # Exportierte Dateien suchen (wie bisher)
        if hasattr(self, 'last_export_path') and self.last_export_path:
            base_path = self.last_export_path
            # Füge Dateien aus Export-Log hinzu
            for datei in datei_matches:
                for category_name in self.categories.values():
                    potential_path = os.path.join(base_path, category_name, datei)
                    if os.path.exists(potential_path):
                        paths.append(("file", potential_path, datei))
                        break
            # Füge Dateien aus Fehler/Warnung hinzu
            for datei in vorlage_matches:
                for category_name in self.categories.values():
                    potential_path = os.path.join(base_path, category_name, datei)
                    if os.path.exists(potential_path):
                        paths.append(("file", potential_path, datei))
                        break
            # Füge Ordner hinzu
            for ordner in ordner_matches:
                potential_path = os.path.join(base_path, ordner)
                if os.path.exists(potential_path):
                    paths.append(("folder", potential_path, ordner))

        # Vorlagen-Dateien suchen
        # Suche in self.paths['vorlagen_ordner'] und allen Unterordnern nach passenden Dateinamen
        if hasattr(self, 'paths') and 'vorlagen_ordner' in self.paths:
            vorlagen_root = self.paths['vorlagen_ordner']
            alle_vorlagen = set(datei_matches + vorlage_matches)
            for suchname in alle_vorlagen:
                for root, _, files in os.walk(vorlagen_root):
                    for file in files:
                        if file == suchname:
                            full_path = os.path.join(root, file)
                            paths.append(("template", full_path, suchname))
        return paths

    def extract_missing_placeholders_from_log_line(self, line_text):
        """Extrahiert fehlende Platzhalter/Textmarken aus einer Log-Zeile."""
        import re
        # Sucht nach: Fehlende Platzhalter in Excel: {'foo', 'bar'}
        match = re.search(r"Fehlende Platzhalter in Excel: (\{.*?\})", line_text)
        if match:
            try:
                # Sichere Auswertung des Sets
                marken = eval(match.group(1), {"__builtins__": None}, {})
                if isinstance(marken, set):
                    return sorted(marken)
            except Exception:
                pass
        return []

    def extract_missing_size_placeholders_from_log_line(self, line_text):
        """Extrahiert fehlende Size-Textmarken aus einer Log-Zeile."""
        import re
        # NEU: Erkenne auch die neue Info-Log-Zeile
        # Beispiel: Hinweis: Für Platzhalter 'ba_logo_img' wird die Standardgröße verwendet, da kein 'ba_logo_img_size' oder 'ba_logo_img_Size' in Excel/Fallback gefunden wurde.
        match_alt = re.search(r"Keine Größenangabe\s*\(([^)]+)\).*?für Platzhalter", line_text)
        match_neu = re.search(
            r"Hinweis: Für Platzhalter '([^']+)' wird die Standardgröße verwendet, da kein '([^']+)' oder '([^']+)' in Excel/Fallback gefunden wurde",
            line_text
        )
        if match_alt:
            inhalt = match_alt.group(1)
            self.append_html_log(f'<span style="color:#888">[DEBUG] Klammer-Inhalt: {inhalt}</span><br>')
            marken = re.findall(r"'([^']+)'", inhalt)
            if marken:
                self.append_html_log(
                    f'<span style="color:#888">[DEBUG] Erkannte Size-Textmarken: {", ".join(marken)}</span><br>'
                )
            else:
                self.append_html_log(f'<span style="color:#888">[DEBUG] Keine Size-Textmarken erkannt.</span><br>')
            return sorted(marken)
        elif match_neu:
            # Extrahiere die beiden _size-Varianten
            size1 = match_neu.group(2)
            size2 = match_neu.group(3)
            marken = [size1, size2]
            self.append_html_log(
                f'<span style="color:#888">[DEBUG] Erkannte Size-Textmarken (neu): {", ".join(marken)}</span><br>'
            )
            return sorted(marken)
        self.append_html_log(f'<span style="color:#888">[DEBUG] Keine Size-Textmarken erkannt.</span><br>')
        return []

    def add_placeholders_to_excel(self, placeholders):
        """Fügt die angegebenen Platzhalter als neue Zeilen im zweiten Blatt der aktuellen Excel-Datei hinzu (Spalte 2=Textmarke, Spalte 3=leer)."""
        from openpyxl import load_workbook
        import os
        excel_path = self.paths.get('excel_path')
        if not excel_path or not os.path.exists(excel_path):
            QMessageBox.warning(self, "Fehler", "Excel-Datei nicht gefunden!")
            return
        wb = None
        try:
            wb = load_workbook(excel_path)
            if len(wb.sheetnames) < 2:
                QMessageBox.warning(self, "Fehler", "Die Excel-Datei hat kein zweites Blatt!")
                return
            ws = wb[wb.sheetnames[1]]
            # Nur eine Size-Variante pro Basis-Textmarke einfügen
            filtered = {}
            for marke in placeholders:
                if marke.lower().endswith('_size'):
                    base = marke[:-5].lower()
                    # Bevorzuge die Variante mit kleinem 's'
                    if base not in filtered or marke.endswith('_size'):
                        filtered[base] = marke
                else:
                    filtered[marke] = marke
            for marke in filtered.values():
                ws.append([None, marke, ""])
            wb.save(excel_path)
            QMessageBox.information(
                self,
                "Erfolg",
                f"{len(filtered)} Textmarke(n) wurden im zweiten Blatt hinzugefügt."
            )
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte Textmarken nicht hinzufügen: {e}")
        finally:
            if wb is not None:
                wb.close()

    def open_file(self, file_path):
        """Öffnet eine Datei mit der Standard-Anwendung."""
        try:
            os.startfile(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte die Datei nicht öffnen: {e}")

    def open_folder(self, folder_path):
        """Öffnet einen Ordner im Explorer."""
        try:
            os.startfile(folder_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte den Ordner nicht öffnen: {e}")

    def open_template_file(self, file_path):
        """Öffnet eine Word-Vorlage mit der Standard-Anwendung."""
        try:
            os.startfile(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte die Vorlage nicht öffnen: {e}")

    def copy_selected_text(self):
        """Kopiert den ausgewählten Text in die Zwischenablage."""
        cursor = self.log_text.textCursor()
        if cursor.hasSelection():
            text = cursor.selectedText()
            clipboard = QApplication.clipboard()
            clipboard.setText(text)

    def clear_log(self):
        """Löscht den gesamten Log-Inhalt."""
        reply = QMessageBox.question(
            self,
            "Log löschen",
            "Möchten Sie wirklich den gesamten Log löschen?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.log_text.clear()

    def show_excel_data(self):
        """
        Versucht, die Daten aus der angegebenen Excel-Datei zu laden und
        in der Tabelle (QTableView) in der GUI anzuzeigen.
        Bei gesperrter Datei (z. B. in Excel geöffnet) wird unter Windows
        ein COM-Fallback verwendet.
        """
        path = self.paths.get('excel_path')
        header_row = self.settings.get('header_row', 3)
        if path and os.path.exists(path):
            try:
                df = pd.read_excel(path, header=header_row - 1)
                self.table.setModel(PandasModel(df))
            except (PermissionError, OSError) as e:
                if getattr(e, 'errno', None) == 13 or isinstance(e, PermissionError):
                    try:
                        from core.excel_com import read_excel_sheet1_via_com
                        df = read_excel_sheet1_via_com(path, header_row)
                        if df is not None:
                            self.table.setModel(PandasModel(df))
                            return
                    except Exception:
                        pass
                self.table.setModel(None)
                QMessageBox.critical(self, "Fehler beim Lesen der Excel-Datei", str(e))
            except Exception as e:
                self.table.setModel(None)  # Clear table on error
                QMessageBox.critical(self, "Fehler beim Lesen der Excel-Datei", str(e))

    def open_export_folder(self):
        """Öffnet den zuletzt verwendeten Export-Ordner im Datei-Explorer."""
        if self.last_export_path and os.path.isdir(self.last_export_path):
            try:
                os.startfile(self.last_export_path)
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Konnte den Ordner nicht öffnen: {e}")

    def open_excel_file(self):
        excel_path = self.paths.get('excel_path')
        if excel_path and os.path.exists(excel_path):
            try:
                os.startfile(excel_path)
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Excel konnte nicht geöffnet werden: {e}")
        else:
            QMessageBox.warning(self, "Fehler", "Excel-Datei nicht gefunden!")
