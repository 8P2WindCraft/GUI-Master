"""
Projektdateien (*.dta.json), portable Pfade ({username}) und Vorschlagsdateinamen.
"""
import json
import os
import re
from datetime import datetime

import pandas as pd
from PySide6.QtGui import QAction
from PySide6.QtWidgets import QFileDialog, QMessageBox

from ..constants import MAX_RECENT_PROJECTS, PROJECT_FILE_PATH_KEYS, VERSION


class ProjectFileMixin:
    """
    Ersetzt Windows-Profilpfade durch ``{username}``, liest/schreibt Projekt-JSON,
    Menü „Letzte Projekte“, Default-Dateiname für „Speichern unter…“.
    """

    _RE_USERS_PROFILE = re.compile(
        r'(?i)([a-z]:)([\\/])Users\2([^\\/]+)',
    )
    _RE_USERS_PLACEHOLDER = re.compile(
        r'(?i)([a-z]:)([\\/])Users\2\{username\}',
    )

    def _project_root_dir(self):
        """Liefert das konfigurierte zentrale Projektverzeichnis (lokal aufgelöst)."""
        root = (self.settings.get('project_root_dir', '') or '').strip()
        if not root and hasattr(self, 'project_root_dir_edit'):
            root = self.project_root_dir_edit.text().strip()
        return self._denormalize_path(root)

    def on_project_root_dir_change(self):
        """Speichert Änderungen am zentralen Projektverzeichnis."""
        if not hasattr(self, 'project_root_dir_edit'):
            return
        value = self.project_root_dir_edit.text().strip()
        self.settings['project_root_dir'] = value
        self.save_all_settings()
        self.validate_path(self.project_root_dir_edit)

    def browse_project_root_dir(self):
        """Ordnerdialog für das zentrale Projektverzeichnis."""
        start_dir = self._project_root_dir() or os.getcwd()
        picked = QFileDialog.getExistingDirectory(self, "Projektverzeichnis auswählen", start_dir)
        if not picked:
            return
        if hasattr(self, 'project_root_dir_edit'):
            self.project_root_dir_edit.setText(picked)
        self.on_project_root_dir_change()

    def _project_dialog_start_dir(self):
        """
        Startordner für Projektdatei-Dialoge:
        1) Ordner der aktuell geöffneten Projektdatei,
        2) zentrales Projektverzeichnis,
        3) cwd.
        """
        if self.current_project_path:
            cur_dir = os.path.dirname(self.current_project_path)
            if cur_dir and os.path.isdir(cur_dir):
                return cur_dir
        root = self._project_root_dir()
        if root and os.path.isdir(root):
            return root
        return os.getcwd()

    def _normalize_path(self, path):
        """
        Normalisiert einen Pfad für portables Speichern:
        Ersetzt den Profilordner unter …\\Users\\<name> durch {username}.
        Laufwerk und Schreibweise von „Users“ werden case-insensitive erkannt.
        """
        if not path:
            return path

        def _repl(m):
            return f"{m.group(1)}{m.group(2)}Users{m.group(2)}{{username}}"

        return self._RE_USERS_PROFILE.sub(_repl, path)

    def _denormalize_path(self, path):
        """
        Denormalisiert einen Pfad beim Laden:
        Ersetzt {username} durch den aktuellen Benutzernamen (c:/users/… ebenfalls).
        """
        if not path:
            return path

        def _repl(m):
            return f"{m.group(1)}{m.group(2)}Users{m.group(2)}{self.current_username}"

        return self._RE_USERS_PLACEHOLDER.sub(_repl, path)

    def normalize_all_paths(self):
        """
        Normalisiert alle aktuellen Pfade in der UI (Button-Callback).
        Zeigt danach eine Bestätigung an.
        """
        changed_count = 0
        for key, value in self.paths.items():
            if value and '{username}' not in value:
                normalized = self._normalize_path(value)
                if normalized != value:
                    self.paths[key] = normalized
                    if hasattr(self, f"{key}_edit"):
                        getattr(self, f"{key}_edit").setText(normalized)
                        self.validate_path(getattr(self, f"{key}_edit"))
                    changed_count += 1
        root = (self.settings.get('project_root_dir', '') or '').strip()
        if root and '{username}' not in root:
            normalized_root = self._normalize_path(root)
            if normalized_root != root:
                self.settings['project_root_dir'] = normalized_root
                if hasattr(self, 'project_root_dir_edit'):
                    self.project_root_dir_edit.setText(normalized_root)
                    self.validate_path(self.project_root_dir_edit)
                changed_count += 1

        if changed_count > 0:
            self.save_all_settings()
            QMessageBox.information(
                self,
                "Pfade normalisiert",
                f"{changed_count} Pfade wurden normalisiert (Benutzername → '{{username}}').\n\n"
                "Die Pfade funktionieren jetzt auf allen Rechnern!"
            )
        else:
            QMessageBox.information(
                self,
                "Keine Normalisierung",
                "Entweder ist '{{username}}' bereits gesetzt, oder es wurde kein Windows-Profilpfad "
                "im Muster Laufwerk:\\Users\\<Benutzername> erkannt.",
            )

    def denormalize_all_paths(self):
        """Ersetzt {username} in allen Pfaden durch den angemeldeten Benutzer (lokale Nutzung)."""
        changed_count = 0
        for key, value in list(self.paths.items()):
            if not value or '{username}' not in value:
                continue
            denorm = self._denormalize_path(value)
            if denorm != value:
                self.paths[key] = denorm
                if hasattr(self, f"{key}_edit"):
                    getattr(self, f"{key}_edit").setText(denorm)
                    self.validate_path(getattr(self, f"{key}_edit"))
                changed_count += 1
        root = (self.settings.get('project_root_dir', '') or '').strip()
        if root and '{username}' in root:
            denorm_root = self._denormalize_path(root)
            if denorm_root != root:
                self.settings['project_root_dir'] = denorm_root
                if hasattr(self, 'project_root_dir_edit'):
                    self.project_root_dir_edit.setText(denorm_root)
                    self.validate_path(self.project_root_dir_edit)
                changed_count += 1

        if changed_count > 0:
            self.save_all_settings()
            QMessageBox.information(
                self,
                "Pfade lokal gesetzt",
                f"{changed_count} Pfad(e): '{{username}}' wurde durch '{self.current_username}' ersetzt.",
            )
        else:
            QMessageBox.information(
                self,
                "Keine Änderung",
                "Kein '{{username}}' in den Pfaden gefunden (nichts zu ersetzen).",
            )

    def _collect_project_dict(self):
        """
        Baut das serialisierbare Dict für eine Projektdatei (*.dta.json).

        Enthält: Version, Pfade (mit {username}), Einstellungen, Kategorien-Map, Vorlagen-Auswahl,
        Regel-Flags und Beschilderungsregeln. Reihenfolge der Pfad-Keys entspricht PROJECT_FILE_PATH_KEYS.
        """
        # UI-Werte vor dem Schreiben in self.settings übernehmen (Spins, Texte, Checkboxen)
        for key in self.settings:
            if hasattr(self, f"{key}_spin"):
                self.settings[key] = getattr(self, f"{key}_spin").value()
        if hasattr(self, 'datetime_utc_format_edit'):
            self.settings['datetime_utc_format'] = self.datetime_utc_format_edit.text()
        if hasattr(self, 'projekt_name_override_edit'):
            self.settings['projekt_name_override'] = self.projekt_name_override_edit.text().strip()
        if hasattr(self, 'export_as_pdf_check'):
            self.settings['export_as_pdf'] = self.export_as_pdf_check.isChecked()

        self._snapshot_template_selection()
        self._snapshot_signage_rules_from_ui()

        # Pfade normalisieren (Benutzername → {username}) für Portabilität; nur definierte Keys
        normalized_paths = {
            k: self._normalize_path(self.paths.get(k, '')) for k in PROJECT_FILE_PATH_KEYS
        }

        return {
            'version': VERSION,
            'paths': normalized_paths,
            'settings': self.settings,
            'categories': self.categories,
            'selected_template_rel_paths': self.selected_template_rel_paths,
            'rules_enabled': self.rules_enabled,
            'reuse_lageplan_from_last_export': self.reuse_lageplan_from_last_export,
            'signage_rules': self.signage_rules,
            'saved_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    def _apply_project_dict(self, data):
        """
        Wendet ein geladenes Projekt-JSON auf Zustand und UI an.

        - Pfade: immer genau PROJECT_FILE_PATH_KEYS; fehlende Keys in der Datei werden als leer geladen.
        - settings: flach mergen.
        - categories: nur wenn nicht leer; mergen in bestehende self.categories, Präfix p_ bleibt garantiert.
        Anschließend: Felder setzen, Theme, Kategorien-Widgets neu aufbauen, Excel-Vorschau, Speichern.
        """
        loaded_paths = data.get('paths', self.paths) or {}
        self.paths = {
            k: self._denormalize_path(loaded_paths.get(k, ''))
            for k in PROJECT_FILE_PATH_KEYS
        }

        self.settings = {**self.settings, **data.get('settings', {})}
        if 'selected_template_rel_paths' in data:
            self.selected_template_rel_paths = data.get('selected_template_rel_paths')
        else:
            self.selected_template_rel_paths = None
        if 'rules_enabled' in data:
            self.rules_enabled = bool(data.get('rules_enabled'))
        if 'reuse_lageplan_from_last_export' in data:
            self.reuse_lageplan_from_last_export = bool(data.get('reuse_lageplan_from_last_export'))
        if 'signage_rules' in data:
            self.signage_rules = data.get('signage_rules') or []

        if isinstance(data.get('categories'), dict) and data.get('categories'):
            # Datei-Inhalt hat Vorrang vor dem bisherigen Speicher-Stand (z. B. settings.json)
            self.categories = {**self.categories, **data['categories']}
            self.categories.setdefault('p_', 'Pläne')

        # UI: die vier Pfadzeilen
        for key, value in self.paths.items():
            if hasattr(self, f"{key}_edit"):
                getattr(self, f"{key}_edit").setText(value)
                self.validate_path(getattr(self, f"{key}_edit"))
        # Settings (Spins)
        for key in ['header_row', 'svg_scale', 'png_compression']:
            if hasattr(self, f"{key}_spin"):
                getattr(self, f"{key}_spin").setValue(
                    self.settings.get(key, getattr(self, f"{key}_spin").value())
                )
        # UTC Format
        if hasattr(self, 'datetime_utc_format_edit'):
            self.datetime_utc_format_edit.setText(
                self.settings.get('datetime_utc_format', self.datetime_utc_format_edit.text())
            )
        # Projektname (Export-Ordner)
        if hasattr(self, 'projekt_name_override_edit'):
            self.projekt_name_override_edit.setText(self.settings.get('projekt_name_override', ''))
        if hasattr(self, 'project_root_dir_edit'):
            project_root = self._denormalize_path(self.settings.get('project_root_dir', ''))
            self.settings['project_root_dir'] = project_root
            self.project_root_dir_edit.setText(project_root)
            self.validate_path(self.project_root_dir_edit)
        # PDF-Export
        if hasattr(self, 'export_as_pdf_check'):
            self.export_as_pdf_check.setChecked(self.settings.get('export_as_pdf', False))
        # Theme
        self.apply_theme(self.settings.get('theme', 'Light'))
        # Kategorien-UI neu aufbauen
        if hasattr(self, 'categories_ui_layout'):
            # Bestehende Widgets entfernen
            for layout_item, _triplet in list(self.category_widgets.items()):
                while layout_item.count():
                    child = layout_item.takeAt(0)
                    if child.widget():
                        child.widget().deleteLater()
                self.categories_ui_layout.removeItem(layout_item)
                layout_item.deleteLater()
                del self.category_widgets[layout_item]
            # Neu hinzufügen
            for prefix, name in self.categories.items():
                self.add_category_widget(prefix, name)
        # Excel-Vorschau aktualisieren
        self.show_excel_data()
        if hasattr(self, 'refresh_template_checkboxes'):
            self.refresh_template_checkboxes()
        if hasattr(self, '_apply_signage_rules_to_ui'):
            self._apply_signage_rules_to_ui()
        # Einstellungen persistieren
        self.save_all_settings()
        if hasattr(self, 'refresh_datenansicht'):
            self.refresh_datenansicht()
        if hasattr(self, 'refresh_media_sizes_table'):
            self.refresh_media_sizes_table()

    def _remember_recent_project(self, file_path):
        """Merkt sich ein Projekt in der Liste zuletzt verwendeter Projekte."""
        try:
            if not file_path:
                return
            if file_path in self.recent_projects:
                self.recent_projects.remove(file_path)
            self.recent_projects.insert(0, file_path)
            self.recent_projects = self.recent_projects[:10]
            self.current_project_path = file_path
            # in settings.json speichern
            self.save_all_settings()
        except Exception:
            pass

    def _refresh_recent_projects_menu(self):
        """Aktualisiert das Untermenü 'Letzte Projekte' mit den zuletzt geöffneten Projekten."""
        self.recent_projects_menu.clear()
        recent = getattr(self, 'recent_projects', [])[:MAX_RECENT_PROJECTS]
        for path in recent:
            if not os.path.isfile(path):
                continue
            display_name = os.path.basename(path)
            action = QAction(display_name, self)
            action.setToolTip(path)
            action.triggered.connect(lambda checked, p=path: self.project_open_path(p))
            self.recent_projects_menu.addAction(action)
        if not self.recent_projects_menu.actions():
            none_action = QAction("(Keine)", self)
            none_action.setEnabled(False)
            self.recent_projects_menu.addAction(none_action)

    def project_new(self):
        """Setzt ein leeres Projekt (ohne Pfade), behält aber Kategorien/Settings-Grundwerte bei."""
        self.paths = {k: '' for k in PROJECT_FILE_PATH_KEYS}
        self.selected_template_rel_paths = None
        self.rules_enabled = False
        self.reuse_lageplan_from_last_export = False
        self.signage_rules = []
        for key in ['header_row', 'svg_scale', 'png_compression']:
            if hasattr(self, f"{key}_spin"):
                getattr(self, f"{key}_spin").setValue(
                    self.settings.get(key, getattr(self, f"{key}_spin").minimum())
                )
        if hasattr(self, 'datetime_utc_format_edit'):
            self.datetime_utc_format_edit.setText('%Y-%m-%d %H:%M:%S UTC')
        self.apply_theme('Light')
        for key in self.paths:
            if hasattr(self, f"{key}_edit"):
                getattr(self, f"{key}_edit").setText('')
                self.validate_path(getattr(self, f"{key}_edit"))
        self.current_project_path = None
        if hasattr(self, 'refresh_template_checkboxes'):
            self.refresh_template_checkboxes()
        if hasattr(self, '_apply_signage_rules_to_ui'):
            self._apply_signage_rules_to_ui()
        if hasattr(self, 'refresh_media_sizes_table'):
            self.refresh_media_sizes_table()
        self.save_all_settings()

    def project_open_path(self, file_path):
        """Lädt ein Projekt aus dem angegebenen Pfad (ohne Dateidialog)."""
        if not file_path or not os.path.isfile(file_path):
            if file_path:
                QMessageBox.warning(self, "Fehler", f"Projektdatei nicht gefunden:\n{file_path}")
            return
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Neues Format: verschachtelte "paths" + "settings"; Alt: flache Keys wie bei settings.json
            if 'paths' in data and 'settings' in data:
                self._apply_project_dict(data)
            else:
                # Flaches Format: Top-Level-Keys → dieselbe Struktur wie neues Format
                self._apply_project_dict({
                    'paths': {k: data.get(k, '') for k in PROJECT_FILE_PATH_KEYS},
                    'settings': {k: data.get(k, self.settings.get(k)) for k in self.settings.keys()},
                    'categories': data.get('categories', {}),
                    'rules_enabled': data.get('rules_enabled', False),
                    'reuse_lageplan_from_last_export': data.get('reuse_lageplan_from_last_export', False),
                    'signage_rules': data.get('signage_rules', []),
                })
            self._remember_recent_project(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Projekt konnte nicht geladen werden: {e}")

    def project_open(self):
        """Öffnet eine Projektdatei (*.dta.json) per Dateidialog."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Projekt öffnen", self._project_dialog_start_dir(), "Projektdateien (*.dta.json *.json)"
        )
        if file_path:
            self.project_open_path(file_path)

    def project_save(self):
        """Speichert das aktuelle Projekt an den zuletzt genutzten Projektpfad oder fragt nach einem neuen Pfad."""
        if not self.current_project_path:
            return self.project_save_as()
        try:
            data = self._collect_project_dict()
            with open(self.current_project_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            QMessageBox.information(self, "Gespeichert", f"Projekt gespeichert: {self.current_project_path}")
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Projekt konnte nicht gespeichert werden: {e}")

    def project_save_as(self):
        """Speichert das aktuelle Projekt unter einem neuen Dateinamen (*.dta.json)."""
        default_name = self._build_default_project_filename()
        start_dir = self._project_dialog_start_dir()
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Projekt speichern unter",
            os.path.join(start_dir, default_name),
            "Projektdateien (*.dta.json)"
        )
        if not file_path:
            return
        try:
            data = self._collect_project_dict()
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            self._remember_recent_project(file_path)
            QMessageBox.information(self, "Gespeichert", f"Projekt gespeichert: {file_path}")
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Projekt konnte nicht gespeichert werden: {e}")

    def _sanitize_filename_part(self, value):
        """Bereinigt einen String für sichere Windows-Dateinamen (verbotene Zeichen → _)."""
        text = str(value or "").strip()
        if not text:
            return ""
        sanitized = "".join(c if c not in r'\/:*?"<>|' else "_" for c in text)
        sanitized = sanitized.replace("\n", " ").replace("\r", " ").strip(" .")
        return sanitized

    def _base_name_for_project_dta_file(self, value):
        """
        Liefert den Basisnamen für …/name.dta.json (ohne Endung).

        Entfernt mehrfach angehängtes .dta bzw. .dta.json (z. B. aus Excel-Zelle „Park.dta“),
        damit der Vorschlagsname nicht „Park.dta.dta.json“ wird.
        """
        s = self._sanitize_filename_part(value)
        if not s:
            return ""
        while True:
            low = s.lower()
            if low.endswith('.dta.json'):
                s = s[:-9].rstrip(' .')
            elif low.endswith('.dta'):
                s = s[:-4].rstrip(' .')
            else:
                break
        return s

    def _read_name_from_excel(self):
        """
        Liest einen möglichen Winpark-/Projektnamen aus der aktuell gewählten Excel-Datei.
        Unterstützt Blatt 1 (Spalten) und Blatt 2 (Fallback-Marken).
        """
        excel_path = self.paths.get('excel_path', '').strip()
        if not excel_path or not os.path.isfile(excel_path):
            return ""

        key_priority = ['winpark_name', 'windpark_name', 'winpark', 'windpark', 'projekt_name']
        header_row = self.settings.get('header_row', 3)

        try:
            df = pd.read_excel(excel_path, header=header_row - 1)
            if not df.empty:
                normalized_cols = {str(col).strip().lower(): col for col in df.columns}
                for key in key_priority:
                    col = normalized_cols.get(key)
                    if col is None:
                        continue
                    series = df[col].dropna()
                    if not series.empty:
                        candidate = self._sanitize_filename_part(series.iloc[0])
                        if candidate:
                            return candidate
        except Exception:
            pass

        wb = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            if len(wb.sheetnames) > 1:
                ws = wb[wb.sheetnames[1]]
                fallback = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 3 and row[1]:
                        fallback_key = str(row[1]).strip().lower()
                        fallback_value = str(row[2]).strip() if row[2] is not None else ""
                        fallback[fallback_key] = fallback_value
                for key in key_priority:
                    candidate = self._sanitize_filename_part(fallback.get(key, ""))
                    if candidate:
                        return candidate
        except Exception:
            pass
        finally:
            if wb is not None:
                wb.close()

        return ""

    def _build_default_project_filename(self):
        """
        Erzeugt den Standard-Dateinamen für 'Speichern unter...':
        1) aktueller Projektdateiname (falls vorhanden),
        2) Projektname-Override,
        3) Winpark-/Projektname aus Excel,
        4) Fallback.
        """
        if self.current_project_path:
            current_name = os.path.basename(self.current_project_path)
            if current_name:
                return current_name

        override = ""
        if hasattr(self, 'projekt_name_override_edit'):
            override = self._sanitize_filename_part(self.projekt_name_override_edit.text())
        if not override:
            override = self._sanitize_filename_part(self.settings.get('projekt_name_override', ''))
        if override:
            base = self._base_name_for_project_dta_file(override)
            return f"{base}.dta.json" if base else 'projekt.dta.json'

        excel_name = self._read_name_from_excel()
        if excel_name:
            base = self._base_name_for_project_dta_file(excel_name)
            return f"{base}.dta.json" if base else 'projekt.dta.json'

        return 'projekt.dta.json'  # letzter Fallback wenn weder Override noch Excel-Name
