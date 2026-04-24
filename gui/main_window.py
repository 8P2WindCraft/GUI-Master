import io
import json
import os
import re
import sys
import tempfile
import traceback

import pandas as pd
from datetime import datetime

from PySide6.QtCore import Qt, QAbstractTableModel, QThread, Signal
from PySide6.QtGui import QIcon, QAction, QActionGroup, QTextCursor, QPixmap, QImage
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QFileDialog, QProgressBar, QTableView, QMessageBox, QFrame,
    QSpinBox, QDoubleSpinBox, QTextEdit, QMainWindow, QTabWidget, QGroupBox, QScrollArea, QCheckBox,
    QMenu, QComboBox, QTableWidget, QTableWidgetItem, QSplitter, QListWidget,
)

from core.logic import (
    _convert_docx_to_pdf_in_folder,
    liste_textmarken_aus_docx,
    sammle_vorlagen_pfade,
    verarbeite_vorlagen,
    verarbeite_vorlagen_preview,
)
from core.logging import _log_handler
from core.media_sizes import (
    gruppier_nach_basis_platzhalter,
    sammle_media_groessen_uebersicht,
    schreibe_groessen_nach_excel_blatt2,
)
from core.utils import resource_path


# ==============================================================================
# VERSION
# ==============================================================================
VERSION = "7.0.1"
BUILD_DATE = "2025-01-27"

# Unterstützte Bildformate für den Bilder-Vorschau-Tab
BILDER_VORSCHAU_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.svg', '.bmp'}

# (value, label) für Textbedingungen im Regeleditor
RULE_CONDITION_CHOICES = (
    ("not_empty", "Nicht leer (nach Trim)"),
    ("length_gt", "Text länger als … Zeichen (>)"),
    ("length_gte", "Mindestlänge … Zeichen (≥)"),
    ("contains", "Enthält Text …"),
    ("equals", "Exakt gleich …"),
    ("equals_ignorecase", "Gleich … (Groß/Klein egal)"),
    ("regex", "Erfüllt Regex …"),
)

# Maximale Anzahl Einträge im Untermenü "Letzte Projekte"
MAX_RECENT_PROJECTS = 6

# Pfade in Projekt-JSON: immer exakt diese vier Keys, damit UI und Logik stabil bleiben
PROJECT_FILE_PATH_KEYS = (
    'excel_path', 'vorlagen_ordner', 'bilder_ordner', 'export_ordner',
)


class PandasModel(QAbstractTableModel):
    """
    Ein benutzerdefiniertes Tabellenmodell für PySide6, das einen
    Pandas DataFrame als Datenquelle für eine QTableView verwenden kann.
    Dies ermöglicht die Anzeige der Excel-Daten direkt in der GUI.
    """
    def __init__(self, df):
        super().__init__()
        self._df = df

    def rowCount(self, parent=None):
        return self._df.shape[0]

    def columnCount(self, parent=None):
        return self._df.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid() and role == Qt.DisplayRole:
            return str(self._df.iloc[index.row(), index.column()])

    def headerData(self, col, orient, role=Qt.DisplayRole):
        if orient == Qt.Horizontal and role == Qt.DisplayRole:
            return self._df.columns[col]


class Worker(QThread):
    """
    Ein QThread-Worker, der die zeitintensive Dokumentenerstellung in einem
    separaten Thread ausführt. Dies verhindert, dass die GUI während des
    Prozesses "einfriert" und nicht mehr reagiert.

    Signale:
    - log: Sendet Log-Nachrichten an die GUI.
    - progress: Sendet den Fortschritt (aktueller Wert, Maximalwert).
    - finished: Signalisiert das Ende der Verarbeitung.
    - current_file: Sendet den Namen der aktuell bearbeiteten Datei.
    """
    log = Signal(str)  # html_message
    progress = Signal(int, int)
    finished = Signal(bool, str)  # success, export_path
    current_file = Signal(str)

    def __init__(self, **kwargs):
        super().__init__()
        self.params = kwargs
        self.dry_run = kwargs.get('dry_run', False)
        self.preview_run = kwargs.get('preview_run', False)

    def run(self):
        """Startet die Verarbeitung durch Aufruf der `verarbeite_vorlagen` Funktion."""
        export_path = None
        try:
            thread_safe_params = self.params.copy()
            thread_safe_params['log_callback'] = self.log.emit
            thread_safe_params['progress_callback'] = self.progress.emit
            thread_safe_params['file_callback'] = self.current_file.emit
            thread_safe_params['worker_thread'] = self
            thread_safe_params.pop('theme', None)
            preview_run = thread_safe_params.pop('preview_run', False)
            preview_template_abs = thread_safe_params.pop('preview_template_abs', None)

            if preview_run:
                thread_safe_params.pop('dry_run', None)
                result = verarbeite_vorlagen_preview(
                    **thread_safe_params,
                    preview_template_abs=preview_template_abs,
                )
                if not self.isInterruptionRequested():
                    self.finished.emit(bool(result), result)
                else:
                    self.finished.emit(False, None)
            else:
                result = verarbeite_vorlagen(**thread_safe_params)

                if self.dry_run:
                    self.finished.emit(result, None)  # Bei Trockenlauf ist das Ergebnis ein Boolean
                elif not self.isInterruptionRequested():
                    self.finished.emit(bool(result), result)  # Bei echtem Lauf ist es der Pfad
                else:
                    self.finished.emit(False, None)

        except Exception as e:
            error_msg = f"FATALER FEHLER im Worker-Thread: {e}\n{traceback.format_exc()}"
            _log_handler(error_msg, "FATAL", self.log.emit)
            self.finished.emit(False, None)


class PdfWorker(QThread):
    """
    Worker für die manuelle PDF-Konvertierung des letzten Export-Ordners.
    Läuft im Hintergrund, damit die GUI während der Konvertierung responsiv bleibt.
    """
    log = Signal(str)  # html_message
    progress = Signal(int, int)
    current_file = Signal(str)
    finished = Signal(object)  # result dict

    def __init__(self, export_path, is_dark_mode=False):
        super().__init__()
        self.export_path = export_path
        self.is_dark_mode = is_dark_mode

    def run(self):
        try:
            def _worker_log(msg, level="INFO"):
                _log_handler(msg, level, self.log.emit, self.is_dark_mode)

            result = _convert_docx_to_pdf_in_folder(
                self.export_path,
                _worker_log,
                progress_callback=self.progress.emit,
                file_callback=self.current_file.emit
            )
            self.finished.emit(result)
        except Exception as e:
            self.finished.emit(
                {
                    "docx_count": 0,
                    "success_count": 0,
                    "failure_count": 0,
                    "error": "unknown_error",
                    "error_message": str(e),
                }
            )


class MainWindow(QMainWindow):
    """
    Das Hauptfenster der Anwendung.
    Es initialisiert die Benutzeroberfläche, verwaltet die Benutzereingaben,
    lädt und speichert Einstellungen und startet den Worker-Thread.
    """
    def __init__(self):
        # #region agent log
        log_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.cursor', 'debug.log')
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                import json
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"C","location":"gui/main_window.py:109","message":"MainWindow.__init__ entry","data":{},"timestamp":int(__import__('time').time()*1000)}) + '\n')
        except: pass
        # #endregion
        super().__init__()
        self.setWindowTitle(f'DocxTpl Automatisierung v{VERSION} - Build {BUILD_DATE}')
        self.category_widgets = {}
        self.theme_action_group = None
        self.last_export_path = None
        # Aktuellen Windows-Benutzernamen ermitteln (für portable Pfade)
        self.current_username = os.environ.get('USERNAME', os.environ.get('USER', 'UnknownUser'))
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                import json
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"C","location":"gui/main_window.py:118","message":"Before load_settings_and_init_vars","data":{"username":self.current_username},"timestamp":int(__import__('time').time()*1000)}) + '\n')
        except: pass
        # #endregion
        self.load_settings_and_init_vars()
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                import json
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"gui/main_window.py:120","message":"Before setup_ui","data":{"paths_keys":list(self.paths.keys())},"timestamp":int(__import__('time').time()*1000)}) + '\n')
        except: pass
        # #endregion
        self.setup_ui()
        # #region agent log
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                import json
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"gui/main_window.py:122","message":"After setup_ui","data":{"has_bilder_ordner_edit":hasattr(self, 'bilder_ordner_edit')},"timestamp":int(__import__('time').time()*1000)}) + '\n')
        except: pass
        # #endregion

    def load_settings_and_init_vars(self):
        """Lädt die zuletzt verwendeten Einstellungen (Pfade, Optionen) aus einer JSON-Datei."""
        last = self.load_settings()
        self.paths = {
            'excel_path': last.get('excel_path', ''),
            'vorlagen_ordner': last.get('vorlagen_ordner', ''),
            'bilder_ordner': last.get('bilder_ordner', ''),
            'export_ordner': last.get('export_ordner', '')
        }
        # Gespeicherte portable Pfade ({username}) sofort für dieses Konto auflösen
        self.paths = {k: self._denormalize_path(v) for k, v in self.paths.items()}
        self.settings = {
            'header_row': last.get('header_row', 3),
            'svg_scale': last.get('svg_scale', 3),
            'png_compression': last.get('png_compression', -1),
            'theme': last.get('theme', 'Light'),
            'datetime_utc_format': last.get('datetime_utc_format', '%Y-%m-%d %H:%M:%S UTC'),
            'projekt_name_override': last.get('projekt_name_override', ''),
            'export_as_pdf': last.get('export_as_pdf', False)
        }
        self.categories = last.get(
            'categories',
            {'b_': 'Beschilderung', 'ba_': 'Betriebsanweisung', 'p_': 'Pläne'},
        )
        self.categories.setdefault('p_', 'Pläne')
        # None = alle Vorlagen (wie bisher); Liste = explizite Teilauswahl
        if 'selected_template_rel_paths' in last:
            self.selected_template_rel_paths = last.get('selected_template_rel_paths')
        else:
            self.selected_template_rel_paths = None
        self.rules_enabled = last.get('rules_enabled', False)
        self.reuse_lageplan_from_last_export = last.get('reuse_lageplan_from_last_export', False)
        self.signage_rules = last.get('signage_rules', [])
        # Projektverwaltung: zuletzt verwendete Projekte und aktiver Projektpfad
        self.recent_projects = last.get('recent_projects', [])
        self.current_project_path = last.get('last_project_path', None)

    def setup_ui(self):
        """Erstellt und arrangiert alle UI-Elemente im Hauptfenster."""
        self.resize(800, 700)
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        tab_widget = QTabWidget()
        main_layout.addWidget(tab_widget)
        self.tab_widget = tab_widget

        # Tab 1: Hauptsteuerung
        controls_tab = QWidget()
        tab_widget.addTab(controls_tab, "Hauptsteuerung")
        self.setup_controls_tab(controls_tab)

        # Tab 2: Vorlagen-Auswahl
        templates_tab = QWidget()
        tab_widget.addTab(templates_tab, "Vorlagen")
        self.vorlagen_tab_index = tab_widget.indexOf(templates_tab)
        self.setup_templates_tab(templates_tab)

        media_sizes_tab = QWidget()
        tab_widget.addTab(media_sizes_tab, "Bilder & QR-Größen")
        self.media_sizes_tab_index = tab_widget.indexOf(media_sizes_tab)
        self.setup_media_sizes_tab(media_sizes_tab)

        # Tab: Regeln (optionale Schilder)
        rules_tab = QWidget()
        tab_widget.addTab(rules_tab, "Regeln")
        self.rules_tab_index = tab_widget.indexOf(rules_tab)
        self.setup_rules_tab(rules_tab)

        # Tab 4: Kategorien
        categories_tab = QWidget()
        tab_widget.addTab(categories_tab, "Kategorien")
        self.setup_categories_tab(categories_tab)

        # Tab 5: Bilder-Vorschau
        bilder_vorschau_tab = QWidget()
        tab_widget.addTab(bilder_vorschau_tab, "Bilder-Vorschau")
        self.bilder_vorschau_tab_index = tab_widget.indexOf(bilder_vorschau_tab)
        self.setup_bilder_vorschau_tab(bilder_vorschau_tab)

        # Tab 6: PDF
        pdf_tab = QWidget()
        tab_widget.addTab(pdf_tab, "PDF")
        self.pdf_tab_index = tab_widget.indexOf(pdf_tab)
        self.setup_pdf_tab(pdf_tab)

        # Tab 7: Logs
        logs_tab = QWidget()
        tab_widget.addTab(logs_tab, "Logs")
        tab_widget.currentChanged.connect(self._on_tab_changed)
        self.setup_logs_tab(logs_tab)

        self.setup_menu()
        self.apply_theme(self.settings.get('theme', 'Light'))

        for key in self.paths:
            self.validate_path(getattr(self, f"{key}_edit"))

        self.show_excel_data()

        # Zeige Versions-Info in den Logs
        try:
            import PySide6
            pyside_version = PySide6.__version__
        except ImportError:
            pyside_version = "Nicht verfügbar"

        self.log_text.append(
            f'<div style="color: #666; text-align: center; padding: 10px; border: 1px solid #ccc; margin: 10px 0; background-color: #f9f9f9;">'
            f'<strong>DocxTpl Automatisierung v{VERSION}</strong><br>'
            f'Build: {BUILD_DATE}<br>'
            f'Python: {sys.version.split()[0]} | PySide6: {pyside_version}</div>'
        )

    def setup_controls_tab(self, tab):
        """Erstellt den Inhalt des 'Hauptsteuerung'-Tabs."""
        layout = QVBoxLayout(tab)
        path_map = {
            'excel_path': 'Excel-Datei',
            'vorlagen_ordner': 'Vorlagen-Ordner',
            'bilder_ordner': 'Bilder-Ordner (optional)',
            'export_ordner': 'Export-Ordner'
        }
        for key, name in path_map.items():
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{name}:"))
            path_edit = QLineEdit(self.paths.get(key, ''))
            path_edit.editingFinished.connect(lambda k=key, p=path_edit: self.on_path_change(k, p.text()))
            path_edit.textChanged.connect(lambda text, widget=path_edit: self.validate_path(widget))
            setattr(self, f"{key}_edit", path_edit)
            row.addWidget(path_edit)
            browse_btn = QPushButton("...")
            browse_btn.clicked.connect(lambda c, k=key: self.browse(k))
            row.addWidget(browse_btn)
            if key == 'export_ordner':
                layout.addLayout(row)
                proj_row = QHBoxLayout()
                proj_row.addWidget(QLabel("Projektname (Export-Ordner):"))
                self.projekt_name_override_edit = QLineEdit(self.settings.get('projekt_name_override', ''))
                self.projekt_name_override_edit.setPlaceholderText("Leer = aus Excel")
                self.projekt_name_override_edit.editingFinished.connect(self.save_all_settings)
                proj_row.addWidget(self.projekt_name_override_edit)
                layout.addLayout(proj_row)
                continue
            # Excel-Icon-Button nur beim Excel-Pfad
            if key == 'excel_path':
                excel_icon_btn = QPushButton()
                excel_icon_btn.setIcon(QIcon(resource_path('Pictures/excel_icon.png')))
                excel_icon_btn.setToolTip("Excel-Datei öffnen")
                excel_icon_btn.setFixedWidth(32)
                excel_icon_btn.clicked.connect(self.open_excel_file)
                row.addWidget(excel_icon_btn)
            layout.addLayout(row)

        settings_layout = QHBoxLayout()
        setting_map = {
            'header_row': ('Header-Zeile', (1, 100), 3),
            'svg_scale': ('SVG-Skala', (1, 10), 3),
            'png_compression': ('PNG-Komp. (0=Max)', (-1, 100), -1)
        }
        for key, (name, r, default) in setting_map.items():
            settings_layout.addWidget(QLabel(name))
            spin = QSpinBox()
            spin.setRange(*r)
            spin.setValue(self.settings.get(key, default))
            spin.valueChanged.connect(self.save_all_settings)
            setattr(self, f"{key}_spin", spin)
            settings_layout.addWidget(spin)
        layout.addLayout(settings_layout)

        # Layout für UTC-Format
        utc_format_layout = QHBoxLayout()
        utc_format_layout.addWidget(QLabel("Zeitstempel-Format (datetime_utc):"))
        self.datetime_utc_format_edit = QLineEdit(self.settings.get('datetime_utc_format'))
        self.datetime_utc_format_edit.editingFinished.connect(self.save_all_settings)
        utc_format_layout.addWidget(self.datetime_utc_format_edit)
        layout.addLayout(utc_format_layout)

        # Versions-Info und Pfad-Normalisierung
        version_layout = QHBoxLayout()
        version_layout.addWidget(QLabel(f"Version: {VERSION}"))
        version_layout.addWidget(QLabel(f"Build: {BUILD_DATE}"))
        version_layout.addStretch()
        # Button: Pfade normalisieren für Team-Portabilität
        normalize_btn = QPushButton("🔄 Pfade normalisieren")
        normalize_btn.setToolTip(
            f"Wandelt den lokalen Benutzerordner unter …\\Users\\… in '{{username}}' um,\n"
            "damit Einstellungen/Projekte auf allen Rechnern geteilt werden können."
        )
        normalize_btn.clicked.connect(self.normalize_all_paths)
        version_layout.addWidget(normalize_btn)
        denorm_btn = QPushButton("Lokalen Benutzer einsetzen")
        denorm_btn.setToolTip(
            f"Ersetzt '{{username}}' in allen Pfaden durch '{self.current_username}' "
            "(nach Import oder „Pfade normalisieren“)."
        )
        denorm_btn.clicked.connect(self.denormalize_all_paths)
        version_layout.addWidget(denorm_btn)
        layout.addLayout(version_layout)

        layout.addWidget(QFrame(frameShape=QFrame.HLine))
        self.table = QTableView()
        layout.addWidget(self.table)
        self.progress = QProgressBar()
        layout.addWidget(self.progress)
        self.current_file_label = QLabel("Bereit zum Starten...")
        self.current_file_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.current_file_label)

        self.open_export_folder_btn = QPushButton("Export-Ordner öffnen")
        self.open_export_folder_btn.clicked.connect(self.open_export_folder)
        self.open_export_folder_btn.setVisible(False)
        layout.addWidget(self.open_export_folder_btn)

        self.export_as_pdf_check = QCheckBox("Word-Dokumente nach Generierung in PDF umwandeln")
        self.export_as_pdf_check.setChecked(self.settings.get('export_as_pdf', False))
        self.export_as_pdf_check.stateChanged.connect(self.save_all_settings)
        layout.addWidget(self.export_as_pdf_check)

        btn_row = QHBoxLayout()
        self.dry_run_btn = QPushButton("Konfiguration prüfen")
        self.preview_btn = QPushButton("Vorschau")
        self.preview_btn.setToolTip(
            "Erzeugt pro ausgewählter Vorlage ein Beispieldokument mit der ersten Excel-Datenzeile "
            "(Ordner …_Vorschau unter dem Export-Pfad)."
        )
        self.start_btn = QPushButton("Start")
        self.close_btn = QPushButton("Schließen")

        self.dry_run_btn.clicked.connect(self.start_dry_run)
        self.preview_btn.clicked.connect(self.start_preview_batch)
        self.start_btn.clicked.connect(self.start)
        self.close_btn.clicked.connect(self.handle_close_or_cancel)

        btn_row.addWidget(self.dry_run_btn)
        btn_row.addWidget(self.preview_btn)
        btn_row.addWidget(self.start_btn)
        btn_row.addWidget(self.close_btn)
        layout.addLayout(btn_row)

    def setup_templates_tab(self, tab):
        """Tab: Welche Word-Vorlagen (anlagen/allgemein) für Export und Trockenlauf genutzt werden."""
        layout = QVBoxLayout(tab)
        hint_top = QLabel(
            "Die Pfade zu Excel und Vorlagen-Ordner legst du unter <b>Hauptsteuerung</b> fest. "
            "Hier wählst du die .docx-Dateien für den nächsten Lauf."
        )
        hint_top.setWordWrap(True)
        hint_top.setStyleSheet("color: #555;")
        layout.addWidget(hint_top)

        vorlagen_pick = QGroupBox("Vorlagen für diesen Lauf")
        vorlagen_pick_layout = QVBoxLayout(vorlagen_pick)
        self._template_hint_label = QLabel(
            "Nur Unterordner …\\anlagen\\ und …\\allgemein\\ werden verwendet."
        )
        self._template_hint_label.setWordWrap(True)
        self._template_hint_label.setStyleSheet("color: #555;")
        vorlagen_pick_layout.addWidget(self._template_hint_label)
        tpl_btn_row = QHBoxLayout()
        refresh_tpl_btn = QPushButton("Vorlagen einlesen")
        refresh_tpl_btn.setToolTip("Liste aus dem Vorlagen-Ordner neu aufbauen")
        refresh_tpl_btn.clicked.connect(self.refresh_template_checkboxes)
        tpl_btn_row.addWidget(refresh_tpl_btn)
        all_tpl_btn = QPushButton("Alle auswählen")
        all_tpl_btn.clicked.connect(self._template_select_all)
        tpl_btn_row.addWidget(all_tpl_btn)
        none_tpl_btn = QPushButton("Keine auswählen")
        none_tpl_btn.clicked.connect(self._template_select_none)
        tpl_btn_row.addWidget(none_tpl_btn)
        tpl_btn_row.addStretch()
        vorlagen_pick_layout.addLayout(tpl_btn_row)
        self._template_preview_status = QLabel("")
        self._template_preview_status.setStyleSheet("color: #444; font-size: 12px;")
        self._template_preview_status.setWordWrap(True)
        self._template_preview_status.setVisible(False)
        self._template_preview_progress = QProgressBar()
        self._template_preview_progress.setVisible(False)
        self._template_preview_progress.setMinimumHeight(16)
        self._template_preview_progress.setTextVisible(True)
        self._template_preview_progress.setFormat("%p%")
        vorlagen_pick_layout.addWidget(self._template_preview_status)
        vorlagen_pick_layout.addWidget(self._template_preview_progress)
        self._template_row_preview_active = False
        tpl_scroll = QScrollArea()
        tpl_scroll.setWidgetResizable(True)
        tpl_scroll.setMinimumHeight(200)
        tpl_scroll_content = QWidget()
        self._template_list_layout = QVBoxLayout(tpl_scroll_content)
        self._template_list_layout.addStretch()
        tpl_scroll.setWidget(tpl_scroll_content)
        vorlagen_pick_layout.addWidget(tpl_scroll)
        layout.addWidget(vorlagen_pick)
        layout.addStretch()

        self.refresh_template_checkboxes()

    def setup_media_sizes_tab(self, tab):
        """Tab: Bild-/QR-Platzhalter aus Vorlagen, Größen aus Excel, Vorschau, Schreiben nach Blatt 2."""
        layout = QVBoxLayout(tab)
        intro = QLabel(
            "Standard: Gruppierung nach Basis-Platzhalter – eine Größenänderung gilt für alle gelisteten "
            "Vorlagen. Referenz und Größe wie erste Anlagenzeile + Fallback (Blatt 2). "
            "Ohne Häkchen: jede Vorlage einzeln. Speichern schreibt nach Blatt 2."
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color: #555;")
        layout.addWidget(intro)

        btn_row = QHBoxLayout()
        self.media_sizes_refresh_btn = QPushButton("Aktualisieren")
        self.media_sizes_refresh_btn.setToolTip("Excel und Vorlagen neu einlesen")
        self.media_sizes_refresh_btn.clicked.connect(self.refresh_media_sizes_table)
        btn_row.addWidget(self.media_sizes_refresh_btn)
        self.media_sizes_save_btn = QPushButton("Änderungen in Excel speichern")
        self.media_sizes_save_btn.setToolTip("Nur geänderte cm-Werte → Blatt 2, Spalte B/C")
        self.media_sizes_save_btn.clicked.connect(self._save_media_sizes_excel)
        btn_row.addWidget(self.media_sizes_save_btn)
        self.media_sizes_group_check = QCheckBox("Nach Basisplatzhalter gruppieren (Standard)")
        self.media_sizes_group_check.setChecked(True)
        self.media_sizes_group_check.setToolTip(
            "Eine Zeile pro Platzhalter (z. B. ba_logo_img). Größe und Referenz gelten für alle "
            "gelisteten Vorlagen gleich – Speichern schreibt eine Textmarke in Blatt 2."
        )
        self.media_sizes_group_check.toggled.connect(self._on_media_sizes_group_toggled)
        btn_row.addWidget(self.media_sizes_group_check)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self.media_sizes_status_label = QLabel("")
        self.media_sizes_status_label.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(self.media_sizes_status_label)

        self.media_sizes_detail_table = QTableWidget()
        self.media_sizes_detail_table.setColumnCount(8)
        self.media_sizes_detail_table.setHorizontalHeaderLabels(
            [
                "Dokument",
                "Basis-Platzhalter",
                "Typ",
                "Referenz",
                "Vorschau",
                "Größen-Textmarke",
                "Wert (cm)",
                "Quelle",
            ]
        )
        self.media_sizes_detail_table.horizontalHeader().setStretchLastSection(True)
        self.media_sizes_detail_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.media_sizes_detail_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.media_sizes_detail_table.hide()
        layout.addWidget(self.media_sizes_detail_table)

        self.media_sizes_group_table = QTableWidget()
        self.media_sizes_group_table.setColumnCount(8)
        self.media_sizes_group_table.setHorizontalHeaderLabels(
            [
                "Basis-Platzhalter",
                "Typ",
                "Referenz",
                "Vorschau",
                "Größen-Textmarke",
                "Wert (cm)",
                "Quelle",
                "Dokumente (Vorlagen)",
            ]
        )
        self.media_sizes_group_table.horizontalHeader().setStretchLastSection(True)
        self.media_sizes_group_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.media_sizes_group_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.media_sizes_group_table)

        self._media_sizes_detail_rows = []
        self._media_sizes_row_meta = []
        self._media_sizes_basis_row_meta = []

    def _on_media_sizes_group_toggled(self, checked: bool):
        self.media_sizes_detail_table.setVisible(not checked)
        self.media_sizes_group_table.setVisible(checked)
        if checked and self._media_sizes_detail_rows:
            self._fill_media_sizes_basis_group_table()
        elif not checked and self._media_sizes_detail_rows:
            self._fill_media_sizes_detail_table(self._media_sizes_detail_rows)

    def _media_sizes_preview_pixmap(self, kind: str, reference: str, bilder_ordner: str) -> QPixmap | None:
        ref = (reference or "").strip()
        if not ref:
            pix = QPixmap(56, 56)
            pix.fill(Qt.lightGray)
            return pix
        if kind == "Bild" and bilder_ordner:
            full = os.path.normpath(os.path.join(bilder_ordner, ref))
            if not os.path.isfile(full):
                return None
            ext = os.path.splitext(full)[1].lower()
            if ext == ".svg":
                try:
                    from core.utils import svg_to_png_file_pyside

                    fd, tmp_png = tempfile.mkstemp(suffix=".png")
                    os.close(fd)
                    try:
                        if svg_to_png_file_pyside(full, tmp_png, scale=2, compression=-1):
                            pix = QPixmap(tmp_png)
                        else:
                            return None
                    finally:
                        try:
                            os.unlink(tmp_png)
                        except OSError:
                            pass
                except Exception:
                    return None
            else:
                pix = QPixmap(full)
            if pix.isNull():
                return None
            return pix.scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        if kind == "QR-Link":
            try:
                import qrcode

                qr = qrcode.make(ref)
                buf = io.BytesIO()
                qr.save(buf, format="PNG")
                buf.seek(0)
                img = QImage.fromData(buf.read())
                if img.isNull():
                    return None
                return QPixmap.fromImage(img).scaled(
                    56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
            except Exception:
                return None
        return None

    def _fill_media_sizes_detail_table(self, rows: list):
        self._media_sizes_row_meta = []
        table = self.media_sizes_detail_table
        table.setRowCount(0)
        bilder_ordner = (self.paths.get("bilder_ordner") or "").strip()
        for i, r in enumerate(rows):
            table.insertRow(i)
            table.setItem(i, 0, QTableWidgetItem(r.get("template_rel", "")))
            table.setItem(i, 1, QTableWidgetItem(r.get("base_key", "")))
            table.setItem(i, 2, QTableWidgetItem(r.get("kind", "")))
            ref_full = r.get("reference") or ""
            ref_disp = ref_full if len(ref_full) <= 100 else ref_full[:97] + "…"
            it_ref = QTableWidgetItem(ref_disp)
            it_ref.setToolTip(ref_full if len(ref_full) > 100 else "")
            table.setItem(i, 3, it_ref)
            table.setItem(i, 5, QTableWidgetItem(r.get("size_key", "")))
            spin = QDoubleSpinBox()
            spin.setRange(0.1, 999.0)
            spin.setDecimals(2)
            spin.setValue(float(r.get("cm_value", 0)))
            spin.setEnabled(not r.get("cm_variable"))
            table.setCellWidget(i, 6, spin)
            src = r.get("cm_source", "")
            it_src = QTableWidgetItem(src)
            if r.get("cm_variable"):
                it_src.setToolTip("Werte in Blatt 1 unterscheiden sich – bitte in Excel anpassen.")
            table.setItem(i, 7, it_src)

            lbl = QLabel()
            lbl.setAlignment(Qt.AlignCenter)
            pix = self._media_sizes_preview_pixmap(r.get("kind", ""), ref_full, bilder_ordner)
            if pix is not None and not pix.isNull():
                lbl.setPixmap(pix)
            else:
                lbl.setText("—")
                if r.get("kind") == "Bild" and ref_full and bilder_ordner:
                    lbl.setToolTip("Bilddatei im Bilder-Ordner nicht gefunden.")
            lbl.setFixedSize(60, 60)
            table.setCellWidget(i, 4, lbl)
            table.setRowHeight(i, 68)

            self._media_sizes_row_meta.append(
                {
                    "size_key": r.get("size_key", ""),
                    "cm_variable": bool(r.get("cm_variable")),
                    "orig_cm": float(r.get("cm_value", 0)),
                }
            )

    def _fill_media_sizes_basis_group_table(self):
        detail = self._media_sizes_detail_rows
        groups = gruppier_nach_basis_platzhalter(detail)
        t = self.media_sizes_group_table
        t.setRowCount(0)
        self._media_sizes_basis_row_meta = []
        bilder_ordner = (self.paths.get("bilder_ordner") or "").strip()
        for i, g in enumerate(groups):
            t.insertRow(i)
            base_key = g.get("base_key", "")
            t.setItem(i, 0, QTableWidgetItem(base_key))
            t.setItem(i, 1, QTableWidgetItem(g.get("kind", "")))

            ref_full = g.get("reference") or ""
            ref_disp = ref_full if len(ref_full) <= 100 else ref_full[:97] + "…"
            it_ref = QTableWidgetItem(ref_disp)
            tip_ref = ref_full
            if g.get("reference_conflict"):
                tip_ref = (tip_ref + "\n\n" if tip_ref else "") + "Hinweis: Unterschiedliche Referenzwerte je Vorlage – bitte Excel prüfen."
            it_ref.setToolTip(tip_ref)
            t.setItem(i, 2, it_ref)

            lbl = QLabel()
            lbl.setAlignment(Qt.AlignCenter)
            pix = self._media_sizes_preview_pixmap(g.get("kind", ""), ref_full, bilder_ordner)
            if pix is not None and not pix.isNull():
                lbl.setPixmap(pix)
            else:
                lbl.setText("—")
            lbl.setFixedSize(60, 60)
            t.setCellWidget(i, 3, lbl)

            t.setItem(i, 4, QTableWidgetItem(g.get("size_key", "")))

            spin = QDoubleSpinBox()
            spin.setRange(0.1, 999.0)
            spin.setDecimals(2)
            spin.setValue(float(g.get("cm_value", 0)))
            spin.setEnabled(not g.get("cm_variable"))
            t.setCellWidget(i, 5, spin)

            it_src = QTableWidgetItem(g.get("cm_source", ""))
            if g.get("cm_variable"):
                it_src.setToolTip("Werte in Excel weichen ab – bitte zuerst in Excel bereinigen oder Detailansicht nutzen.")
            t.setItem(i, 6, it_src)

            doc_compact = g.get("templates_compact", "")
            it_doc = QTableWidgetItem(doc_compact)
            it_doc.setToolTip(g.get("templates_tooltip") or doc_compact)
            t.setItem(i, 7, it_doc)

            t.setRowHeight(i, 68)

            self._media_sizes_basis_row_meta.append(
                {
                    "size_key": g.get("size_key", ""),
                    "cm_variable": bool(g.get("cm_variable")),
                    "orig_cm": float(g.get("cm_value", 0)),
                }
            )

    def refresh_media_sizes_table(self):
        vo = (self.paths.get("vorlagen_ordner") or "").strip()
        excel_path = (self.paths.get("excel_path") or "").strip()
        hr = self.settings.get("header_row", 3)
        sel = self._template_paths_for_worker()

        if not excel_path or not os.path.isfile(excel_path):
            self._media_sizes_detail_rows = []
            if self.media_sizes_group_check.isChecked():
                self._fill_media_sizes_basis_group_table()
            else:
                self._fill_media_sizes_detail_table([])
            self.media_sizes_status_label.setText("Bitte eine Excel-Datei unter „Hauptsteuerung“ wählen.")
            return
        if not vo or not os.path.isdir(vo):
            self._media_sizes_detail_rows = []
            if self.media_sizes_group_check.isChecked():
                self._fill_media_sizes_basis_group_table()
            else:
                self._fill_media_sizes_detail_table([])
            self.media_sizes_status_label.setText("Bitte einen Vorlagen-Ordner wählen.")
            return
        if sel == []:
            self._media_sizes_detail_rows = []
            if self.media_sizes_group_check.isChecked():
                self._fill_media_sizes_basis_group_table()
            else:
                self._fill_media_sizes_detail_table([])
            self.media_sizes_status_label.setText("Keine Vorlagen ausgewählt (Tab „Vorlagen“).")
            return

        try:
            rows = sammle_media_groessen_uebersicht(vo, excel_path, hr, sel, None)
        except Exception as e:
            self._media_sizes_detail_rows = []
            if self.media_sizes_group_check.isChecked():
                self._fill_media_sizes_basis_group_table()
            else:
                self._fill_media_sizes_detail_table([])
            self.media_sizes_status_label.setText(f"Fehler beim Einlesen: {e}")
            QMessageBox.warning(self, "Bilder & QR-Größen", str(e))
            return

        self._media_sizes_detail_rows = rows
        n_basis = len(gruppier_nach_basis_platzhalter(rows)) if rows else 0
        self.media_sizes_status_label.setText(
            f"{len(rows)} Einträge (Vorlagen × Platzhalter), {n_basis} Basis-Platzhalter."
        )
        if self.media_sizes_group_check.isChecked():
            self._fill_media_sizes_basis_group_table()
        else:
            self._fill_media_sizes_detail_table(rows)

    def _save_media_sizes_excel(self):
        excel_path = (self.paths.get("excel_path") or "").strip()
        if not excel_path or not os.path.isfile(excel_path):
            QMessageBox.warning(self, "Excel", "Keine gültige Excel-Datei.")
            return

        updates = []
        if self.media_sizes_group_check.isChecked():
            table = self.media_sizes_group_table
            for i in range(table.rowCount()):
                if i >= len(self._media_sizes_basis_row_meta):
                    break
                meta = self._media_sizes_basis_row_meta[i]
                if meta.get("cm_variable"):
                    continue
                w = table.cellWidget(i, 5)
                if isinstance(w, QDoubleSpinBox) and abs(w.value() - meta["orig_cm"]) > 1e-9:
                    updates.append((meta["size_key"], w.value()))
        else:
            table = self.media_sizes_detail_table
            for i in range(table.rowCount()):
                if i >= len(self._media_sizes_row_meta):
                    break
                meta = self._media_sizes_row_meta[i]
                if meta.get("cm_variable"):
                    continue
                w = table.cellWidget(i, 6)
                if isinstance(w, QDoubleSpinBox) and abs(w.value() - meta["orig_cm"]) > 1e-9:
                    updates.append((meta["size_key"], w.value()))

        if not updates:
            QMessageBox.information(self, "Excel", "Keine geänderten Größen zum Speichern.")
            return

        try:
            schreibe_groessen_nach_excel_blatt2(excel_path, updates)
        except PermissionError:
            QMessageBox.warning(
                self,
                "Excel",
                "Die Datei ist gesperrt (z. B. in Excel geöffnet).",
            )
            return
        except Exception as e:
            QMessageBox.warning(self, "Excel", f"Speichern fehlgeschlagen:\n{e}")
            return

        QMessageBox.information(self, "Excel", f"{len(updates)} Größe(n) wurden in Blatt 2 gespeichert.")
        self.refresh_media_sizes_table()
        self.show_excel_data()

    def setup_rules_tab(self, tab):
        """Tab: Optionale Beschilderungs-Regeln (z. B. Notausgang: Text vs. Bild-Zweig)."""
        layout = QVBoxLayout(tab)
        intro = QLabel(
            "Wenn aktiviert, werden nur für passende Excel-Zeilen die angegebenen Vorlagen erzeugt "
            "(Anlagen-Vorlagen). Ohne Aktivierung bleibt alles wie bisher."
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color: #555;")
        layout.addWidget(intro)

        self.rules_enabled_check = QCheckBox("Regeln für optionale Schilder aktivieren")
        self.rules_enabled_check.setChecked(self.rules_enabled)
        self.rules_enabled_check.stateChanged.connect(self.save_all_settings)
        layout.addWidget(self.rules_enabled_check)

        self.reuse_lageplan_from_last_export_check = QCheckBox(
            "Lagepläne (rl_) aus letztem Export übernehmen"
        )
        self.reuse_lageplan_from_last_export_check.setChecked(self.reuse_lageplan_from_last_export)
        self.reuse_lageplan_from_last_export_check.stateChanged.connect(self.save_all_settings)
        layout.addWidget(self.reuse_lageplan_from_last_export_check)
        lageplan_hint = QLabel(
            "Wenn aktiv, werden Anlagen-Vorlagen mit Präfix „rl_“ nicht neu gerendert, sondern aus dem "
            "zuletzt erfolgreichen Export kopiert (Unterordner wie in „Kategorien“ für rl_, sonst "
            "„Lageplan“ oder „Plan“). Fehlt die Datei, wird wie gewohnt aus der Vorlage erzeugt."
        )
        lageplan_hint.setWordWrap(True)
        lageplan_hint.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(lageplan_hint)

        self._rules_loading_ui = False
        self._rules_prev_selected_row = -2

        split = QSplitter(Qt.Horizontal)
        left_panel = QWidget()
        left_l = QVBoxLayout(left_panel)
        left_l.setContentsMargins(0, 0, 0, 0)
        self.rules_list_widget = QListWidget()
        self.rules_list_widget.setMinimumWidth(170)
        self.rules_list_widget.currentRowChanged.connect(self._on_rules_list_row_changed)
        left_l.addWidget(self.rules_list_widget)
        list_btn_row = QHBoxLayout()
        rules_add_btn = QPushButton("Neue Regel")
        rules_add_btn.clicked.connect(self._rules_add_clicked)
        rules_del_btn = QPushButton("Löschen")
        rules_del_btn.clicked.connect(self._rules_delete_clicked)
        list_btn_row.addWidget(rules_add_btn)
        list_btn_row.addWidget(rules_del_btn)
        left_l.addLayout(list_btn_row)

        self.rules_editor_widget = QWidget()
        box_l = QVBoxLayout(self.rules_editor_widget)
        self.rules_editor_box = QGroupBox("Regel bearbeiten")
        editor_inner = QVBoxLayout(self.rules_editor_box)

        self.rules_rule_enabled_check = QCheckBox("Diese Regel anwenden")
        self.rules_rule_enabled_check.setChecked(True)
        self.rules_rule_enabled_check.stateChanged.connect(self._rules_mark_dirty_save)
        editor_inner.addWidget(self.rules_rule_enabled_check)

        row_name = QHBoxLayout()
        row_name.addWidget(QLabel("Name:"))
        self.rules_rule_name_edit = QLineEdit()
        self.rules_rule_name_edit.setPlaceholderText("z. B. Notausgang")
        self.rules_rule_name_edit.editingFinished.connect(self._rules_mark_dirty_save)
        self.rules_rule_name_edit.editingFinished.connect(self._rules_sync_list_item_title)
        row_name.addWidget(self.rules_rule_name_edit)
        editor_inner.addLayout(row_name)

        ref_row = QHBoxLayout()
        ref_row.addWidget(QLabel("1. Bezugs-Vorlage (Textmarken):"))
        self.rules_reference_template_combo = QComboBox()
        self.rules_reference_template_combo.setMinimumWidth(280)
        self.rules_reference_template_combo.setToolTip(
            "Zuerst die Word-Vorlage wählen, aus der die Textmarken gelesen werden."
        )
        self.rules_reference_template_combo.currentIndexChanged.connect(self._rules_on_reference_template_changed)
        ref_row.addWidget(self.rules_reference_template_combo)
        editor_inner.addLayout(ref_row)

        ref_hint = QLabel(
            "Nach der Auswahl erscheinen die Platzhalter der Vorlage in der Spalte „Textmarke“. "
            "Leer lassen = Excel-Spaltenname ist gleichzeitig der Kontext-Schlüssel."
        )
        ref_hint.setWordWrap(True)
        ref_hint.setStyleSheet("color: #666; font-size: 11px;")
        editor_inner.addWidget(ref_hint)

        editor_inner.addWidget(QLabel("2. Excel-Spalte und zugehörige Textmarke (mind. eine Zeile mit Inhalt):"))
        self.rules_fields_table = QTableWidget(0, 2)
        self.rules_fields_table.setHorizontalHeaderLabels(["Excel-Spalte", "Textmarke"])
        self.rules_fields_table.horizontalHeader().setStretchLastSection(True)
        self.rules_fields_table.setMinimumHeight(120)
        self.rules_fields_table.setMaximumHeight(220)
        editor_inner.addWidget(self.rules_fields_table)
        field_btn_row = QHBoxLayout()
        rules_field_add_btn = QPushButton("Zeile hinzufügen")
        rules_field_add_btn.clicked.connect(self._rules_field_row_add)
        rules_field_del_btn = QPushButton("Zeile entfernen")
        rules_field_del_btn.clicked.connect(self._rules_field_row_remove)
        field_btn_row.addWidget(rules_field_add_btn)
        field_btn_row.addWidget(rules_field_del_btn)
        field_btn_row.addStretch()
        editor_inner.addLayout(field_btn_row)

        cond_row = QHBoxLayout()
        cond_row.addWidget(QLabel("3. Bedingung (für die Zeilen oben, mindestens eine muss passen):"))
        self.rules_condition_combo = QComboBox()
        self.rules_condition_combo.setMinimumWidth(280)
        for val, label in RULE_CONDITION_CHOICES:
            self.rules_condition_combo.addItem(label, val)
        self.rules_condition_combo.currentIndexChanged.connect(self._rules_on_condition_mode_changed)
        cond_row.addWidget(self.rules_condition_combo)
        editor_inner.addLayout(cond_row)

        param_row = QHBoxLayout()
        param_row.addWidget(QLabel("Zahl / Länge:"))
        self.rules_min_len_spin = QSpinBox()
        self.rules_min_len_spin.setRange(0, 9999)
        self.rules_min_len_spin.setValue(3)
        self.rules_min_len_spin.setToolTip("Für „länger als“ und „Mindestlänge (≥)“.")
        self.rules_min_len_spin.valueChanged.connect(self._rules_mark_dirty_save)
        param_row.addWidget(self.rules_min_len_spin)
        param_row.addWidget(QLabel("enthält / gleich:"))
        self.rules_needle_edit = QLineEdit()
        self.rules_needle_edit.setPlaceholderText("Suchtext oder exakter Vergleichswert")
        self.rules_needle_edit.editingFinished.connect(self._rules_mark_dirty_save)
        param_row.addWidget(self.rules_needle_edit)
        editor_inner.addLayout(param_row)

        regex_row = QHBoxLayout()
        regex_row.addWidget(QLabel("Regex:"))
        self.rules_regex_edit = QLineEdit()
        self.rules_regex_edit.setPlaceholderText("z. B. ^[A-Z]{2}-[0-9]+$")
        self.rules_regex_edit.editingFinished.connect(self._rules_mark_dirty_save)
        regex_row.addWidget(self.rules_regex_edit)
        editor_inner.addLayout(regex_row)

        row_img = QHBoxLayout()
        row_img.addWidget(QLabel("Bild-Spalte (optional, Dateiname im Bilder-Ordner):"))
        self.rules_image_col_edit = QLineEdit()
        self.rules_image_col_edit.setPlaceholderText("z. B. notausgang_img")
        self.rules_image_col_edit.editingFinished.connect(self._rules_mark_dirty_save)
        row_img.addWidget(self.rules_image_col_edit)
        editor_inner.addLayout(row_img)

        row_if = QHBoxLayout()
        row_if.addWidget(QLabel("Vorlage wenn Bedingung / Bild-Zweig „ja“:"))
        self.rules_template_if_combo = QComboBox()
        self.rules_template_if_combo.setMinimumWidth(280)
        self.rules_template_if_combo.currentIndexChanged.connect(self._rules_mark_dirty_save)
        row_if.addWidget(self.rules_template_if_combo)
        editor_inner.addLayout(row_if)

        row_else = QHBoxLayout()
        row_else.addWidget(QLabel("Vorlage wenn Bedingung / Bild-Zweig „nein“:"))
        self.rules_template_else_combo = QComboBox()
        self.rules_template_else_combo.setMinimumWidth(280)
        self.rules_template_else_combo.currentIndexChanged.connect(self._rules_mark_dirty_save)
        row_else.addWidget(self.rules_template_else_combo)
        editor_inner.addLayout(row_else)

        hint = QLabel(
            "Beide Ziel-Vorlagen müssen unter „Vorlagen“ angehakt sein. "
            "Die Bezugs-Vorlage dient nur zum Auslesen der Textmarken und kann dieselbe Datei sein wie IF oder ELSE."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #666; font-size: 11px;")
        editor_inner.addWidget(hint)

        box_l.addWidget(self.rules_editor_box)
        box_l.addStretch()

        split.addWidget(left_panel)
        split.addWidget(self.rules_editor_widget)
        split.setStretchFactor(1, 1)
        layout.addWidget(split)
        layout.addStretch()

        self._rules_on_condition_mode_changed()
        self._apply_signage_rules_to_ui()

    def _rules_mark_dirty_save(self, *args):
        if getattr(self, "_rules_loading_ui", False):
            return
        self.save_all_settings()

    def _rules_sync_list_item_title(self):
        row = self.rules_list_widget.currentRow()
        if row < 0:
            return
        it = self.rules_list_widget.item(row)
        if it:
            name = (self.rules_rule_name_edit.text() or f"Regel {row + 1}").strip()
            it.setText(name)

    def _set_rule_editor_enabled(self, enabled: bool):
        if hasattr(self, "rules_editor_box"):
            self.rules_editor_box.setEnabled(enabled)

    def _default_signage_rule_dict(self):
        return {
            "id": "",
            "name": "Neue Regel",
            "enabled": True,
            "reference_template": "",
            "templates_if": [],
            "templates_else": [],
            "when": {
                "text_condition": {
                    "mode": "not_empty",
                    "min": 3,
                    "needle": "",
                    "equals_value": "",
                    "regex_pattern": "",
                    "fields": [{"column": "", "textmarke": ""}],
                }
            },
            "branch_on_image": {"column": ""},
        }

    def _ensure_signage_rules_length(self, n: int):
        while len(self.signage_rules) < n:
            self.signage_rules.append(self._default_signage_rule_dict())

    def _pick_template_combo(self, cb: QComboBox, rel: str):
        if not rel:
            cb.setCurrentIndex(0)
            return
        idx = cb.findData(rel)
        if idx >= 0:
            cb.setCurrentIndex(idx)
        else:
            cb.addItem(rel, rel)
            cb.setCurrentIndex(cb.count() - 1)

    def _on_rules_list_row_changed(self, row: int):
        if getattr(self, "_rules_loading_ui", False):
            return
        prev = getattr(self, "_rules_prev_selected_row", -2)
        if prev >= 0 and prev != row:
            self._ensure_signage_rules_length(prev + 1)
            self.signage_rules[prev] = self._collect_rule_dict_from_editor()
            it = self.rules_list_widget.item(prev)
            if it:
                it.setText(self.signage_rules[prev].get("name") or f"Regel {prev + 1}")
        self._rules_prev_selected_row = row
        if row < 0:
            self._set_rule_editor_enabled(False)
            return
        self._set_rule_editor_enabled(True)
        self._ensure_signage_rules_length(row + 1)
        self._load_rule_editor_from_index(row)

    def _rules_add_clicked(self):
        row = self.rules_list_widget.currentRow()
        if row >= 0:
            self._ensure_signage_rules_length(row + 1)
            self.signage_rules[row] = self._collect_rule_dict_from_editor()
        self.signage_rules.append(self._default_signage_rule_dict())
        self.rules_list_widget.addItem(self.signage_rules[-1]["name"])
        self.rules_list_widget.setCurrentRow(self.rules_list_widget.count() - 1)
        self.save_all_settings()

    def _rules_delete_clicked(self):
        row = self.rules_list_widget.currentRow()
        if row < 0 or row >= len(self.signage_rules):
            return
        del self.signage_rules[row]
        self.rules_list_widget.takeItem(row)
        if self.rules_list_widget.count():
            self.rules_list_widget.setCurrentRow(min(row, self.rules_list_widget.count() - 1))
        else:
            self._rules_prev_selected_row = -1
            self._set_rule_editor_enabled(False)
        self.save_all_settings()

    def _rules_field_row_add(self):
        self._rules_add_fields_table_row("", "")
        self._rules_mark_dirty_save()

    def _rules_field_row_remove(self):
        r = self.rules_fields_table.currentRow()
        if r < 0 and self.rules_fields_table.rowCount() > 0:
            r = self.rules_fields_table.rowCount() - 1
        if r >= 0:
            self.rules_fields_table.removeRow(r)
        self._rules_mark_dirty_save()

    def _rules_add_fields_table_row(self, column: str, textmarke: str):
        r = self.rules_fields_table.rowCount()
        self.rules_fields_table.insertRow(r)
        le = QLineEdit()
        le.setText(column)
        le.setPlaceholderText("Excel-Überschrift")
        le.editingFinished.connect(self._rules_mark_dirty_save)
        self.rules_fields_table.setCellWidget(r, 0, le)
        cb = QComboBox()
        cb.setEditable(True)
        cb.lineEdit().setPlaceholderText("Textmarke aus Vorlage oder leer")
        cb.setCurrentText(textmarke)
        cb.currentTextChanged.connect(self._rules_mark_dirty_save)
        self.rules_fields_table.setCellWidget(r, 1, cb)
        self._rules_fill_textmarke_combo_items(cb)

    def _rules_fill_textmarke_combo_items(self, cb: QComboBox):
        vo = self.paths.get("vorlagen_ordner", "").strip()
        rel = self.rules_reference_template_combo.currentData() or ""
        cur = cb.currentText()
        cb.blockSignals(True)
        cb.clear()
        marken = []
        if vo and rel:
            abs_p = os.path.normpath(os.path.join(vo, rel))
            marken = liste_textmarken_aus_docx(abs_p)
        for m in marken:
            cb.addItem(m)
        cb.setEditText(cur)
        cb.blockSignals(False)

    def _rules_refresh_all_field_textmarken(self):
        for r in range(self.rules_fields_table.rowCount()):
            w = self.rules_fields_table.cellWidget(r, 1)
            if isinstance(w, QComboBox):
                self._rules_fill_textmarke_combo_items(w)

    def _rules_on_reference_template_changed(self):
        self._rules_refresh_all_field_textmarken()
        self._rules_mark_dirty_save()

    def _rules_on_condition_mode_changed(self, *_):
        if not hasattr(self, "rules_condition_combo"):
            return
        mode = self.rules_condition_combo.currentData()
        show_len = mode in ("length_gt", "length_gte")
        show_needle = mode in ("contains", "equals", "equals_ignorecase")
        show_rx = mode == "regex"
        self.rules_min_len_spin.setVisible(show_len)
        self.rules_needle_edit.setVisible(show_needle)
        self.rules_regex_edit.setVisible(show_rx)
        self._rules_mark_dirty_save()

    def _clear_rule_fields_table(self):
        self.rules_fields_table.setRowCount(0)

    def _load_rule_editor_from_index(self, row: int):
        self._rules_loading_ui = True
        try:
            rule = (
                self.signage_rules[row]
                if row < len(self.signage_rules)
                else self._default_signage_rule_dict()
            )
            self.rules_rule_enabled_check.setChecked(bool(rule.get("enabled", True)))
            self.rules_rule_name_edit.setText(str(rule.get("name", "")))

            when = rule.get("when") or {}
            mode = "not_empty"
            min_v = 3
            needle = ""
            eq_v = ""
            rx_v = ""
            fields = [{"column": "", "textmarke": ""}]

            if "text_condition" in when:
                tc = when.get("text_condition") or {}
                mode = str(tc.get("mode") or "not_empty").strip().lower()
                min_v = int(tc.get("min", tc.get("min_len", 3)))
                needle = str(tc.get("needle", tc.get("contains", "")))
                eq_v = str(tc.get("equals_value", tc.get("equals", "")))
                rx_v = str(tc.get("regex_pattern", tc.get("regex", "")))
                fields = tc.get("fields")
                if not fields:
                    cols = tc.get("columns") or []
                    if isinstance(cols, str):
                        cols = [c.strip() for c in cols.split(",") if c.strip()]
                    fields = [{"column": c, "textmarke": ""} for c in cols]
                if not fields:
                    fields = [{"column": "", "textmarke": ""}]
            elif "text_fields_any_length_gt" in when:
                old = when.get("text_fields_any_length_gt") or {}
                mode = "length_gt"
                min_v = int(old.get("min", 3))
                cols = old.get("columns") or []
                if isinstance(cols, str):
                    cols = [c.strip() for c in cols.split(",") if c.strip()]
                fields = [{"column": c, "textmarke": ""} for c in cols] or [
                    {"column": "", "textmarke": ""}
                ]

            idx = self.rules_condition_combo.findData(mode)
            self.rules_condition_combo.setCurrentIndex(idx if idx >= 0 else 0)
            self.rules_min_len_spin.setValue(min_v)
            if mode == "contains":
                self.rules_needle_edit.setText(needle)
            elif mode in ("equals", "equals_ignorecase"):
                self.rules_needle_edit.setText(eq_v)
            else:
                self.rules_needle_edit.setText("")
            self.rules_regex_edit.setText(rx_v)

            img_col = (rule.get("branch_on_image") or {}).get("column", "")
            self.rules_image_col_edit.setText(str(img_col or ""))

            self.refresh_rules_template_combos()
            ref = str(rule.get("reference_template", "") or "")
            self._pick_template_combo(self.rules_reference_template_combo, ref)
            tif_list = rule.get("templates_if") or []
            tel_list = rule.get("templates_else") or []
            self._pick_template_combo(
                self.rules_template_if_combo, tif_list[0] if tif_list else ""
            )
            self._pick_template_combo(
                self.rules_template_else_combo, tel_list[0] if tel_list else ""
            )

            self._clear_rule_fields_table()
            for f in fields:
                if isinstance(f, str):
                    self._rules_add_fields_table_row(f.strip(), "")
                else:
                    fd = f or {}
                    self._rules_add_fields_table_row(
                        str(fd.get("column", "")),
                        str(fd.get("textmarke", "")),
                    )
            if self.rules_fields_table.rowCount() == 0:
                self._rules_add_fields_table_row("", "")

            self._rules_on_condition_mode_changed()
        finally:
            self._rules_loading_ui = False
        self._rules_refresh_all_field_textmarken()

    def _collect_rule_dict_from_editor(self) -> dict:
        mode = self.rules_condition_combo.currentData() or "not_empty"
        needle = (self.rules_needle_edit.text() or "").strip()
        eq_val = needle if mode in ("equals", "equals_ignorecase") else ""
        needle_val = needle if mode == "contains" else ""
        fields = []
        for r in range(self.rules_fields_table.rowCount()):
            le_w = self.rules_fields_table.cellWidget(r, 0)
            cb_w = self.rules_fields_table.cellWidget(r, 1)
            col = le_w.text().strip() if isinstance(le_w, QLineEdit) else ""
            tm = ""
            if isinstance(cb_w, QComboBox):
                tm = cb_w.currentText().strip()
            if not col and not tm:
                continue
            fields.append({"column": col, "textmarke": tm})
        if not fields:
            fields = [{"column": "", "textmarke": ""}]

        tif = self.rules_template_if_combo.currentData() or self.rules_template_if_combo.currentText().strip()
        tel = self.rules_template_else_combo.currentData() or self.rules_template_else_combo.currentText().strip()
        ref = self.rules_reference_template_combo.currentData() or self.rules_reference_template_combo.currentText().strip()

        row = self.rules_list_widget.currentRow()
        rid = ""
        if 0 <= row < len(self.signage_rules):
            rid = str((self.signage_rules[row] or {}).get("id") or "")
        if not rid:
            rid = f"rule_{row + 1}" if row >= 0 else "rule_1"

        return {
            "id": rid,
            "name": (self.rules_rule_name_edit.text() or f"Regel {row + 1}").strip()
            or f"Regel {row + 1}",
            "enabled": self.rules_rule_enabled_check.isChecked(),
            "reference_template": ref,
            "templates_if": [tif] if tif else [],
            "templates_else": [tel] if tel else [],
            "when": {
                "text_condition": {
                    "mode": mode,
                    "min": int(self.rules_min_len_spin.value()),
                    "needle": needle_val,
                    "equals_value": eq_val,
                    "regex_pattern": (self.rules_regex_edit.text() or "").strip(),
                    "fields": fields,
                }
            },
            "branch_on_image": {"column": (self.rules_image_col_edit.text() or "").strip()},
        }

    def refresh_rules_template_combos(self):
        """Füllt die Vorlagen-Combos (Bezug, IF, ELSE) aus dem aktuellen Vorlagen-Ordner."""
        if not hasattr(self, "rules_template_if_combo"):
            return
        vo = self.paths.get("vorlagen_ordner", "").strip()
        preserve_ref = self.rules_reference_template_combo.currentData()
        preserve_if = self.rules_template_if_combo.currentData()
        preserve_else = self.rules_template_else_combo.currentData()
        for cb in (
            self.rules_reference_template_combo,
            self.rules_template_if_combo,
            self.rules_template_else_combo,
        ):
            cb.blockSignals(True)
            cb.clear()
            cb.addItem("(keine Auswahl)", "")
        if vo and os.path.isdir(vo):
            anlagen, allgemein = sammle_vorlagen_pfade(vo)
            for abs_p in sorted(anlagen + allgemein):
                rel = os.path.normpath(os.path.relpath(abs_p, vo))
                self.rules_reference_template_combo.addItem(rel, rel)
                self.rules_template_if_combo.addItem(rel, rel)
                self.rules_template_else_combo.addItem(rel, rel)
        for cb, preserve in (
            (self.rules_reference_template_combo, preserve_ref),
            (self.rules_template_if_combo, preserve_if),
            (self.rules_template_else_combo, preserve_else),
        ):
            if preserve:
                idx = cb.findData(preserve)
                if idx >= 0:
                    cb.setCurrentIndex(idx)
            cb.blockSignals(False)

    def _snapshot_signage_rules_from_ui(self):
        """Übernimmt die Regel-UI nach self.signage_rules (alle Einträge)."""
        if not hasattr(self, "rules_enabled_check"):
            return
        self.rules_enabled = self.rules_enabled_check.isChecked()
        if hasattr(self, "reuse_lageplan_from_last_export_check"):
            self.reuse_lageplan_from_last_export = (
                self.reuse_lageplan_from_last_export_check.isChecked()
            )
        if not hasattr(self, "rules_list_widget"):
            return
        row = self.rules_list_widget.currentRow()
        if row >= 0:
            self._ensure_signage_rules_length(row + 1)
            self.signage_rules[row] = self._collect_rule_dict_from_editor()
            it = self.rules_list_widget.item(row)
            if it:
                it.setText(self.signage_rules[row].get("name") or f"Regel {row + 1}")

    def _apply_signage_rules_to_ui(self):
        """Lädt self.signage_rules in Liste und Editor."""
        if not hasattr(self, "rules_enabled_check"):
            return
        self.rules_enabled_check.setChecked(self.rules_enabled)
        if hasattr(self, "reuse_lageplan_from_last_export_check"):
            self.reuse_lageplan_from_last_export_check.setChecked(
                self.reuse_lageplan_from_last_export
            )
        self._rules_loading_ui = True
        try:
            self.rules_list_widget.clear()
            if not isinstance(self.signage_rules, list):
                self.signage_rules = []
            for i, rule in enumerate(self.signage_rules):
                name = (rule or {}).get("name") or f"Regel {i + 1}"
                self.rules_list_widget.addItem(str(name))
            if self.signage_rules:
                self.rules_list_widget.setCurrentRow(0)
                self._rules_prev_selected_row = 0
                self._set_rule_editor_enabled(True)
                self._load_rule_editor_from_index(0)
            else:
                self._rules_prev_selected_row = -1
                self._set_rule_editor_enabled(False)
                self._clear_rule_fields_table()
        finally:
            self._rules_loading_ui = False
        if not self.signage_rules:
            self.refresh_rules_template_combos()

    def _vorlage_kategorie_label(self, vorlage_basename):
        """Kategorie-Anzeigename wie beim Export (Präfix aus Tab „Kategorien“), sonst Sonstiges."""
        name_lower = vorlage_basename.lower()
        for prefix, cat_name in self.categories.items():
            if name_lower.startswith(prefix.lower()):
                return cat_name
        return "Sonstiges"

    def _gruppiere_vorlagen_nach_kategorie(self, abs_paths):
        """
        Gruppiert absolute Vorlagen-Pfade nach Kategorie, sortiert Gruppen und Dateien.
        Liefert [(kategorie_name, [abs_pfad, ...]), ...].
        """
        buckets = {}
        for p in abs_paths:
            lab = self._vorlage_kategorie_label(os.path.basename(p))
            buckets.setdefault(lab, []).append(p)
        for lab in buckets:
            buckets[lab].sort(key=lambda x: os.path.basename(x).lower())

        all_labels = set(buckets.keys())
        category_order = []
        for _pref, lab in self.categories.items():
            if lab not in category_order:
                category_order.append(lab)
        ordered = [l for l in category_order if l in all_labels]
        rest = sorted(all_labels - set(ordered) - {"Sonstiges"})
        ordered.extend(rest)
        if "Sonstiges" in all_labels:
            ordered.append("Sonstiges")
        return [(l, buckets[l]) for l in ordered]

    def refresh_template_checkboxes(self):
        """Baut die Checkbox-Liste aus dem aktuellen Vorlagen-Ordner auf."""
        while self._template_list_layout.count():
            item = self._template_list_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        self._template_checkbox_by_rel = {}
        vo = self.paths.get('vorlagen_ordner', '').strip()
        if not vo or not os.path.isdir(vo):
            self._template_hint_label.setText("Bitte einen gültigen Vorlagen-Ordner wählen.")
            return

        saved = self.selected_template_rel_paths
        saved_set = None
        if saved is not None:
            saved_set = {os.path.normpath(s) for s in saved}

        def add_checkbox_for_abs(abs_p):
            rel = os.path.normpath(os.path.relpath(abs_p, vo))
            row = QWidget()
            row_l = QHBoxLayout(row)
            row_l.setContentsMargins(0, 2, 0, 2)
            cb = QCheckBox(rel)
            if saved_set is None:
                cb.setChecked(True)
            else:
                cb.setChecked(rel in saved_set)
            cb.stateChanged.connect(self.save_all_settings)
            prev_btn = QPushButton("Vorschau")
            prev_btn.setToolTip("Nur diese Vorlage mit der ersten Excel-Zeile erzeugen (Beispiel).")
            prev_btn.setFixedWidth(88)
            prev_btn.clicked.connect(lambda checked=False, p=abs_p: self.start_preview_single_template(p))
            row_l.addWidget(cb, 1)
            row_l.addWidget(prev_btn)
            self._template_checkbox_by_rel[rel] = cb
            self._template_list_layout.addWidget(row)

        def add_group_mit_kategorien(title, paths):
            if not paths:
                return
            self._template_list_layout.addWidget(QLabel(f"<b>{title}</b>"))
            for cat_label, path_list in self._gruppiere_vorlagen_nach_kategorie(paths):
                sub = QLabel(f"  {cat_label}")
                sub.setStyleSheet("color: #444; margin-top: 4px;")
                self._template_list_layout.addWidget(sub)
                for abs_p in path_list:
                    add_checkbox_for_abs(abs_p)

        anlagen, allgemein = sammle_vorlagen_pfade(vo)
        add_group_mit_kategorien("Anlagen", anlagen)
        add_group_mit_kategorien("Allgemein", allgemein)

        if not self._template_checkbox_by_rel:
            self._template_hint_label.setText(
                "Keine .docx-Vorlagen unter …\\anlagen\\ oder …\\allgemein\\ gefunden."
            )
        else:
            self._template_hint_label.setText(
                "Nur angehakte Vorlagen werden erzeugt. Sortierung nach Kategorien wie unter „Kategorien“ "
                "(Dateinamen-Präfixe, z. B. b_, ba_); ohne passendes Präfix: „Sonstiges“."
            )

    def _snapshot_template_selection(self):
        if not hasattr(self, '_template_checkbox_by_rel') or not self._template_checkbox_by_rel:
            return
        checked = [r for r, cb in self._template_checkbox_by_rel.items() if cb.isChecked()]
        all_rels = list(self._template_checkbox_by_rel.keys())
        if not all_rels:
            return
        self.selected_template_rel_paths = None if len(checked) == len(all_rels) else checked

    def _template_select_all(self):
        if not hasattr(self, '_template_checkbox_by_rel'):
            return
        for cb in self._template_checkbox_by_rel.values():
            cb.setChecked(True)
        self.save_all_settings()

    def _template_select_none(self):
        if not hasattr(self, '_template_checkbox_by_rel'):
            return
        for cb in self._template_checkbox_by_rel.values():
            cb.setChecked(False)
        self.save_all_settings()

    def _template_paths_for_worker(self):
        """None = alle Vorlagen; Liste = nur diese absoluten Pfade; [] = keine / nichts gewählt."""
        vo = self.paths.get('vorlagen_ordner', '').strip()
        if not hasattr(self, '_template_checkbox_by_rel') or not self._template_checkbox_by_rel:
            return []
        all_rels = list(self._template_checkbox_by_rel.keys())
        checked = [r for r, cb in self._template_checkbox_by_rel.items() if cb.isChecked()]
        if not checked:
            return []
        if len(checked) == len(all_rels):
            return None
        return [os.path.normpath(os.path.join(vo, r)) for r in checked]

    def setup_categories_tab(self, tab):
        """Erstellt den Inhalt des 'Kategorien'-Tabs für die Verwaltung der Ausgabeordner."""
        layout = QVBoxLayout(tab)

        # Versions-Info
        version_info = QLabel(f"Version {VERSION} - Build {BUILD_DATE}")
        version_info.setStyleSheet("color: #666; font-size: 10px; padding: 5px;")
        version_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(version_info)

        group_box = QGroupBox("Dokumentkategorien")
        layout.addWidget(group_box)
        group_layout = QVBoxLayout(group_box)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        self.categories_ui_layout = QVBoxLayout(scroll_content)
        scroll.setWidget(scroll_content)

        for prefix, name in self.categories.items():
            self.add_category_widget(prefix, name)

        btn_layout = QHBoxLayout()
        add_btn = QPushButton("+")
        add_btn.setFixedWidth(30)
        add_btn.setToolTip("Neue Kategorie hinzufügen")
        add_btn.clicked.connect(self.add_category)
        remove_btn = QPushButton("-")
        remove_btn.setFixedWidth(30)
        remove_btn.setToolTip("Ausgewählte Kategorie entfernen")
        remove_btn.clicked.connect(self.remove_category)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(remove_btn)
        btn_layout.addStretch()

        group_layout.addLayout(btn_layout)
        group_layout.addWidget(scroll)

        save_btn = QPushButton("Kategorien speichern")
        save_btn.clicked.connect(self.save_categories)
        layout.addWidget(save_btn)
        layout.addStretch()

    def setup_bilder_vorschau_tab(self, tab):
        """Erstellt den Inhalt des 'Bilder-Vorschau'-Tabs: Dateiliste des Bilder-Ordners und Bildvorschau."""
        layout = QVBoxLayout(tab)

        self.bilder_vorschau_hint_label = QLabel("Bitte in der Hauptsteuerung einen Bilder-Ordner wählen.")
        self.bilder_vorschau_hint_label.setStyleSheet("color: #666; padding: 8px;")
        layout.addWidget(self.bilder_vorschau_hint_label)

        splitter = QSplitter(Qt.Horizontal)

        self.bilder_vorschau_table = QTableWidget()
        self.bilder_vorschau_table.setColumnCount(2)
        self.bilder_vorschau_table.setHorizontalHeaderLabels(["Dateiname", "Dateityp"])
        self.bilder_vorschau_table.horizontalHeader().setStretchLastSection(True)
        self.bilder_vorschau_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.bilder_vorschau_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.bilder_vorschau_table.itemSelectionChanged.connect(self._on_bilder_vorschau_selection_changed)
        splitter.addWidget(self.bilder_vorschau_table)

        preview_frame = QFrame()
        preview_frame.setFrameShape(QFrame.StyledPanel)
        preview_layout = QVBoxLayout(preview_frame)
        preview_layout.addWidget(QLabel("Vorschau:"))
        self.bilder_vorschau_preview_label = QLabel("Keine Vorschau")
        self.bilder_vorschau_preview_label.setAlignment(Qt.AlignCenter)
        self.bilder_vorschau_preview_label.setMinimumSize(200, 200)
        self.bilder_vorschau_preview_label.setMaximumSize(400, 400)
        self.bilder_vorschau_preview_label.setStyleSheet("border: 1px solid #ccc; background: #f5f5f5; padding: 8px;")
        self.bilder_vorschau_preview_label.setScaledContents(False)
        preview_layout.addWidget(self.bilder_vorschau_preview_label)
        preview_layout.addStretch()
        splitter.addWidget(preview_frame)

        splitter.setSizes([400, 400])
        layout.addWidget(splitter)

    def _on_tab_changed(self, index):
        """Lädt die Bilder-Vorschau-Liste neu, wenn der Bilder-Vorschau-Tab aktiv wird."""
        if hasattr(self, 'bilder_vorschau_tab_index') and index == self.bilder_vorschau_tab_index:
            self.refresh_bilder_vorschau()
        if hasattr(self, 'vorlagen_tab_index') and index == self.vorlagen_tab_index:
            self.refresh_template_checkboxes()
        if hasattr(self, 'rules_tab_index') and index == self.rules_tab_index:
            self.refresh_rules_template_combos()
        if hasattr(self, 'pdf_tab_index') and index == self.pdf_tab_index:
            self.refresh_pdf_tab()

    def refresh_bilder_vorschau(self):
        """Liest den Bilder-Ordner ein und füllt die Tabelle; setzt die Vorschau zurück."""
        if not hasattr(self, 'bilder_vorschau_table'):
            return
        path = (self.bilder_ordner_edit.text() if hasattr(self, 'bilder_ordner_edit') else "").strip()
        self.bilder_vorschau_table.setRowCount(0)
        self.bilder_vorschau_preview_label.clear()
        self.bilder_vorschau_preview_label.setText("Keine Vorschau")
        if not path:
            self.bilder_vorschau_hint_label.setVisible(True)
            self.bilder_vorschau_hint_label.setText("Bitte in der Hauptsteuerung einen Bilder-Ordner wählen.")
            return
        if not os.path.isdir(path):
            self.bilder_vorschau_hint_label.setVisible(True)
            self.bilder_vorschau_hint_label.setText(f"Ordner existiert nicht: {path}")
            return
        self.bilder_vorschau_hint_label.setVisible(False)
        try:
            names = sorted(os.listdir(path))
        except OSError:
            self.bilder_vorschau_hint_label.setVisible(True)
            self.bilder_vorschau_hint_label.setText("Ordner konnte nicht gelesen werden.")
            return
        for name in names:
            full = os.path.join(path, name)
            if os.path.isfile(full):
                ext = os.path.splitext(name)[1].lower()
                if ext in BILDER_VORSCHAU_EXTENSIONS:
                    row = self.bilder_vorschau_table.rowCount()
                    self.bilder_vorschau_table.insertRow(row)
                    self.bilder_vorschau_table.setItem(row, 0, QTableWidgetItem(name))
                    self.bilder_vorschau_table.setItem(row, 1, QTableWidgetItem(ext or "—"))
                    self.bilder_vorschau_table.item(row, 0).setData(Qt.UserRole, full)
                    self.bilder_vorschau_table.item(row, 1).setData(Qt.UserRole, full)

    def _on_bilder_vorschau_selection_changed(self):
        """Zeigt bei Auswahl einer Zeile die Bildvorschau an."""
        if not hasattr(self, 'bilder_vorschau_preview_label'):
            return
        row = self.bilder_vorschau_table.currentRow()
        if row < 0:
            self.bilder_vorschau_preview_label.clear()
            self.bilder_vorschau_preview_label.setText("Keine Vorschau")
            return
        item = self.bilder_vorschau_table.item(row, 0)
        file_path = item.data(Qt.UserRole) if item else None
        if not file_path or not os.path.isfile(file_path):
            self.bilder_vorschau_preview_label.clear()
            self.bilder_vorschau_preview_label.setText("Keine Vorschau")
            return
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.svg':
            try:
                from core.utils import svg_to_png_file_pyside
                import tempfile
                fd, tmp_png = tempfile.mkstemp(suffix='.png')
                os.close(fd)
                try:
                    if svg_to_png_file_pyside(file_path, tmp_png, scale=2, compression=-1):
                        pix = QPixmap(tmp_png)
                        if not pix.isNull():
                            scaled = pix.scaled(400, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            self.bilder_vorschau_preview_label.setPixmap(scaled)
                            self.bilder_vorschau_preview_label.setText("")
                        else:
                            self.bilder_vorschau_preview_label.clear()
                            self.bilder_vorschau_preview_label.setText("SVG konnte nicht geladen werden.")
                    else:
                        self.bilder_vorschau_preview_label.clear()
                        self.bilder_vorschau_preview_label.setText("Keine Vorschau")
                finally:
                    try:
                        os.unlink(tmp_png)
                    except OSError:
                        pass
            except Exception:
                self.bilder_vorschau_preview_label.clear()
                self.bilder_vorschau_preview_label.setText("Keine Vorschau")
        else:
            pix = QPixmap(file_path)
            if not pix.isNull():
                scaled = pix.scaled(400, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.bilder_vorschau_preview_label.setPixmap(scaled)
                self.bilder_vorschau_preview_label.setText("")
            else:
                self.bilder_vorschau_preview_label.clear()
                self.bilder_vorschau_preview_label.setText("Keine Vorschau")

    def setup_pdf_tab(self, tab):
        """Erstellt den separaten PDF-Reiter für die manuelle Konvertierung des letzten Exports."""
        layout = QVBoxLayout(tab)

        info = QLabel("PDF-Umwandlung für den letzten erfolgreichen Export")
        info.setStyleSheet("font-weight: 600;")
        layout.addWidget(info)

        self.pdf_last_export_label = QLabel("Letzter Export: —")
        layout.addWidget(self.pdf_last_export_label)

        self.pdf_progress = QProgressBar()
        self.pdf_progress.setValue(0)
        layout.addWidget(self.pdf_progress)

        self.pdf_current_file_label = QLabel("Bereit.")
        self.pdf_current_file_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.pdf_current_file_label)

        self.pdf_summary_label = QLabel("Noch keine PDF-Konvertierung gestartet.")
        self.pdf_summary_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.pdf_summary_label)

        btn_row = QHBoxLayout()
        self.pdf_convert_btn = QPushButton("Letzten Export in PDF umwandeln")
        self.pdf_convert_btn.clicked.connect(self.convert_last_export_to_pdf)
        btn_row.addWidget(self.pdf_convert_btn)

        self.pdf_open_folder_btn = QPushButton("Letzten Export-Ordner öffnen")
        self.pdf_open_folder_btn.clicked.connect(self.open_export_folder)
        btn_row.addWidget(self.pdf_open_folder_btn)

        self.pdf_select_folder_btn = QPushButton("Ordner auswählen...")
        self.pdf_select_folder_btn.clicked.connect(self.select_pdf_source_folder)
        btn_row.addWidget(self.pdf_select_folder_btn)
        layout.addLayout(btn_row)

        self.pdf_log_text = QTextEdit()
        self.pdf_log_text.setReadOnly(True)
        layout.addWidget(self.pdf_log_text)

        self.refresh_pdf_tab()

    def setup_logs_tab(self, tab):
        """Erstellt den Inhalt des 'Logs'-Tabs mit Filter-Dropdown."""
        layout = QVBoxLayout(tab)

        # Versions-Info (immer sichtbar)
        version_info = QLabel(f"DocxTpl Automatisierung v{VERSION} - Build {BUILD_DATE}")
        version_info.setStyleSheet("color: #666; font-size: 10px; padding: 5px;")
        version_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(version_info)

        # Filter-Dropdown
        self.log_filter_combo = QComboBox()
        self.log_filter_combo.addItems([
            "ALLE", "INFO", "WARN", "ERROR", "SUCCESS", "FATAL", "SEP"
        ])
        self.log_filter_combo.currentTextChanged.connect(self.apply_log_filter)
        layout.addWidget(self.log_filter_combo)
        # Log-Textfeld
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setContextMenuPolicy(Qt.CustomContextMenu)
        self.log_text.customContextMenuRequested.connect(self.show_log_context_menu)
        self.log_lines = []  # Speichert alle Log-Zeilen als (level, html) Tupel
        layout.addWidget(self.log_text)

    def setup_menu(self):
        """Erstellt die Menüleiste für die Anwendung, inkl. Projekt- und Theme-Menüs."""
        menu_bar = self.menuBar()
        # Projekt-Menü
        project_menu = menu_bar.addMenu("Projekt")
        act_new = QAction("Neu", self)
        act_open = QAction("Öffnen...", self)
        act_save = QAction("Speichern", self)
        act_save_as = QAction("Speichern unter...", self)
        project_menu.addAction(act_new)
        project_menu.addAction(act_open)
        self.recent_projects_menu = project_menu.addMenu("Letzte Projekte")
        self.recent_projects_menu.aboutToShow.connect(self._refresh_recent_projects_menu)
        project_menu.addSeparator()
        project_menu.addAction(act_save)
        project_menu.addAction(act_save_as)

        act_new.triggered.connect(self.project_new)
        act_open.triggered.connect(self.project_open)
        act_save.triggered.connect(self.project_save)
        act_save_as.triggered.connect(self.project_save_as)

        self._refresh_recent_projects_menu()

        # Ansicht/Theme
        view_menu = menu_bar.addMenu("Ansicht")
        theme_menu = view_menu.addMenu("Theme")

        self.theme_action_group = QActionGroup(self)
        self.theme_action_group.setExclusive(True)

        actions = {
            "Light": QAction("Hell (Standard)", self, checkable=True),
            "Dark": QAction("Dark Mode", self, checkable=True),
            "Girly": QAction("Girly Mode", self, checkable=True)
        }

        for name, action in actions.items():
            action.setData(name)
            self.theme_action_group.addAction(action)
            theme_menu.addAction(action)

        self.theme_action_group.triggered.connect(self.on_theme_selected)

    # -----------------------------
    # Projektverwaltung - Methoden
    # -----------------------------
    _RE_USERS_PROFILE = re.compile(
        r'(?i)([a-z]:)([\\/])Users\2([^\\/]+)',
    )
    _RE_USERS_PLACEHOLDER = re.compile(
        r'(?i)([a-z]:)([\\/])Users\2\{username\}',
    )

    def _normalize_path(self, path):
        """
        Normalisiert einen Pfad für portables Speichern:
        Ersetzt den Profilordner unter …\\Users\\<name> durch {username}.
        Laufwerk und Schreibweise von „Users“ werden case-insensitive erkannt.
        """
        if not path:
            return path

        def _repl(m):
            return f"{m.group(1)}{m.group(2)}Users{m.group(2)}{{username}}"

        return self._RE_USERS_PROFILE.sub(_repl, path)

    def _denormalize_path(self, path):
        """
        Denormalisiert einen Pfad beim Laden:
        Ersetzt {username} durch den aktuellen Benutzernamen (c:/users/… ebenfalls).
        """
        if not path:
            return path

        def _repl(m):
            return f"{m.group(1)}{m.group(2)}Users{m.group(2)}{self.current_username}"

        return self._RE_USERS_PLACEHOLDER.sub(_repl, path)

    def normalize_all_paths(self):
        """
        Normalisiert alle aktuellen Pfade in der UI (Button-Callback).
        Zeigt danach eine Bestätigung an.
        """
        changed_count = 0
        for key, value in self.paths.items():
            if value and '{username}' not in value:
                normalized = self._normalize_path(value)
                if normalized != value:
                    self.paths[key] = normalized
                    if hasattr(self, f"{key}_edit"):
                        getattr(self, f"{key}_edit").setText(normalized)
                        self.validate_path(getattr(self, f"{key}_edit"))
                    changed_count += 1

        if changed_count > 0:
            self.save_all_settings()
            QMessageBox.information(
                self,
                "Pfade normalisiert",
                f"{changed_count} Pfade wurden normalisiert (Benutzername → '{{username}}').\n\n"
                "Die Pfade funktionieren jetzt auf allen Rechnern!"
            )
        else:
            QMessageBox.information(
                self,
                "Keine Normalisierung",
                "Entweder ist '{{username}}' bereits gesetzt, oder es wurde kein Windows-Profilpfad "
                "im Muster Laufwerk:\\Users\\<Benutzername> erkannt.",
            )

    def denormalize_all_paths(self):
        """Ersetzt {username} in allen Pfaden durch den angemeldeten Benutzer (lokale Nutzung)."""
        changed_count = 0
        for key, value in list(self.paths.items()):
            if not value or '{username}' not in value:
                continue
            denorm = self._denormalize_path(value)
            if denorm != value:
                self.paths[key] = denorm
                if hasattr(self, f"{key}_edit"):
                    getattr(self, f"{key}_edit").setText(denorm)
                    self.validate_path(getattr(self, f"{key}_edit"))
                changed_count += 1

        if changed_count > 0:
            self.save_all_settings()
            QMessageBox.information(
                self,
                "Pfade lokal gesetzt",
                f"{changed_count} Pfad(e): '{{username}}' wurde durch '{self.current_username}' ersetzt.",
            )
        else:
            QMessageBox.information(
                self,
                "Keine Änderung",
                "Kein '{{username}}' in den Pfaden gefunden (nichts zu ersetzen).",
            )

    def _collect_project_dict(self):
        """Sammelt den aktuellen Zustand für ein Projekt-JSON."""
        # Stelle sicher, dass Settings aus Spins/Textfeldern aktuell sind
        for key in self.settings:
            if hasattr(self, f"{key}_spin"):
                self.settings[key] = getattr(self, f"{key}_spin").value()
        if hasattr(self, 'datetime_utc_format_edit'):
            self.settings['datetime_utc_format'] = self.datetime_utc_format_edit.text()
        if hasattr(self, 'projekt_name_override_edit'):
            self.settings['projekt_name_override'] = self.projekt_name_override_edit.text().strip()
        if hasattr(self, 'export_as_pdf_check'):
            self.settings['export_as_pdf'] = self.export_as_pdf_check.isChecked()

        self._snapshot_template_selection()
        self._snapshot_signage_rules_from_ui()

        # Pfade normalisieren (Benutzername → {username}) für Portabilität
        normalized_paths = {k: self._normalize_path(v) for k, v in self.paths.items()}

        return {
            'version': VERSION,
            'paths': normalized_paths,
            'settings': self.settings,
            'categories': self.categories,
            'selected_template_rel_paths': self.selected_template_rel_paths,
            'rules_enabled': self.rules_enabled,
            'reuse_lageplan_from_last_export': self.reuse_lageplan_from_last_export,
            'signage_rules': self.signage_rules,
            'saved_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    def _apply_project_dict(self, data):
        """Wendet ein geladenes Projekt auf UI und Zustand an."""
        # Pfade denormalisieren ({username} → aktueller Benutzername), immer alle vier Keys
        loaded_paths = data.get('paths', self.paths) or {}
        self.paths = {
            k: self._denormalize_path(loaded_paths.get(k, ''))
            for k in PROJECT_FILE_PATH_KEYS
        }

        self.settings = {**self.settings, **data.get('settings', {})}
        if 'selected_template_rel_paths' in data:
            self.selected_template_rel_paths = data.get('selected_template_rel_paths')
        else:
            self.selected_template_rel_paths = None
        if 'rules_enabled' in data:
            self.rules_enabled = bool(data.get('rules_enabled'))
        if 'reuse_lageplan_from_last_export' in data:
            self.reuse_lageplan_from_last_export = bool(data.get('reuse_lageplan_from_last_export'))
        if 'signage_rules' in data:
            self.signage_rules = data.get('signage_rules') or []

        if isinstance(data.get('categories'), dict) and data.get('categories'):
            self.categories = {**self.categories, **data['categories']}
            self.categories.setdefault('p_', 'Pläne')

        # UI aktualisieren: Pfade
        for key, value in self.paths.items():
            if hasattr(self, f"{key}_edit"):
                getattr(self, f"{key}_edit").setText(value)
                self.validate_path(getattr(self, f"{key}_edit"))
        # Settings (Spins)
        for key in ['header_row', 'svg_scale', 'png_compression']:
            if hasattr(self, f"{key}_spin"):
                getattr(self, f"{key}_spin").setValue(
                    self.settings.get(key, getattr(self, f"{key}_spin").value())
                )
        # UTC Format
        if hasattr(self, 'datetime_utc_format_edit'):
            self.datetime_utc_format_edit.setText(
                self.settings.get('datetime_utc_format', self.datetime_utc_format_edit.text())
            )
        # Projektname (Export-Ordner)
        if hasattr(self, 'projekt_name_override_edit'):
            self.projekt_name_override_edit.setText(self.settings.get('projekt_name_override', ''))
        # PDF-Export
        if hasattr(self, 'export_as_pdf_check'):
            self.export_as_pdf_check.setChecked(self.settings.get('export_as_pdf', False))
        # Theme
        self.apply_theme(self.settings.get('theme', 'Light'))
        # Kategorien-UI neu aufbauen
        if hasattr(self, 'categories_ui_layout'):
            # Bestehende Widgets entfernen
            for layout_item, _triplet in list(self.category_widgets.items()):
                while layout_item.count():
                    child = layout_item.takeAt(0)
                    if child.widget():
                        child.widget().deleteLater()
                self.categories_ui_layout.removeItem(layout_item)
                layout_item.deleteLater()
                del self.category_widgets[layout_item]
            # Neu hinzufügen
            for prefix, name in self.categories.items():
                self.add_category_widget(prefix, name)
        # Excel-Vorschau aktualisieren
        self.show_excel_data()
        if hasattr(self, 'refresh_template_checkboxes'):
            self.refresh_template_checkboxes()
        if hasattr(self, '_apply_signage_rules_to_ui'):
            self._apply_signage_rules_to_ui()
        # Einstellungen persistieren
        self.save_all_settings()

    def _remember_recent_project(self, file_path):
        """Merkt sich ein Projekt in der Liste zuletzt verwendeter Projekte."""
        try:
            if not file_path:
                return
            if file_path in self.recent_projects:
                self.recent_projects.remove(file_path)
            self.recent_projects.insert(0, file_path)
            self.recent_projects = self.recent_projects[:10]
            self.current_project_path = file_path
            # in settings.json speichern
            self.save_all_settings()
        except Exception:
            pass

    def _refresh_recent_projects_menu(self):
        """Aktualisiert das Untermenü 'Letzte Projekte' mit den zuletzt geöffneten Projekten."""
        self.recent_projects_menu.clear()
        recent = getattr(self, 'recent_projects', [])[:MAX_RECENT_PROJECTS]
        for path in recent:
            if not os.path.isfile(path):
                continue
            display_name = os.path.basename(path)
            action = QAction(display_name, self)
            action.setToolTip(path)
            action.triggered.connect(lambda checked, p=path: self.project_open_path(p))
            self.recent_projects_menu.addAction(action)
        if not self.recent_projects_menu.actions():
            none_action = QAction("(Keine)", self)
            none_action.setEnabled(False)
            self.recent_projects_menu.addAction(none_action)

    def project_new(self):
        """Setzt ein leeres Projekt (ohne Pfade), behält aber Kategorien/Settings-Grundwerte bei."""
        self.paths = {k: '' for k in self.paths.keys()}
        self.selected_template_rel_paths = None
        self.rules_enabled = False
        self.reuse_lageplan_from_last_export = False
        self.signage_rules = []
        for key in ['header_row', 'svg_scale', 'png_compression']:
            if hasattr(self, f"{key}_spin"):
                getattr(self, f"{key}_spin").setValue(
                    self.settings.get(key, getattr(self, f"{key}_spin").minimum())
                )
        if hasattr(self, 'datetime_utc_format_edit'):
            self.datetime_utc_format_edit.setText('%Y-%m-%d %H:%M:%S UTC')
        self.apply_theme('Light')
        for key in self.paths:
            if hasattr(self, f"{key}_edit"):
                getattr(self, f"{key}_edit").setText('')
                self.validate_path(getattr(self, f"{key}_edit"))
        self.current_project_path = None
        if hasattr(self, 'refresh_template_checkboxes'):
            self.refresh_template_checkboxes()
        if hasattr(self, '_apply_signage_rules_to_ui'):
            self._apply_signage_rules_to_ui()
        self.save_all_settings()

    def project_open_path(self, file_path):
        """Lädt ein Projekt aus dem angegebenen Pfad (ohne Dateidialog)."""
        if not file_path or not os.path.isfile(file_path):
            if file_path:
                QMessageBox.warning(self, "Fehler", f"Projektdatei nicht gefunden:\n{file_path}")
            return
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Erlaube beide Formate: altes Flat-Settings.json oder neues Projektformat
            if 'paths' in data and 'settings' in data:
                self._apply_project_dict(data)
            else:
                # Altes Format: direkt anwenden
                self._apply_project_dict({
                    'paths': {k: data.get(k, '') for k in self.paths.keys()},
                    'settings': {k: data.get(k, self.settings.get(k)) for k in self.settings.keys()},
                    'categories': data.get('categories', {}),
                    'rules_enabled': data.get('rules_enabled', False),
                    'reuse_lageplan_from_last_export': data.get('reuse_lageplan_from_last_export', False),
                    'signage_rules': data.get('signage_rules', []),
                })
            self._remember_recent_project(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Projekt konnte nicht geladen werden: {e}")

    def project_open(self):
        """Öffnet eine Projektdatei (*.dta.json) per Dateidialog."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Projekt öffnen", os.getcwd(), "Projektdateien (*.dta.json *.json)"
        )
        if file_path:
            self.project_open_path(file_path)

    def project_save(self):
        """Speichert das aktuelle Projekt an den zuletzt genutzten Projektpfad oder fragt nach einem neuen Pfad."""
        if not self.current_project_path:
            return self.project_save_as()
        try:
            data = self._collect_project_dict()
            with open(self.current_project_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            QMessageBox.information(self, "Gespeichert", f"Projekt gespeichert: {self.current_project_path}")
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Projekt konnte nicht gespeichert werden: {e}")

    def project_save_as(self):
        """Speichert das aktuelle Projekt unter einem neuen Dateinamen (*.dta.json)."""
        default_name = self._build_default_project_filename()
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Projekt speichern unter",
            os.path.join(os.getcwd(), default_name),
            "Projektdateien (*.dta.json)"
        )
        if not file_path:
            return
        try:
            data = self._collect_project_dict()
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            self._remember_recent_project(file_path)
            QMessageBox.information(self, "Gespeichert", f"Projekt gespeichert: {file_path}")
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Projekt konnte nicht gespeichert werden: {e}")

    def _sanitize_filename_part(self, value):
        """Bereinigt einen String für sichere Dateinamen."""
        text = str(value or "").strip()
        if not text:
            return ""
        sanitized = "".join(c if c not in r'\/:*?"<>|' else "_" for c in text)
        sanitized = sanitized.replace("\n", " ").replace("\r", " ").strip(" .")
        return sanitized

    def _base_name_for_project_dta_file(self, value):
        """Bereinigt den Basisnamen; entfernt angehängte .dta / .dta.json, damit nicht .dta.dta.json entsteht."""
        s = self._sanitize_filename_part(value)
        if not s:
            return ""
        while True:
            low = s.lower()
            if low.endswith('.dta.json'):
                s = s[:-9].rstrip(' .')
            elif low.endswith('.dta'):
                s = s[:-4].rstrip(' .')
            else:
                break
        return s

    def _read_name_from_excel(self):
        """
        Liest einen möglichen Winpark-/Projektnamen aus der aktuell gewählten Excel-Datei.
        Unterstützt Blatt 1 (Spalten) und Blatt 2 (Fallback-Marken).
        """
        excel_path = self.paths.get('excel_path', '').strip()
        if not excel_path or not os.path.isfile(excel_path):
            return ""

        key_priority = ['winpark_name', 'windpark_name', 'winpark', 'windpark', 'projekt_name']
        header_row = self.settings.get('header_row', 3)

        try:
            df = pd.read_excel(excel_path, header=header_row - 1)
            if not df.empty:
                normalized_cols = {str(col).strip().lower(): col for col in df.columns}
                for key in key_priority:
                    col = normalized_cols.get(key)
                    if col is None:
                        continue
                    series = df[col].dropna()
                    if not series.empty:
                        candidate = self._sanitize_filename_part(series.iloc[0])
                        if candidate:
                            return candidate
        except Exception:
            pass

        wb = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            if len(wb.sheetnames) > 1:
                ws = wb[wb.sheetnames[1]]
                fallback = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 3 and row[1]:
                        fallback_key = str(row[1]).strip().lower()
                        fallback_value = str(row[2]).strip() if row[2] is not None else ""
                        fallback[fallback_key] = fallback_value
                for key in key_priority:
                    candidate = self._sanitize_filename_part(fallback.get(key, ""))
                    if candidate:
                        return candidate
        except Exception:
            pass
        finally:
            if wb is not None:
                wb.close()

        return ""

    def _build_default_project_filename(self):
        """
        Erzeugt den Standard-Dateinamen für 'Speichern unter...':
        1) aktueller Projektdateiname (falls vorhanden),
        2) Projektname-Override,
        3) Winpark-/Projektname aus Excel,
        4) Fallback.
        """
        if self.current_project_path:
            current_name = os.path.basename(self.current_project_path)
            if current_name:
                return current_name

        override = ""
        if hasattr(self, 'projekt_name_override_edit'):
            override = self._sanitize_filename_part(self.projekt_name_override_edit.text())
        if not override:
            override = self._sanitize_filename_part(self.settings.get('projekt_name_override', ''))
        if override:
            base = self._base_name_for_project_dta_file(override)
            return f"{base}.dta.json" if base else 'projekt.dta.json'

        excel_name = self._read_name_from_excel()
        if excel_name:
            base = self._base_name_for_project_dta_file(excel_name)
            return f"{base}.dta.json" if base else 'projekt.dta.json'

        return 'projekt.dta.json'

    def on_theme_selected(self, action):
        """Wird aufgerufen, wenn ein Theme aus dem Menü ausgewählt wird."""
        theme_name = action.data()
        self.apply_theme(theme_name)
        self.settings['theme'] = theme_name
        self.save_all_settings()

    def apply_theme(self, theme_name):
        """Wendet das ausgewählte Stylesheet an und aktualisiert die Menü-Auswahl. Setzt im Dark Mode den Log-Text auf weiß."""
        stylesheet = self.get_stylesheet(theme_name)
        self.setStyleSheet(stylesheet)  # Apply to main window
        for action in self.theme_action_group.actions():
            if action.data() == theme_name:
                action.setChecked(True)
                break
        # Im Dark Mode: Log-Text maximal hell und kontrastreich
        self.is_dark_mode = (theme_name == 'Dark')
        if self.is_dark_mode:
            self.log_text.setStyleSheet("color: #FFF; background-color: #181818; font-weight: 500;")
        else:
            self.log_text.setStyleSheet("")

    def get_stylesheet(self, theme_name):
        """Gibt den QSS-Stylesheet-String für das angeforderte Theme zurück."""
        theme_map = {
            'Dark': 'dark.qss',
            'Girly': 'girly.qss'
        }
        qss_file = theme_map.get(theme_name)
        if qss_file:
            qss_path = resource_path(qss_file)
            if os.path.exists(qss_path):
                try:
                    with open(qss_path, 'r') as f:
                        return f.read()
                except IOError:
                    return ""  # Fallback to default
        return ""  # Light/Default Theme

    def on_path_change(self, key, value):
        """Wird aufgerufen, wenn ein Pfad geändert wird, und speichert die Einstellung."""
        self.paths[key] = value
        self.save_all_settings()

        # Aktualisiere Excel-Datenanzeige wenn sich der Excel-Pfad ändert
        if key == 'excel_path':
            self.show_excel_data()
        # Bilder-Vorschau-Tab aktualisieren wenn sich der Bilder-Ordner ändert
        if key == 'bilder_ordner' and hasattr(self, 'bilder_vorschau_table'):
            self.refresh_bilder_vorschau()
        if key == 'vorlagen_ordner':
            if hasattr(self, 'refresh_template_checkboxes'):
                self.refresh_template_checkboxes()
            if hasattr(self, 'refresh_rules_template_combos'):
                self.refresh_rules_template_combos()

    def browse(self, key):
        """Öffnet einen Datei- oder Ordnerdialog und aktualisiert das entsprechende Eingabefeld."""
        path_edit = getattr(self, f"{key}_edit")
        dialog_title = "Ordner auswählen" if 'ordner' in key else "Excel-Datei auswählen"
        start_path = path_edit.text()
        if 'ordner' in key:
            path = QFileDialog.getExistingDirectory(self, dialog_title, start_path)
        else:
            path, _ = QFileDialog.getOpenFileName(self, dialog_title, start_path, "*.xlsx *.xls")
        if path:
            path_edit.setText(path)
            self.on_path_change(key, path)
            self.validate_path(path_edit)

    def validate_path(self, line_edit_widget):
        """Prüft, ob der Pfad in einem QLineEdit existiert und färbt es entsprechend."""
        path = line_edit_widget.text()
        is_optional_empty = 'bilder_ordner' in self.paths and hasattr(self, 'bilder_ordner_edit') and line_edit_widget == self.bilder_ordner_edit and not path

        valid_style = "background-color: #e8e8e8; color: #2d2d2d;"  # Leichtes Grau bei gültigem Pfad
        invalid_style = "background-color: #f8d7da; color: #721c24;"  # Reddish

        if os.path.exists(path) or is_optional_empty:
            line_edit_widget.setStyleSheet(valid_style)
        else:
            line_edit_widget.setStyleSheet(invalid_style)

    def add_category_widget(self, prefix, name):
        """Fügt der UI eine neue Zeile zur Eingabe einer Kategorie hinzu."""
        row_layout = QHBoxLayout()
        checkbox = QCheckBox()
        prefix_edit = QLineEdit(prefix)
        name_edit = QLineEdit(name)

        row_layout.addWidget(checkbox)
        row_layout.addWidget(QLabel("Präfix:"))
        row_layout.addWidget(prefix_edit)
        row_layout.addWidget(QLabel("Ordnername:"))
        row_layout.addWidget(name_edit)

        self.categories_ui_layout.addLayout(row_layout)
        self.category_widgets[row_layout] = (checkbox, prefix_edit, name_edit)

    def add_category(self):
        """Event-Handler, um eine neue, leere Kategorie-Zeile hinzuzufügen."""
        self.add_category_widget("", "")

    def remove_category(self):
        """Event-Handler, um alle ausgewählten Kategorie-Zeilen zu entfernen."""
        to_remove = [l for l, (c, _, _) in self.category_widgets.items() if c.isChecked()]
        for layout in to_remove:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
            self.categories_ui_layout.removeItem(layout)
            layout.deleteLater()
            del self.category_widgets[layout]

    def save_categories(self):
        """Sammelt die Daten aus den Kategorie-Eingabefeldern und speichert sie."""
        self.categories.clear()
        # Werte bestehen aus (checkbox, prefix_edit, name_edit)
        for checkbox, prefix_edit, name_edit in self.category_widgets.values():
            prefix = prefix_edit.text().strip()
            name = name_edit.text().strip()
            if prefix and name:
                self.categories[prefix] = name
        self.save_all_settings()
        QMessageBox.information(self, "Gespeichert", "Kategorien aktualisiert.")

    def save_all_settings(self):
        """Sammelt alle aktuellen Einstellungen und speichert sie in 'settings.json'."""
        for key in self.settings:
            if hasattr(self, f"{key}_spin"):
                self.settings[key] = getattr(self, f"{key}_spin").value()

        if hasattr(self, 'datetime_utc_format_edit'):
            self.settings['datetime_utc_format'] = self.datetime_utc_format_edit.text()
        if hasattr(self, 'projekt_name_override_edit'):
            self.settings['projekt_name_override'] = self.projekt_name_override_edit.text().strip()
        if hasattr(self, 'export_as_pdf_check'):
            self.settings['export_as_pdf'] = self.export_as_pdf_check.isChecked()

        self._snapshot_template_selection()
        self._snapshot_signage_rules_from_ui()

        try:
            payload = {**self.paths, **self.settings, 'categories': self.categories}
            payload['selected_template_rel_paths'] = self.selected_template_rel_paths
            payload['rules_enabled'] = self.rules_enabled
            payload['reuse_lageplan_from_last_export'] = self.reuse_lageplan_from_last_export
            payload['signage_rules'] = self.signage_rules
            # Projektverwaltung: letzte Projekte und aktueller Projektpfad speichern
            if hasattr(self, 'recent_projects'):
                payload['recent_projects'] = self.recent_projects
            if hasattr(self, 'current_project_path') and self.current_project_path:
                payload['last_project_path'] = self.current_project_path
            with open('settings.json', 'w', encoding='utf-8') as f:
                json.dump(payload, f, indent=4, ensure_ascii=False)
        except IOError as e:
            self.log_text.append(f"Speicherfehler: {e}")
            _log_handler(f"Speicherfehler: {e}", "ERROR", self.append_html_log)

    def load_settings(self):
        """Lädt Einstellungen aus 'settings.json', falls die Datei existiert."""
        try:
            if os.path.exists('settings.json'):
                with open('settings.json', 'r', encoding='utf-8') as f:
                    return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
        return {}

    def start_dry_run(self):
        """Startet den Trockenlauf-Prozess."""
        if not all(self.paths.get(k) for k in ['excel_path', 'vorlagen_ordner']):
            QMessageBox.warning(self, "Fehlende Pfade", "Bitte Excel- und Vorlagen-Pfad für den Trockenlauf angeben!")
            return
        if not getattr(self, '_template_checkbox_by_rel', {}):
            QMessageBox.warning(
                self,
                "Keine Vorlagen",
                "Im Vorlagen-Ordner wurden keine verwendbaren Vorlagen (Unterordner anlagen/ oder allgemein/) gefunden.",
            )
            return
        sel = self._template_paths_for_worker()
        if sel == []:
            QMessageBox.warning(
                self,
                "Keine Auswahl",
                "Bitte mindestens eine Vorlage auswählen.",
            )
            return
        self.save_all_settings()
        self.log_text.clear()

        previous_export_root = None
        if self.reuse_lageplan_from_last_export and self.last_export_path and os.path.isdir(self.last_export_path):
            previous_export_root = self.last_export_path

        worker_params = {
            **self.paths,
            **self.settings,
            'categories': self.categories,
            'dry_run': True,
            'is_dark_mode': self.is_dark_mode,
            'selected_template_paths': sel,
            'rules_enabled': self.rules_enabled,
            'signage_rules': list(self.signage_rules) if self.signage_rules else [],
            'reuse_lageplan_from_last_export': self.reuse_lageplan_from_last_export,
            'previous_export_root': previous_export_root,
        }

        self.worker = Worker(**worker_params)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.log.connect(self.append_html_log)

        # Für den Trockenlauf brauchen wir diese Signale nicht, aber der Worker erwartet sie
        self.worker.current_file.connect(lambda: None)
        self.worker.progress.connect(lambda: None)

        self.worker.start()
        self.set_ui_running_state(True)

    @staticmethod
    def _first_pdf_under_dir(root_dir):
        if not root_dir or not os.path.isdir(root_dir):
            return None
        found = []
        for dirpath, _, files in os.walk(root_dir):
            for name in files:
                if name.lower().endswith('.pdf') and not name.startswith('~'):
                    found.append(os.path.join(dirpath, name))
        if not found:
            return None
        found.sort(key=lambda p: p.lower())
        return found[0]

    def start_preview_batch(self):
        """Stapel-Vorschau: je eine Ausgabe pro gewählter Vorlage, erste Excel-Zeile für Anlagen."""
        if not all(self.paths.get(k) for k in ['excel_path', 'vorlagen_ordner', 'export_ordner']):
            QMessageBox.warning(
                self,
                "Fehlende Pfade",
                "Bitte Excel-, Vorlagen- und Export-Pfad für die Vorschau angeben!",
            )
            return
        if not getattr(self, '_template_checkbox_by_rel', {}):
            QMessageBox.warning(
                self,
                "Keine Vorlagen",
                "Im Vorlagen-Ordner wurden keine verwendbaren Vorlagen (Unterordner anlagen/ oder allgemein/) gefunden.",
            )
            return
        sel = self._template_paths_for_worker()
        if sel == []:
            QMessageBox.warning(
                self,
                "Keine Auswahl",
                "Bitte mindestens eine Vorlage auswählen.",
            )
            return
        self.save_all_settings()
        self.log_text.clear()

        previous_export_root = None
        if self.reuse_lageplan_from_last_export and self.last_export_path and os.path.isdir(self.last_export_path):
            previous_export_root = self.last_export_path

        worker_params = {
            **self.paths,
            **self.settings,
            'categories': self.categories,
            'dry_run': False,
            'is_dark_mode': self.is_dark_mode,
            'selected_template_paths': sel,
            'rules_enabled': self.rules_enabled,
            'signage_rules': list(self.signage_rules) if self.signage_rules else [],
            'reuse_lageplan_from_last_export': self.reuse_lageplan_from_last_export,
            'previous_export_root': previous_export_root,
            'preview_run': True,
            'preview_template_abs': None,
        }

        self.worker = Worker(**worker_params)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.log.connect(self.append_html_log)
        self.worker.current_file.connect(self.update_current_file_label)
        self.worker.progress.connect(self.update_progress_bar)
        self.worker.start()
        self.set_ui_running_state(True)

    def start_preview_single_template(self, template_abs):
        """Einzel-Vorschau für eine Vorlage (unabhängig von der Checkbox)."""
        if not all(self.paths.get(k) for k in ['excel_path', 'vorlagen_ordner', 'export_ordner']):
            QMessageBox.warning(
                self,
                "Fehlende Pfade",
                "Bitte Excel-, Vorlagen- und Export-Pfad für die Vorschau angeben!",
            )
            return
        if not template_abs or not os.path.isfile(template_abs):
            QMessageBox.warning(self, "Vorschau", "Vorlagendatei nicht gefunden.")
            return
        self.save_all_settings()
        self.log_text.clear()

        previous_export_root = None
        if self.reuse_lageplan_from_last_export and self.last_export_path and os.path.isdir(self.last_export_path):
            previous_export_root = self.last_export_path

        worker_params = {
            **self.paths,
            **self.settings,
            'categories': self.categories,
            'dry_run': False,
            'is_dark_mode': self.is_dark_mode,
            'selected_template_paths': None,
            'rules_enabled': self.rules_enabled,
            'signage_rules': list(self.signage_rules) if self.signage_rules else [],
            'reuse_lageplan_from_last_export': self.reuse_lageplan_from_last_export,
            'previous_export_root': previous_export_root,
            'preview_run': True,
            'preview_template_abs': template_abs,
        }

        self._show_template_row_preview_ui(os.path.basename(template_abs))
        self.worker = Worker(**worker_params)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.log.connect(self.append_html_log)
        self.worker.current_file.connect(self.update_current_file_label)
        self.worker.current_file.connect(self._on_template_row_preview_file)
        self.worker.progress.connect(self.update_progress_bar)
        self.worker.progress.connect(self._on_template_row_preview_progress)
        self.worker.start()
        self.set_ui_running_state(True)

    def start(self, dry_run=False):
        """
        Startet den Dokumentenerstellungsprozess.
        - Prüft, ob alle notwendigen Pfade angegeben sind.
        - Speichert die aktuellen Einstellungen.
        - Erstellt und startet den Worker-Thread mit allen notwendigen Parametern.
        - Deaktiviert den 'Start'-Knopf, um doppelte Ausführungen zu verhindern.
        """
        if not all(self.paths.get(k) for k in ['excel_path', 'vorlagen_ordner', 'export_ordner']):
            QMessageBox.warning(self, "Fehlende Pfade", "Bitte Excel-, Vorlagen- und Export-Pfad angeben!")
            return
        if not getattr(self, '_template_checkbox_by_rel', {}):
            QMessageBox.warning(
                self,
                "Keine Vorlagen",
                "Im Vorlagen-Ordner wurden keine verwendbaren Vorlagen (Unterordner anlagen/ oder allgemein/) gefunden.",
            )
            return
        sel = self._template_paths_for_worker()
        if sel == []:
            QMessageBox.warning(
                self,
                "Keine Auswahl",
                "Bitte mindestens eine Vorlage auswählen.",
            )
            return
        self.save_all_settings()
        self.log_text.clear()

        previous_export_root = None
        if self.reuse_lageplan_from_last_export and self.last_export_path and os.path.isdir(self.last_export_path):
            previous_export_root = self.last_export_path

        worker_params = {
            **self.paths,
            **self.settings,
            'categories': self.categories,
            'dry_run': False,
            'is_dark_mode': self.is_dark_mode,
            'selected_template_paths': sel,
            'rules_enabled': self.rules_enabled,
            'signage_rules': list(self.signage_rules) if self.signage_rules else [],
            'reuse_lageplan_from_last_export': self.reuse_lageplan_from_last_export,
            'previous_export_root': previous_export_root,
        }

        self.worker = Worker(**worker_params)

        self.worker.finished.connect(self.on_worker_finished)
        self.worker.log.connect(self.append_html_log)
        self.worker.current_file.connect(self.update_current_file_label)
        self.worker.progress.connect(self.update_progress_bar)

        self.worker.start()
        self.set_ui_running_state(True)

    def set_ui_running_state(self, is_running):
        """Aktiviert/Deaktiviert UI-Elemente, während der Worker läuft."""
        self.start_btn.setEnabled(not is_running)
        self.dry_run_btn.setEnabled(not is_running)
        if hasattr(self, 'preview_btn'):
            self.preview_btn.setEnabled(not is_running)
        self.close_btn.setText("Abbrechen" if is_running else "Schließen")
        if not is_running:
            self.current_file_label.setText("Bereit zum Starten...")
        self.open_export_folder_btn.setVisible(False)
        self.update_export_action_buttons(is_running)

    def handle_close_or_cancel(self):
        """Schließt die App oder bricht den Worker ab."""
        if hasattr(self, 'worker') and self.worker.isRunning():
            reply = QMessageBox.question(
                self,
                "Abbrechen?",
                "Möchten Sie den aktuellen Vorgang wirklich abbrechen?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.worker.requestInterruption()
        else:
            self.close()

    def on_worker_finished(self, success, export_path):
        """Wird aufgerufen, wenn der Worker-Thread seine Arbeit beendet hat."""
        template_row_preview = getattr(self, '_template_row_preview_active', False)
        self.set_ui_running_state(False)

        is_dry_run = self.worker.dry_run
        is_preview = getattr(self.worker, 'preview_run', False)
        was_cancelled = self.worker.isInterruptionRequested()

        if not is_dry_run and not is_preview:
            self.last_export_path = export_path if (success and export_path) else None

        if was_cancelled:
            self.current_file_label.setText("Vorgang abgebrochen.")
            QMessageBox.warning(self, "Abgebrochen", "Der Vorgang wurde vom Benutzer abgebrochen.")
        elif is_dry_run:
            if success:
                self.current_file_label.setText("Trockenlauf erfolgreich!")
                QMessageBox.information(self, "Trockenlauf", "Konfigurationsprüfung erfolgreich abgeschlossen.")
            else:
                self.current_file_label.setText("Trockenlauf fehlgeschlagen.")
                QMessageBox.warning(
                    self,
                    "Trockenlauf",
                    "Konfigurationsprüfung hat Fehler gefunden. Bitte Logs prüfen."
                )
        elif is_preview:
            if success and export_path:
                self.current_file_label.setText("Vorschau abgeschlossen.")
                want_pdf = bool(self.settings.get('export_as_pdf'))
                opened = False
                if want_pdf:
                    pdf_p = self._first_pdf_under_dir(export_path)
                    if pdf_p:
                        try:
                            os.startfile(pdf_p)
                            opened = True
                        except OSError:
                            QMessageBox.warning(
                                self,
                                "Vorschau",
                                f"PDF konnte nicht geöffnet werden:\n{pdf_p}",
                            )
                    else:
                        QMessageBox.warning(
                            self,
                            "Vorschau",
                            "Es wurde keine PDF-Datei gefunden. Prüfen Sie, ob Word/pywin32 verfügbar ist "
                            "und die Option „PDF“ aktiviert ist.",
                        )
                if not want_pdf or not opened:
                    try:
                        os.startfile(export_path)
                    except OSError:
                        QMessageBox.information(
                            self,
                            "Vorschau",
                            f"Dateien liegen unter:\n{export_path}",
                        )
                QMessageBox.information(
                    self,
                    "Vorschau",
                    "Die Beispieldokumente wurden erstellt. Vorschau ≠ rechtlicher Prüfstand; "
                    "andere Excel-Zeilen können abweichen.",
                )
            else:
                self.current_file_label.setText("Vorschau mit Fehlern oder ohne Ausgabe.")
                QMessageBox.warning(
                    self,
                    "Vorschau",
                    "Die Vorschau konnte keine Dokumente erzeugen oder ist fehlgeschlagen. Bitte Logs prüfen.",
                )
        elif success and self.last_export_path:
            self.current_file_label.setText("Verarbeitung abgeschlossen!")
            self.update_export_action_buttons(False)
            QMessageBox.information(self, "Fertig", "Alle Dokumente wurden erfolgreich erstellt.")
        else:
            self.current_file_label.setText("Verarbeitung mit Fehlern abgeschlossen.")
            QMessageBox.warning(
                self,
                "Fehler",
                "Die Verarbeitung wurde mit Fehlern abgeschlossen. Bitte Logs prüfen."
            )
        self.update_export_action_buttons(False)
        if template_row_preview:
            self._template_row_preview_hide()

    def update_export_action_buttons(self, is_running=False):
        """Aktualisiert die Buttons für Export-Aktionen basierend auf dem letzten Exportordner."""
        has_export_path = bool(self.last_export_path and os.path.isdir(self.last_export_path))
        is_pdf_running = hasattr(self, 'pdf_worker') and self.pdf_worker and self.pdf_worker.isRunning()
        export_actions_enabled = has_export_path and not is_running and not is_pdf_running
        select_actions_enabled = not is_running and not is_pdf_running
        self.open_export_folder_btn.setVisible(export_actions_enabled)
        if hasattr(self, 'pdf_convert_btn'):
            self.pdf_convert_btn.setEnabled(export_actions_enabled)
        if hasattr(self, 'pdf_open_folder_btn'):
            self.pdf_open_folder_btn.setEnabled(export_actions_enabled)
        if hasattr(self, 'pdf_select_folder_btn'):
            self.pdf_select_folder_btn.setEnabled(select_actions_enabled)
        if hasattr(self, 'pdf_last_export_label'):
            self.pdf_last_export_label.setText(
                f"Letzter Export: {self.last_export_path}" if has_export_path else "Letzter Export: —"
            )

    def refresh_pdf_tab(self):
        """Aktualisiert den PDF-Reiter basierend auf dem letzten Exportpfad."""
        self.update_export_action_buttons(False)
        if not (self.last_export_path and os.path.isdir(self.last_export_path)):
            self.pdf_summary_label.setText(
                "Kein letzter Export gefunden. Bitte Ordner auswählen oder zuerst Dokumente generieren."
            )
            if self.pdf_progress.value() != 0:
                self.pdf_progress.setValue(0)
            if self.pdf_current_file_label.text() == "Bereit.":
                return
            self.pdf_current_file_label.setText("Bereit.")

    def select_pdf_source_folder(self):
        """Erlaubt die manuelle Auswahl eines Export-Ordners für die PDF-Konvertierung."""
        if hasattr(self, 'worker') and self.worker.isRunning():
            QMessageBox.warning(self, "Bitte warten", "Während der Verarbeitung ist keine Ordnerauswahl möglich.")
            return
        if hasattr(self, 'pdf_worker') and self.pdf_worker and self.pdf_worker.isRunning():
            QMessageBox.warning(self, "Bitte warten", "Während der PDF-Konvertierung ist keine Ordnerauswahl möglich.")
            return

        start_dir = ""
        if self.last_export_path and os.path.isdir(self.last_export_path):
            start_dir = self.last_export_path
        elif self.paths.get('export_ordner') and os.path.isdir(self.paths.get('export_ordner')):
            start_dir = self.paths.get('export_ordner')
        else:
            start_dir = os.getcwd()

        selected_dir = QFileDialog.getExistingDirectory(self, "Export-Ordner für PDF auswählen", start_dir)
        if not selected_dir:
            return

        self.last_export_path = selected_dir
        self.update_export_action_buttons(False)
        self.pdf_summary_label.setText("Export-Ordner manuell gesetzt. Bereit zur PDF-Konvertierung.")
        self.pdf_current_file_label.setText("Bereit.")

    def convert_last_export_to_pdf(self):
        """Konvertiert alle DOCX-Dateien aus dem letzten Exportordner nach PDF."""
        if hasattr(self, 'worker') and self.worker.isRunning():
            QMessageBox.warning(self, "Bitte warten", "Während der Verarbeitung ist die PDF-Konvertierung nicht möglich.")
            return
        if hasattr(self, 'pdf_worker') and self.pdf_worker and self.pdf_worker.isRunning():
            QMessageBox.warning(self, "Bitte warten", "Die PDF-Konvertierung läuft bereits.")
            return

        export_path = self.last_export_path
        if not export_path:
            QMessageBox.warning(self, "Kein Export", "Es wurde noch kein Export durchgeführt.")
            return
        if not os.path.isdir(export_path):
            _log_handler(f"Export-Ordner nicht gefunden: {export_path}", "ERROR", self.append_html_log, self.is_dark_mode)
            QMessageBox.warning(self, "Ordner fehlt", f"Der letzte Export-Ordner existiert nicht mehr:\n{export_path}")
            self.update_export_action_buttons(False)
            return

        existing_pdf_count = 0
        for root, _, files in os.walk(export_path):
            existing_pdf_count += sum(1 for name in files if name.lower().endswith('.pdf'))
        if existing_pdf_count > 0:
            QMessageBox.warning(
                self,
                "Hinweis",
                f"Im Ordner sind bereits {existing_pdf_count} PDF-Datei(en) vorhanden. "
                "Die Umwandlung wird erneut ausgeführt."
            )

        self.current_file_label.setText("PDF-Konvertierung läuft...")
        self.pdf_current_file_label.setText("PDF-Konvertierung gestartet...")
        self.pdf_summary_label.setText("Konvertierung läuft...")
        self.pdf_progress.setValue(0)
        self.pdf_log_text.clear()
        self.update_export_action_buttons(True)
        self.pdf_worker = PdfWorker(export_path, self.is_dark_mode)
        self.pdf_worker.log.connect(self.append_html_log)
        self.pdf_worker.log.connect(self.append_pdf_html_log)
        self.pdf_worker.progress.connect(self.update_pdf_progress_bar)
        self.pdf_worker.current_file.connect(self.update_pdf_current_file_label)
        self.pdf_worker.finished.connect(self.on_pdf_worker_finished)

        for callback in (self.append_html_log, self.append_pdf_html_log):
            _log_handler("=" * 60 + "\n", "SEP", callback, self.is_dark_mode)
            _log_handler("Starte manuelle PDF-Konvertierung für letzten Export...", "INFO", callback, self.is_dark_mode)
            _log_handler(f"Export-Ordner: {export_path}", "INFO", callback, self.is_dark_mode)
            _log_handler("=" * 60 + "\n", "SEP", callback, self.is_dark_mode)
        self.pdf_worker.start()

    def append_pdf_html_log(self, html):
        """Fügt HTML-Logs in den PDF-Reiter ein."""
        if not hasattr(self, 'pdf_log_text'):
            return
        self.pdf_log_text.moveCursor(QTextCursor.End)
        self.pdf_log_text.insertHtml(html)
        self.pdf_log_text.ensureCursorVisible()

    def update_pdf_progress_bar(self, current, total):
        """Aktualisiert den Fortschritt der PDF-Konvertierung."""
        if not hasattr(self, 'pdf_progress'):
            return
        if total > 0:
            self.pdf_progress.setValue(int(current / total * 100))
        else:
            self.pdf_progress.setValue(0)

    def update_pdf_current_file_label(self, filename):
        """Zeigt die aktuell konvertierte Datei im PDF-Reiter."""
        if hasattr(self, 'pdf_current_file_label'):
            self.pdf_current_file_label.setText(f"Konvertiere: {filename}")

    def on_pdf_worker_finished(self, result):
        """Wird aufgerufen, wenn die PDF-Konvertierung abgeschlossen wurde."""
        docx_count = result.get("docx_count", 0)
        success_count = result.get("success_count", 0)
        failure_count = result.get("failure_count", 0)
        error_type = result.get("error")

        if error_type == "missing_docx2pdf":
            QMessageBox.warning(
                self,
                "Modul fehlt",
                "Das Modul 'docx2pdf' ist nicht installiert.\nBitte ausführen: pip install docx2pdf"
            )
            self.current_file_label.setText("PDF-Konvertierung nicht möglich (docx2pdf fehlt).")
            self.pdf_summary_label.setText("Fehler: docx2pdf ist nicht installiert.")
            self.pdf_current_file_label.setText("Keine Konvertierung durchgeführt.")
        elif error_type == "missing_pywin32":
            QMessageBox.warning(
                self,
                "Modul fehlt",
                "Das Modul 'pywin32' ist nicht verfügbar oder unvollständig eingerichtet.\n"
                "Bitte in der aktiven Umgebung ausführen:\n"
                "pip install pywin32\n"
                "pywin32_postinstall.py -install"
            )
            self.current_file_label.setText("PDF-Konvertierung nicht möglich (pywin32 fehlt).")
            self.pdf_summary_label.setText("Fehler: pywin32 ist nicht verfügbar.")
            self.pdf_current_file_label.setText("Keine Konvertierung durchgeführt.")
        elif error_type == "unknown_error":
            error_message = result.get("error_message", "Unbekannter Fehler")
            QMessageBox.critical(self, "PDF-Konvertierung fehlgeschlagen", f"Unerwarteter Fehler:\n{error_message}")
            self.current_file_label.setText("PDF-Konvertierung fehlgeschlagen.")
            self.pdf_summary_label.setText("PDF-Konvertierung fehlgeschlagen.")
            self.pdf_current_file_label.setText("Fehler aufgetreten.")
        elif docx_count == 0:
            QMessageBox.information(self, "Keine Word-Dateien", "Im letzten Export wurden keine .docx-Dateien gefunden.")
            self.current_file_label.setText("Keine Word-Dateien für PDF-Konvertierung gefunden.")
            self.pdf_summary_label.setText("Keine .docx-Dateien gefunden.")
            self.pdf_current_file_label.setText("Bereit.")
        elif failure_count == 0:
            QMessageBox.information(self, "PDF-Konvertierung fertig", f"{success_count} Datei(en) erfolgreich konvertiert.")
            self.current_file_label.setText("PDF-Konvertierung abgeschlossen.")
            self.pdf_summary_label.setText(f"Erfolg: {success_count} von {docx_count} Datei(en) konvertiert.")
            self.pdf_current_file_label.setText("Konvertierung abgeschlossen.")
            self.pdf_progress.setValue(100)
        else:
            QMessageBox.warning(
                self,
                "PDF-Konvertierung teilweise fehlgeschlagen",
                f"{success_count} von {docx_count} Datei(en) wurden konvertiert.\n"
                f"{failure_count} Datei(en) sind fehlgeschlagen. Bitte Logs prüfen."
            )
            self.current_file_label.setText("PDF-Konvertierung mit Fehlern abgeschlossen.")
            self.pdf_summary_label.setText(
                f"Teilweise erfolgreich: {success_count}/{docx_count}, Fehler: {failure_count}."
            )
            self.pdf_current_file_label.setText("Konvertierung abgeschlossen (mit Fehlern).")

        self.update_export_action_buttons(False)

    def update_current_file_label(self, filename):
        """Aktualisiert das Label, das den Namen der aktuell verarbeiteten Datei anzeigt."""
        self.current_file_label.setText(f"Verarbeite: {filename}...")

    def update_progress_bar(self, current, total):
        """Aktualisiert die Fortschrittsanzeige."""
        if total > 0:
            self.progress.setValue(int(current / total * 100))
        else:
            self.progress.setValue(0)

    def _show_template_row_preview_ui(self, vorlage_name):
        """Fortschritt im Tab „Vorlagen“ für die Einzel-Vorschau (sichtbar ohne Wechsel zur Hauptsteuerung)."""
        self._template_row_preview_active = True
        if hasattr(self, '_template_preview_status'):
            self._template_preview_status.setVisible(True)
            self._template_preview_status.setText(
                f"Vorschau läuft … ({vorlage_name}) — bitte warten, Word/Excel kann kurz blockieren."
            )
        if hasattr(self, '_template_preview_progress'):
            self._template_preview_progress.setVisible(True)
            self._template_preview_progress.setRange(0, 0)
            self._template_preview_progress.setFormat("Vorschau läuft …")

    def _template_row_preview_hide(self):
        """Blendet die Vorlagen-Vorschau-Fortschrittszeile aus."""
        self._template_row_preview_active = False
        if hasattr(self, '_template_preview_progress'):
            self._template_preview_progress.setRange(0, 100)
            self._template_preview_progress.setValue(0)
            self._template_preview_progress.setFormat("%p%")
            self._template_preview_progress.setVisible(False)
        if hasattr(self, '_template_preview_status'):
            self._template_preview_status.setVisible(False)
            self._template_preview_status.setText("")

    def _on_template_row_preview_progress(self, current, total):
        if not getattr(self, '_template_row_preview_active', False):
            return
        if not hasattr(self, '_template_preview_progress'):
            return
        if total and total > 0:
            self._template_preview_progress.setRange(0, 100)
            self._template_preview_progress.setFormat("%p%")
            self._template_preview_progress.setValue(min(100, int(100 * current / total)))
        else:
            self._template_preview_progress.setRange(0, 0)
            self._template_preview_progress.setFormat("Vorschau läuft …")

    def _on_template_row_preview_file(self, filename):
        if not getattr(self, '_template_row_preview_active', False):
            return
        if hasattr(self, '_template_preview_status'):
            self._template_preview_status.setText(f"Aktuell: {filename}")

    def append_html_log(self, html):
        """Ein Slot, der HTML-formatierten Text sicher an das Log-Fenster anhängt und für Filter speichert."""
        import re
        # Robust: Suche nach [ LEVEL ] mit beliebigen Leerzeichen, unabhängig von HTML-Tags
        m = re.search(r'\[\s*([A-Z]+)\s*\]', html)
        level = m.group(1) if m else "INFO"
        self.log_lines.append((level, html))
        self.apply_log_filter()

    def apply_log_filter(self):
        """Zeigt nur die Log-Zeilen an, die dem gewählten Filter entsprechen."""
        filter_level = self.log_filter_combo.currentText()
        self.log_text.clear()
        for level, html in self.log_lines:
            if filter_level == "ALLE" or level == filter_level:
                self.log_text.moveCursor(QTextCursor.End)
                self.log_text.insertHtml(html)
        self.log_text.ensureCursorVisible()

    def show_log_context_menu(self, position):
        """Zeigt ein Kontextmenü für Log-Einträge mit Optionen zum Öffnen von Ordnern/Dateien/Textmarken."""
        cursor = self.log_text.cursorForPosition(position)
        cursor.select(QTextCursor.WordUnderCursor)
        selected_text = cursor.selectedText()

        # Suche nach Pfaden in der aktuellen Zeile
        line_cursor = self.log_text.cursorForPosition(position)
        line_cursor.movePosition(QTextCursor.StartOfLine)
        line_cursor.movePosition(QTextCursor.EndOfLine, QTextCursor.KeepAnchor)
        line_text = line_cursor.selectedText()

        menu = QMenu(self)

        # Extrahiere mögliche Pfade aus der Zeile
        paths = self.extract_paths_from_log_line(line_text)

        if paths:
            # Füge Menüpunkte für jeden gefundenen Pfad hinzu
            for path_type, path, display_name in paths:
                if path_type == "file":
                    action = menu.addAction(f"📄 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_file(p))
                elif path_type == "folder":
                    action = menu.addAction(f"📁 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_folder(p))
                elif path_type == "template":
                    action = menu.addAction(f"📝 {display_name} öffnen")
                    action.triggered.connect(lambda checked, p=path: self.open_template_file(p))
            menu.addSeparator()

        # Prüfe auf fehlende Platzhalter/Textmarken
        fehlende_marken = self.extract_missing_placeholders_from_log_line(line_text)
        if fehlende_marken:
            action = menu.addAction("➕ Textmarke(n) im Excel hinzufügen")
            action.triggered.connect(lambda checked, marken=fehlende_marken: self.add_placeholders_to_excel(marken))
            menu.addSeparator()

        # Prüfe auf fehlende Size-Textmarken
        fehlende_size_marken = self.extract_missing_size_placeholders_from_log_line(line_text)
        if fehlende_size_marken:
            action = menu.addAction("➕ Size-Textmarke(n) im Excel hinzufügen")
            action.triggered.connect(lambda checked, marken=fehlende_size_marken: self.add_placeholders_to_excel(marken))
            menu.addSeparator()

        # Standard-Menüpunkte
        copy_action = menu.addAction("📋 Ausgewählten Text kopieren")
        copy_action.triggered.connect(self.copy_selected_text)

        clear_action = menu.addAction("🗑️ Log löschen")
        clear_action.triggered.connect(self.clear_log)

        menu.exec_(self.log_text.mapToGlobal(position))

    def extract_paths_from_log_line(self, line_text):
        """Extrahiert mögliche Datei- und Ordnerpfade aus einer Log-Zeile."""
        paths = []
        import re
        # Suche nach Dateinamen in verschiedenen Log-Formaten
        # 1. Export-Log: Dokument: 'Dateiname.docx'
        datei_matches = re.findall(r"Dokument: '([^']+)'", line_text)
        # 2. Fehler/Warnung: Vorlage 'Dateiname.docx'
        vorlage_matches = re.findall(r"Vorlage '([^']+)'", line_text)
        # 3. Export-Log: in Ordner: 'Ordnername'
        ordner_matches = re.findall(r"in Ordner: '([^']+)'", line_text)

        # Exportierte Dateien suchen (wie bisher)
        if hasattr(self, 'last_export_path') and self.last_export_path:
            base_path = self.last_export_path
            # Füge Dateien aus Export-Log hinzu
            for datei in datei_matches:
                for category_name in self.categories.values():
                    potential_path = os.path.join(base_path, category_name, datei)
                    if os.path.exists(potential_path):
                        paths.append(("file", potential_path, datei))
                        break
            # Füge Dateien aus Fehler/Warnung hinzu
            for datei in vorlage_matches:
                for category_name in self.categories.values():
                    potential_path = os.path.join(base_path, category_name, datei)
                    if os.path.exists(potential_path):
                        paths.append(("file", potential_path, datei))
                        break
            # Füge Ordner hinzu
            for ordner in ordner_matches:
                potential_path = os.path.join(base_path, ordner)
                if os.path.exists(potential_path):
                    paths.append(("folder", potential_path, ordner))

        # Vorlagen-Dateien suchen
        # Suche in self.paths['vorlagen_ordner'] und allen Unterordnern nach passenden Dateinamen
        if hasattr(self, 'paths') and 'vorlagen_ordner' in self.paths:
            vorlagen_root = self.paths['vorlagen_ordner']
            alle_vorlagen = set(datei_matches + vorlage_matches)
            for suchname in alle_vorlagen:
                for root, _, files in os.walk(vorlagen_root):
                    for file in files:
                        if file == suchname:
                            full_path = os.path.join(root, file)
                            paths.append(("template", full_path, suchname))
        return paths

    def extract_missing_placeholders_from_log_line(self, line_text):
        """Extrahiert fehlende Platzhalter/Textmarken aus einer Log-Zeile."""
        import re
        # Sucht nach: Fehlende Platzhalter in Excel: {'foo', 'bar'}
        match = re.search(r"Fehlende Platzhalter in Excel: (\{.*?\})", line_text)
        if match:
            try:
                # Sichere Auswertung des Sets
                marken = eval(match.group(1), {"__builtins__": None}, {})
                if isinstance(marken, set):
                    return sorted(marken)
            except Exception:
                pass
        return []

    def extract_missing_size_placeholders_from_log_line(self, line_text):
        """Extrahiert fehlende Size-Textmarken aus einer Log-Zeile."""
        import re
        # NEU: Erkenne auch die neue Info-Log-Zeile
        # Beispiel: Hinweis: Für Platzhalter 'ba_logo_img' wird die Standardgröße verwendet, da kein 'ba_logo_img_size' oder 'ba_logo_img_Size' in Excel/Fallback gefunden wurde.
        match_alt = re.search(r"Keine Größenangabe\s*\(([^)]+)\).*?für Platzhalter", line_text)
        match_neu = re.search(
            r"Hinweis: Für Platzhalter '([^']+)' wird die Standardgröße verwendet, da kein '([^']+)' oder '([^']+)' in Excel/Fallback gefunden wurde",
            line_text
        )
        if match_alt:
            inhalt = match_alt.group(1)
            self.append_html_log(f'<span style="color:#888">[DEBUG] Klammer-Inhalt: {inhalt}</span><br>')
            marken = re.findall(r"'([^']+)'", inhalt)
            if marken:
                self.append_html_log(
                    f'<span style="color:#888">[DEBUG] Erkannte Size-Textmarken: {", ".join(marken)}</span><br>'
                )
            else:
                self.append_html_log(f'<span style="color:#888">[DEBUG] Keine Size-Textmarken erkannt.</span><br>')
            return sorted(marken)
        elif match_neu:
            # Extrahiere die beiden _size-Varianten
            size1 = match_neu.group(2)
            size2 = match_neu.group(3)
            marken = [size1, size2]
            self.append_html_log(
                f'<span style="color:#888">[DEBUG] Erkannte Size-Textmarken (neu): {", ".join(marken)}</span><br>'
            )
            return sorted(marken)
        self.append_html_log(f'<span style="color:#888">[DEBUG] Keine Size-Textmarken erkannt.</span><br>')
        return []

    def add_placeholders_to_excel(self, placeholders):
        """Fügt die angegebenen Platzhalter als neue Zeilen im zweiten Blatt der aktuellen Excel-Datei hinzu (Spalte 2=Textmarke, Spalte 3=leer)."""
        from openpyxl import load_workbook
        import os
        excel_path = self.paths.get('excel_path')
        if not excel_path or not os.path.exists(excel_path):
            QMessageBox.warning(self, "Fehler", "Excel-Datei nicht gefunden!")
            return
        wb = None
        try:
            wb = load_workbook(excel_path)
            if len(wb.sheetnames) < 2:
                QMessageBox.warning(self, "Fehler", "Die Excel-Datei hat kein zweites Blatt!")
                return
            ws = wb[wb.sheetnames[1]]
            # Nur eine Size-Variante pro Basis-Textmarke einfügen
            filtered = {}
            for marke in placeholders:
                if marke.lower().endswith('_size'):
                    base = marke[:-5].lower()
                    # Bevorzuge die Variante mit kleinem 's'
                    if base not in filtered or marke.endswith('_size'):
                        filtered[base] = marke
                else:
                    filtered[marke] = marke
            for marke in filtered.values():
                ws.append([None, marke, ""])
            wb.save(excel_path)
            QMessageBox.information(
                self,
                "Erfolg",
                f"{len(filtered)} Textmarke(n) wurden im zweiten Blatt hinzugefügt."
            )
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte Textmarken nicht hinzufügen: {e}")
        finally:
            if wb is not None:
                wb.close()

    def open_file(self, file_path):
        """Öffnet eine Datei mit der Standard-Anwendung."""
        try:
            os.startfile(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte die Datei nicht öffnen: {e}")

    def open_folder(self, folder_path):
        """Öffnet einen Ordner im Explorer."""
        try:
            os.startfile(folder_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte den Ordner nicht öffnen: {e}")

    def open_template_file(self, file_path):
        """Öffnet eine Word-Vorlage mit der Standard-Anwendung."""
        try:
            os.startfile(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Fehler", f"Konnte die Vorlage nicht öffnen: {e}")

    def copy_selected_text(self):
        """Kopiert den ausgewählten Text in die Zwischenablage."""
        cursor = self.log_text.textCursor()
        if cursor.hasSelection():
            text = cursor.selectedText()
            clipboard = QApplication.clipboard()
            clipboard.setText(text)

    def clear_log(self):
        """Löscht den gesamten Log-Inhalt."""
        reply = QMessageBox.question(
            self,
            "Log löschen",
            "Möchten Sie wirklich den gesamten Log löschen?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.log_text.clear()

    def show_excel_data(self):
        """
        Versucht, die Daten aus der angegebenen Excel-Datei zu laden und
        in der Tabelle (QTableView) in der GUI anzuzeigen.
        Bei gesperrter Datei (z. B. in Excel geöffnet) wird unter Windows
        ein COM-Fallback verwendet.
        """
        path = self.paths.get('excel_path')
        header_row = self.settings.get('header_row', 3)
        if path and os.path.exists(path):
            try:
                df = pd.read_excel(path, header=header_row - 1)
                self.table.setModel(PandasModel(df))
            except (PermissionError, OSError) as e:
                if getattr(e, 'errno', None) == 13 or isinstance(e, PermissionError):
                    try:
                        from core.excel_com import read_excel_sheet1_via_com
                        df = read_excel_sheet1_via_com(path, header_row)
                        if df is not None:
                            self.table.setModel(PandasModel(df))
                            return
                    except Exception:
                        pass
                self.table.setModel(None)
                QMessageBox.critical(self, "Fehler beim Lesen der Excel-Datei", str(e))
            except Exception as e:
                self.table.setModel(None)  # Clear table on error
                QMessageBox.critical(self, "Fehler beim Lesen der Excel-Datei", str(e))

    def open_export_folder(self):
        """Öffnet den zuletzt verwendeten Export-Ordner im Datei-Explorer."""
        if self.last_export_path and os.path.isdir(self.last_export_path):
            try:
                os.startfile(self.last_export_path)
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Konnte den Ordner nicht öffnen: {e}")

    def open_excel_file(self):
        excel_path = self.paths.get('excel_path')
        if excel_path and os.path.exists(excel_path):
            try:
                os.startfile(excel_path)
            except Exception as e:
                QMessageBox.warning(self, "Fehler", f"Excel konnte nicht geöffnet werden: {e}")
        else:
            QMessageBox.warning(self, "Fehler", "Excel-Datei nicht gefunden!")

    def closeEvent(self, event):
        """Stellt sicher, dass der Worker-Thread sauber beendet wird, wenn das Fenster geschlossen wird."""
        if hasattr(self, 'worker') and self.worker.isRunning():
            self.worker.requestInterruption()
            self.worker.wait()
        if hasattr(self, 'pdf_worker') and self.pdf_worker and self.pdf_worker.isRunning():
            self.pdf_worker.wait()
        event.accept()
